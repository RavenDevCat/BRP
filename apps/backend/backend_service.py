from __future__ import annotations

import base64
import importlib
import importlib.util
import io
import json
import math
import os
import sys
import tempfile
import threading
import time
import traceback
from copy import deepcopy
from dataclasses import asdict, fields
from datetime import date, datetime, time as datetime_time, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import Workbook

try:
    from .job_queue import (
        JobQueueManager,
        terminate_worker_process,
        worker_creation_flags,
    )
    from .json_cache_store import load_json_object, save_json_object
    from . import osrm_manager
    from .ai_audit import generate_ai_audit_report as _generate_ai_audit_report
    from .quota_store_sqlite import SqliteQuotaStore
    from .runtime_store_sqlite import SqliteRuntimeStore
    from .planner_core import (
        FINAL_ROUTE_TRAFFIC_CACHE_PATH,
        PlannerConfig,
        _build_assessment_metric_matrices,
        _amap_route_stats,
        _route_amap_points,
        assess_current_plan,
        build_baseline_template_workbook_bytes,
        build_current_plan_map_scenario,
        build_excel_template_bytes,
        infer_traffic_location,
        load_legacy_planner,
        normalize_traffic_coefficient_mode,
        rerender_html_from_structured_results,
        resolve_traffic_profile,
        run_backend_planner_with_prepared_data,
        summarize_live_traffic_samples,
    )
except ImportError:  # pragma: no cover - supports running from apps/backend directly.
    from job_queue import (
        JobQueueManager,
        terminate_worker_process,
        worker_creation_flags,
    )
    from json_cache_store import load_json_object, save_json_object
    import osrm_manager
    from ai_audit import generate_ai_audit_report as _generate_ai_audit_report
    from quota_store_sqlite import SqliteQuotaStore
    from runtime_store_sqlite import SqliteRuntimeStore
    from planner_core import (
        FINAL_ROUTE_TRAFFIC_CACHE_PATH,
        PlannerConfig,
        _build_assessment_metric_matrices,
        _amap_route_stats,
        _route_amap_points,
        assess_current_plan,
        build_baseline_template_workbook_bytes,
        build_current_plan_map_scenario,
        build_excel_template_bytes,
        infer_traffic_location,
        load_legacy_planner,
        normalize_traffic_coefficient_mode,
        rerender_html_from_structured_results,
        resolve_traffic_profile,
        run_backend_planner_with_prepared_data,
        summarize_live_traffic_samples,
    )


BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent.parent
CLIENT_DIR = BASE_DIR.parent / "client"
DEFAULT_JOBS_DIR = REPO_ROOT / "state" / "jobs"
RAW_JOBS_DIR = os.environ.get("BRP_BACKEND_JOBS_DIR", "").strip()
JOBS_DIR = Path(RAW_JOBS_DIR or str(DEFAULT_JOBS_DIR)).expanduser()
DEFAULT_SIDE_TOOLS_DIR = REPO_ROOT / "state" / "side_tools"
RAW_SIDE_TOOLS_DIR = os.environ.get("BRP_SIDE_TOOLS_DIR", "").strip()
SIDE_TOOLS_DIR = Path(RAW_SIDE_TOOLS_DIR or str(DEFAULT_SIDE_TOOLS_DIR)).expanduser()
RAW_RUNTIME_DB_PATH = os.environ.get("BRP_RUNTIME_DB_PATH", "").strip()
RUNTIME_DB_PATH = Path(RAW_RUNTIME_DB_PATH or str(JOBS_DIR.parent / "brp_runtime.sqlite")).expanduser()
JOB_RUNNER_PATH = BASE_DIR / "backend_job_runner.py"
SERVICE_TOKEN = os.environ.get("BRP_BACKEND_SERVICE_TOKEN", "").strip()
DEV_USER_EMAIL = os.environ.get("BRP_DEV_USER_EMAIL", "local@brp.dev").strip().lower()
AUTH_PROVIDER = (
    (
        os.environ.get("BRP_AUTH_PROVIDER")
        or os.environ.get("BRP_AUTH_MODE")
        or "cloudflare_header"
    )
    .strip()
    .lower()
)
AUTH_LOGIN_URL = os.environ.get("BRP_AUTH_LOGIN_URL", "").strip()
AUTH_LOGOUT_URL = os.environ.get("BRP_AUTH_LOGOUT_URL", "").strip()
AUTH_DISPLAY_NAME = os.environ.get("BRP_AUTH_DISPLAY_NAME", "").strip()
try:
    MAX_CONCURRENT_JOBS = max(
        0, int(os.environ.get("BRP_MAX_CONCURRENT_JOBS", "0") or "0")
    )
except ValueError:
    MAX_CONCURRENT_JOBS = 0
RAW_JOB_CONCURRENCY_DIR = os.environ.get("BRP_JOB_CONCURRENCY_DIR", "").strip()
JOB_CONCURRENCY_DIR = Path(
    RAW_JOB_CONCURRENCY_DIR or str(REPO_ROOT / "state" / "job_concurrency")
).expanduser()
try:
    JOB_QUEUE_POLL_SECONDS = max(
        1.0, float(os.environ.get("BRP_JOB_QUEUE_POLL_SECONDS", "5") or "5")
    )
except ValueError:
    JOB_QUEUE_POLL_SECONDS = 5.0
try:
    JOB_SLOT_ATTACH_STALE_SECONDS = max(
        30.0,
        float(os.environ.get("BRP_JOB_SLOT_ATTACH_STALE_SECONDS", "300") or "300"),
    )
except ValueError:
    JOB_SLOT_ATTACH_STALE_SECONDS = 300.0
MAX_WORKBOOK_UPLOAD_BYTES = int(
    os.environ.get("BRP_MAX_WORKBOOK_UPLOAD_BYTES", str(20 * 1024 * 1024))
)
ADMIN_EMAILS = {
    item.strip().lower()
    for item in os.environ.get("BRP_ADMIN_EMAILS", "").split(",")
    if item.strip()
}
MAP_ARTIFACT_KEYS = {
    "current_plan": "current_plan",
    "original": "original",
    "free_optimization": "original",
    "subway": "subway",
    "nearby": "nearby",
    "time_constrained": "time_constrained",
    "time_constrained_optimization": "time_constrained",
    "exception_preserving": "exception_preserving",
    "exception_preserving_optimization": "exception_preserving",
    "ep15min": "ep15min",
    "ep15min_optimization": "ep15min",
    "further_most": "further_most",
    "further_most_nearby": "further_most_nearby",
}
MAP_SCENARIO_LABELS = {
    "current_plan": "Current Plan",
    "original": "Free Optimization Baseline",
    "subway": "Subway Aggregated",
    "nearby": "Nearby Aggregated",
    "time_constrained": "15-Minute Constrained",
    "exception_preserving": "Exception Preserving",
    "ep15min": "EP 15-Minute",
    "further_most": "Further Most",
    "further_most_nearby": "Further Most + Nearby Aggregate",
}
MAP_ARTIFACT_TOP_LEVEL_KEYS = {
    "current_plan": "current_plan_html",
    "original": "original_html",
    "subway": "subway_html",
    "nearby": "nearby_html",
    "time_constrained": "time_constrained_html",
    "exception_preserving": "exception_preserving_html",
    "ep15min": "ep15min_html",
    "further_most": "further_most_html",
    "further_most_nearby": "further_most_nearby_html",
}
WORKBOOK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES = float(
    os.environ.get("BRP_TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES", "15") or 15
)
MAX_DISTANCE_CHECKER_JOBS = 80
CLIENT_CACHE_DIR = Path(
    os.environ.get("BRP_CLIENT_CACHE_DIR", str(CLIENT_DIR / "cache"))
).expanduser()
MAP_TILE_CACHE_DIR = Path(
    os.environ.get("BRP_MAP_TILE_CACHE_DIR", str(BASE_DIR / "cache" / "map_tiles"))
).expanduser()
DISTANCE_CHECKER_JOBS_PATH = CLIENT_CACHE_DIR / "distance_checker_jobs.json"
GOOGLE_GEOCODE_USAGE_PATH = CLIENT_CACHE_DIR / "google_geocode_usage.json"
GOOGLE_GEOCODE_USAGE_PROVIDER = "google_geocode"
GOOGLE_GEOCODE_USAGE_COUNTER = "geocode"
GOOGLE_GEOCODE_USAGE_PROVIDER_LABEL = "Google Geocoding"
GOOGLE_GEOCODE_USAGE_SKU_ESTIMATE = "geocoding"
MAP_TILE_UPSTREAM_TEMPLATE = os.environ.get(
    "BRP_MAP_TILE_UPSTREAM_TEMPLATE",
    "https://tile.openstreetmap.de/{z}/{x}/{y}.png",
).strip()
MAP_TILE_FALLBACK_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
)
GOOGLE_GEOCODE_MONTHLY_LIMIT = 10_000


def _env_flag(name: str, default: bool = False) -> bool:
    raw_value = os.environ.get(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {"1", "true", "yes", "on"}


def _default_scheduled_jobs_enabled() -> bool:
    root_text = str(BASE_DIR).replace("\\", "/").lower()
    if "users/bus.eim/brp" in root_text:
        return False
    return "/staging/" in root_text or "/prod/" in root_text


GOOGLE_GEOCODE_USAGE_VISIBLE = _env_flag("BRP_SHOW_GOOGLE_GEOCODE_USAGE", False)
ENABLE_LANGUAGE_SWITCH = not _env_flag("BRP_DISABLE_LANGUAGE_SWITCH", False)
SCHEDULED_JOBS_ENABLED = _env_flag(
    "BRP_ENABLE_SCHEDULED_JOBS",
    _default_scheduled_jobs_enabled(),
)
DEFAULT_TRAFFIC_COEFFICIENT_MODE = normalize_traffic_coefficient_mode(
    os.environ.get("BRP_DEFAULT_TRAFFIC_COEFFICIENT_MODE", "legacy")
)
AMAP_DISPLAY_GEOMETRY_ENABLED = _env_flag("BRP_AMAP_DISPLAY_GEOMETRY_ENABLED", True)
AMAP_DISPLAY_GEOMETRY_CACHE_PATH = Path(
    os.environ.get(
        "BRP_AMAP_DISPLAY_GEOMETRY_CACHE_PATH",
        str(BASE_DIR / "cache" / "amap_display_geometry.json"),
    )
).expanduser()
try:
    AMAP_DISPLAY_GEOMETRY_MAX_WAYPOINTS = max(
        0, int(os.environ.get("BRP_AMAP_DISPLAY_GEOMETRY_MAX_WAYPOINTS", "16") or "16")
    )
except ValueError:
    AMAP_DISPLAY_GEOMETRY_MAX_WAYPOINTS = 16
try:
    AMAP_DISPLAY_GEOMETRY_REQUEST_INTERVAL_S = max(
        0.0,
        float(
            os.environ.get("BRP_AMAP_DISPLAY_GEOMETRY_REQUEST_INTERVAL_S", "0.36")
            or "0.36"
        ),
    )
except ValueError:
    AMAP_DISPLAY_GEOMETRY_REQUEST_INTERVAL_S = 0.36
AMAP_DISPLAY_GEOMETRY_VERSION = "amap-cn-display-v1"
_AMAP_DISPLAY_CACHE_LOCK = threading.Lock()
_AMAP_DISPLAY_REQUEST_LOCK = threading.Lock()
_AMAP_DISPLAY_LAST_REQUEST_AT = 0.0


def _amap_display_api_key() -> str:
    for env_key in ("AMAP_API_KEY", "BRP_AMAP_API_KEY"):
        value = os.environ.get(env_key, "").strip()
        if value:
            return value
    try:
        try:
            from .BusingProblem import AMAP_KEY as existing_amap_key
        except ImportError:
            from BusingProblem import AMAP_KEY as existing_amap_key
    except Exception:
        return ""
    return str(existing_amap_key or "").strip()


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


BRP_LOCAL_TZ = timezone(timedelta(hours=8))


def _parse_iso_datetime(value: object) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _next_scheduled_job_trigger(config_payload: dict[str, Any], scheduled_date: Any = None) -> tuple[str, str]:
    direction = str(config_payload.get("service_direction") or "").strip().lower()
    window_start, _ = _parse_clock_payload(
        config_payload.get("time_window_start"),
        "time_window_start",
        "15:40" if direction == "from school" else "06:30",
    )
    hour, minute = (int(part) for part in window_start.split(":", 1))
    label = f"{window_start} {'PM' if direction == 'from school' else 'AM'} window"
    now = datetime.now(BRP_LOCAL_TZ).replace(microsecond=0)
    if scheduled_date:
        try:
            target_day = date.fromisoformat(str(scheduled_date).strip())
        except ValueError as exc:
            raise ValueError("scheduled_date must use YYYY-MM-DD format.") from exc
        target = datetime.combine(target_day, datetime_time(hour, minute), BRP_LOCAL_TZ)
    else:
        target = now.replace(hour=hour, minute=minute, second=0)
    if target <= now:
        if scheduled_date:
            raise ValueError("Scheduled date/time has already passed.")
        target += timedelta(days=1)
    return target.isoformat(), label


_RUNTIME_SQLITE_STORE: SqliteRuntimeStore | None = None
_RUNTIME_SQLITE_STORE_LOCK = threading.Lock()


def _runtime_sqlite_store() -> SqliteRuntimeStore:
    global _RUNTIME_SQLITE_STORE
    with _RUNTIME_SQLITE_STORE_LOCK:
        if _RUNTIME_SQLITE_STORE is None:
            _RUNTIME_SQLITE_STORE = SqliteRuntimeStore(RUNTIME_DB_PATH)
            _RUNTIME_SQLITE_STORE.initialize()
        return _RUNTIME_SQLITE_STORE


def _save_runtime_job(record: dict[str, Any]) -> None:
    store = _runtime_sqlite_store()
    try:
        store.upsert_job(record)
    except Exception:
        traceback.print_exc()


def _delete_runtime_job(job_id: str) -> bool:
    store = _runtime_sqlite_store()
    try:
        return bool(store.delete_job(job_id))
    except Exception:
        traceback.print_exc()
        return False


def _save_runtime_side_tool(tool_key: str, record: dict[str, Any]) -> None:
    store = _runtime_sqlite_store()
    try:
        store.upsert_side_tool_run(tool_key, record)
    except Exception:
        traceback.print_exc()


def _delete_runtime_side_tool(tool_key: str, run_id: str) -> bool:
    store = _runtime_sqlite_store()
    try:
        return bool(store.delete_side_tool_run(tool_key, run_id))
    except Exception:
        traceback.print_exc()
        return False


def _build_planner_config(config_payload: dict[str, Any]) -> PlannerConfig:
    allowed_field_names = {field.name for field in fields(PlannerConfig)}
    filtered_payload = {
        key: value
        for key, value in dict(config_payload or {}).items()
        if key in allowed_field_names
    }
    return PlannerConfig(**filtered_payload)


def _parse_clock_payload(value: Any, field_name: str, default_value: str) -> tuple[str, int]:
    raw = str(value or default_value).strip()
    parts = raw.split(":")
    if len(parts) < 2:
        raise ValueError(f"{field_name} must use HH:MM format.")
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
    except ValueError as exc:
        raise ValueError(f"{field_name} must use HH:MM format.") from exc
    if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
        raise ValueError(f"{field_name} must use HH:MM format.")
    return f"{hours:02d}:{minutes:02d}", hours * 60 + minutes


def _parse_route_stop_limit_payload(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        limit = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("route_stop_limit must be a whole number.") from exc
    if limit < 1 or limit > 200:
        raise ValueError("route_stop_limit must be between 1 and 200.")
    return limit


def _parse_minimum_vehicle_reduction_payload(value: Any) -> int:
    if value is None or value == "":
        return 2
    try:
        reduction = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("minimum_vehicle_reduction must be a whole number.") from exc
    if reduction < 0 or reduction > 200:
        raise ValueError("minimum_vehicle_reduction must be between 0 and 200.")
    return reduction


def _parse_time_impact_limit_payload(value: Any) -> int:
    if value is None or value == "":
        return int(TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)
    try:
        minutes = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("time_impact_limit_minutes must be a whole number.") from exc
    if minutes < 0 or minutes > 240:
        raise ValueError("time_impact_limit_minutes must be between 0 and 240.")
    return minutes


def _planner_config_payload(config_payload: dict[str, Any]) -> dict[str, Any]:
    normalized_input = dict(config_payload or {})
    window_start, start_minutes = _parse_clock_payload(
        normalized_input.get("time_window_start"),
        "time_window_start",
        "06:30",
    )
    window_end, end_minutes = _parse_clock_payload(
        normalized_input.get("time_window_end"),
        "time_window_end",
        "08:00",
    )
    if end_minutes <= start_minutes:
        raise ValueError("time_window_start must be before time_window_end.")
    normalized_input["time_window_start"] = window_start
    normalized_input["time_window_end"] = window_end
    normalized_input["route_stop_limit"] = _parse_route_stop_limit_payload(
        normalized_input.get("route_stop_limit")
    )
    normalized_input["minimum_vehicle_reduction"] = _parse_minimum_vehicle_reduction_payload(
        normalized_input.get("minimum_vehicle_reduction")
    )
    normalized_input["time_impact_limit_minutes"] = _parse_time_impact_limit_payload(
        normalized_input.get("time_impact_limit_minutes")
    )
    payload = asdict(_build_planner_config(normalized_input))
    payload["traffic_coefficient_mode"] = normalize_traffic_coefficient_mode(
        payload.get("traffic_coefficient_mode")
    )
    return payload


def _client_core_module() -> Any:
    client_dir = str(CLIENT_DIR)
    if client_dir not in sys.path:
        sys.path.insert(0, client_dir)
    return importlib.import_module("client_core")


def _distance_tool_module() -> Any:
    client_dir = str(CLIENT_DIR)
    if client_dir not in sys.path:
        sys.path.insert(0, client_dir)
    return importlib.import_module("distance_tool")


def _client_module(module_name: str) -> Any:
    client_dir = str(CLIENT_DIR)
    if client_dir not in sys.path:
        sys.path.insert(0, client_dir)
    return importlib.import_module(module_name)


def _build_client_planner_config(
    client_core: Any, config_payload: dict[str, Any]
) -> Any:
    allowed_field_names = {field.name for field in fields(client_core.PlannerConfig)}
    filtered_payload = {
        key: value
        for key, value in dict(config_payload or {}).items()
        if key in allowed_field_names
    }
    return client_core.PlannerConfig(**filtered_payload)


def _build_job_display_name(source_label: str, custom_name: str = "") -> str:
    default_name = Path(str(source_label or "")).stem.strip() or "Untitled job"
    normalized_custom_name = " ".join(str(custom_name or "").strip().split())
    if not normalized_custom_name:
        return default_name
    return normalized_custom_name


def _decode_workbook_bytes(payload: dict[str, Any]) -> tuple[str, bytes]:
    source_label = str(
        payload.get("file_name") or payload.get("source_label") or "workbook.xlsx"
    ).strip()
    suffix = Path(source_label).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Workbook upload must be an .xlsx or .xlsm file.")
    raw_base64 = str(payload.get("file_base64") or "").strip()
    if not raw_base64:
        raise ValueError("Missing workbook file_base64 payload.")
    if "," in raw_base64 and raw_base64.split(",", 1)[0].startswith("data:"):
        raw_base64 = raw_base64.split(",", 1)[1]
    workbook_bytes = base64.b64decode(raw_base64, validate=False)
    if not workbook_bytes:
        raise ValueError("Uploaded workbook is empty.")
    if len(workbook_bytes) > MAX_WORKBOOK_UPLOAD_BYTES:
        raise ValueError(
            f"Workbook upload exceeds {MAX_WORKBOOK_UPLOAD_BYTES // (1024 * 1024)} MB."
        )
    return source_label, workbook_bytes


def _with_temp_workbook(payload: dict[str, Any], callback: Any) -> Any:
    source_label, workbook_bytes = _decode_workbook_bytes(payload)
    suffix = Path(source_label).suffix.lower()
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file.write(workbook_bytes)
            temp_path = temp_file.name
        return callback(source_label, temp_path)
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except OSError:
                pass


def _dataframe_records(dataframe: Any) -> list[dict[str, Any]]:
    return json.loads(dataframe.to_json(orient="records", force_ascii=False))


def _distance_checker_jobs() -> list[dict[str, Any]]:
    try:
        payload = json.loads(DISTANCE_CHECKER_JOBS_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, list) else []
    except Exception:
        return []


def _save_distance_checker_job(job: dict[str, Any]) -> None:
    DISTANCE_CHECKER_JOBS_PATH.parent.mkdir(parents=True, exist_ok=True)
    jobs = _distance_checker_jobs()
    jobs.insert(0, job)
    DISTANCE_CHECKER_JOBS_PATH.write_text(
        json.dumps(
            _json_safe(jobs[:MAX_DISTANCE_CHECKER_JOBS]),
            ensure_ascii=False,
            indent=2,
            allow_nan=False,
        ),
        encoding="utf-8",
    )


def _handle_distance_workbook_preview(payload: dict[str, Any]) -> dict[str, Any]:
    distance_tool = _distance_tool_module()

    def read_preview(source_label: str, temp_path: str) -> dict[str, Any]:
        sheet_names = list(distance_tool.get_excel_sheet_names(temp_path))
        if not sheet_names:
            raise ValueError("Workbook has no readable sheets.")
        requested_sheet = str(payload.get("selected_sheet") or "").strip()
        selected_sheet = (
            requested_sheet if requested_sheet in sheet_names else sheet_names[0]
        )
        source_df = distance_tool.read_excel_sheet(temp_path, sheet_name=selected_sheet)
        columns = [str(column) for column in list(source_df.columns)]
        inferred = distance_tool.infer_current_plan_columns(source_df)
        return {
            "source_label": source_label,
            "sheet_names": sheet_names,
            "selected_sheet": selected_sheet,
            "columns": columns,
            "row_count": int(len(source_df)),
            "sample_rows": _dataframe_records(source_df.head(8)),
            "suggested_columns": {
                "address": inferred.get("address") or (columns[0] if columns else ""),
                "city": inferred.get("city") or "",
                "country": inferred.get("country") or "",
                "route": inferred.get("route") or "",
                "sequence": inferred.get("sequence") or "",
                "bus_type": inferred.get("bus_type") or "",
            },
        }

    return _with_temp_workbook(payload, read_preview)


def _handle_reference_distance_check(payload: dict[str, Any]) -> dict[str, Any]:
    distance_tool = _distance_tool_module()
    origin = dict(payload.get("origin") or {})
    origin_country = str(origin.get("country") or "").strip()
    origin_city = str(origin.get("city") or "").strip()
    origin_address = str(origin.get("address") or "").strip()
    distance_mode = str(payload.get("distance_mode") or "road").strip()
    if distance_mode not in {"road", "straight_line"}:
        raise ValueError("distance_mode must be road or straight_line.")
    if not origin_address:
        raise ValueError("Reference stop address is required.")

    def run_check(source_label: str, temp_path: str) -> dict[str, Any]:
        sheet_names = list(distance_tool.get_excel_sheet_names(temp_path))
        requested_sheet = str(payload.get("selected_sheet") or "").strip()
        selected_sheet = (
            requested_sheet
            if requested_sheet in sheet_names
            else (sheet_names[0] if sheet_names else "")
        )
        if not selected_sheet:
            raise ValueError("Workbook has no readable sheets.")
        source_df = distance_tool.read_excel_sheet(temp_path, sheet_name=selected_sheet)
        columns = {str(column) for column in list(source_df.columns)}
        address_column = str(payload.get("address_column") or "").strip()
        city_column = str(payload.get("city_column") or "").strip() or None
        country_column = str(payload.get("country_column") or "").strip() or None
        if address_column not in columns:
            raise ValueError("Select a valid address column.")
        if city_column and city_column not in columns:
            raise ValueError("Select a valid city column or leave it blank.")
        if country_column and country_column not in columns:
            raise ValueError("Select a valid country column or leave it blank.")

        origin_rows, _ = distance_tool.geocode_records_for_distance_tool(
            [
                {
                    "source_excel_row": 1,
                    "country": origin_country,
                    "city": origin_city,
                    "address": origin_address,
                }
            ]
        )
        origin_row = dict(origin_rows[0])
        if origin_row.get("status") != "ok":
            raise RuntimeError(
                str(
                    origin_row.get("warning") or "Reference stop could not be geocoded."
                )
            )

        input_rows = distance_tool.build_distance_input_rows(
            source_df,
            address_column=address_column,
            city_column=city_column,
            country_column=country_column,
            default_city=origin_city,
            default_country=origin_country,
        )
        geocoded_rows, _ = distance_tool.geocode_records_for_distance_tool(input_rows)
        results_df = distance_tool.build_distance_result_dataframe(
            source_df,
            input_rows,
            geocoded_rows,
            origin_record=origin_row,
            origin_point=dict(origin_row["point"]),
            distance_mode=distance_mode,
        )
        records = _dataframe_records(results_df)
        ok_count = sum(1 for row in records if str(row.get("status")) == "ok")
        failed_count = sum(
            1 for row in records if str(row.get("status")) == "geocode_failed"
        )
        blank_count = sum(
            1 for row in records if str(row.get("status")) == "blank_address"
        )
        job = {
            "job_id": uuid4().hex[:12],
            "type": "reference_distance",
            "created_at": utc_now_iso(),
            "label": f"{Path(source_label).stem} from {origin_address}",
            "metadata": {
                "source_label": source_label,
                "selected_sheet": selected_sheet,
                "origin_country": origin_country,
                "origin_city": origin_city,
                "origin_address": origin_address,
                "distance_mode": distance_mode,
                "address_column": address_column,
                "city_column": city_column or "",
                "country_column": country_column or "",
            },
            "results": records,
        }
        _save_distance_checker_job(job)
        return {
            "job": {key: value for key, value in job.items() if key != "results"},
            "summary": {
                "row_count": len(records),
                "resolved_count": ok_count,
                "failed_count": failed_count,
                "blank_count": blank_count,
                "distance_mode": distance_mode,
            },
            "results": records,
        }

    return _with_temp_workbook(payload, run_check)


def _handle_current_plan_route_cost(payload: dict[str, Any]) -> dict[str, Any]:
    distance_tool = _distance_tool_module()
    default_city = str(payload.get("default_city") or "").strip()
    default_country = str(payload.get("default_country") or "").strip()
    diesel_price_per_liter = float(payload.get("diesel_price_per_liter") or 0.0)
    fuel_efficiency_km_per_liter = float(
        payload.get("fuel_efficiency_km_per_liter") or 0.0
    )
    currency_code = str(payload.get("currency_code") or "").strip().upper()
    currency_label = str(payload.get("currency_label") or currency_code or "").strip()
    if diesel_price_per_liter < 0:
        raise ValueError("diesel_price_per_liter must be zero or greater.")
    if fuel_efficiency_km_per_liter <= 0:
        raise ValueError("fuel_efficiency_km_per_liter must be greater than zero.")

    def run_route_cost(source_label: str, temp_path: str) -> dict[str, Any]:
        sheet_names = list(distance_tool.get_excel_sheet_names(temp_path))
        requested_sheet = str(payload.get("selected_sheet") or "").strip()
        selected_sheet = (
            requested_sheet
            if requested_sheet in sheet_names
            else (sheet_names[0] if sheet_names else "")
        )
        if not selected_sheet:
            raise ValueError("Workbook has no readable sheets.")
        source_df = distance_tool.read_excel_sheet(temp_path, sheet_name=selected_sheet)
        columns = {str(column) for column in list(source_df.columns)}
        route_column = str(payload.get("route_column") or "").strip()
        address_column = str(payload.get("address_column") or "").strip()
        sequence_column = str(payload.get("sequence_column") or "").strip() or None
        bus_type_column = str(payload.get("bus_type_column") or "").strip() or None
        city_column = str(payload.get("city_column") or "").strip() or None
        country_column = str(payload.get("country_column") or "").strip() or None
        required_columns = {
            "route_column": route_column,
            "address_column": address_column,
        }
        for label, column in required_columns.items():
            if column not in columns:
                raise ValueError(f"Select a valid {label.replace('_', ' ')}.")
        optional_columns = {
            "sequence_column": sequence_column,
            "bus_type_column": bus_type_column,
            "city_column": city_column,
            "country_column": country_column,
        }
        for label, column in optional_columns.items():
            if column and column not in columns:
                raise ValueError(
                    f"Select a valid {label.replace('_', ' ')} or leave it blank."
                )

        input_rows = distance_tool.build_current_plan_route_input_rows(
            source_df,
            route_column=route_column,
            address_column=address_column,
            sequence_column=sequence_column,
            bus_type_column=bus_type_column,
            city_column=city_column,
            country_column=country_column,
            default_city=default_city,
            default_country=default_country,
        )
        geocoded_rows, _ = distance_tool.geocode_records_for_distance_tool(input_rows)
        route_results_df, leg_results_df = (
            distance_tool.build_current_plan_route_cost_dataframe(
                input_rows,
                geocoded_rows,
                diesel_price_per_liter=diesel_price_per_liter,
                fuel_efficiency_km_per_liter=fuel_efficiency_km_per_liter,
            )
        )
        route_records = _dataframe_records(route_results_df)
        leg_records = _dataframe_records(leg_results_df)
        total_distance = sum(
            float(row.get("route_distance_km") or 0.0) for row in route_records
        )
        total_cost = sum(
            float(row.get("estimated_one_way_fuel_cost") or 0.0)
            for row in route_records
        )
        unresolved_routes = sum(
            1 for row in route_records if float(row.get("failed_stops") or 0.0) > 0
        )
        electric_routes = sum(
            1
            for row in route_records
            if str(row.get("diesel_cost_status") or "") == "skipped_electric_bus"
        )
        job = {
            "job_id": uuid4().hex[:12],
            "type": "route_cost",
            "created_at": utc_now_iso(),
            "label": f"{Path(source_label).stem} route cost",
            "metadata": {
                "source_label": source_label,
                "selected_sheet": selected_sheet,
                "default_city": default_city,
                "default_country": default_country,
                "currency_code": currency_code,
                "currency_label": currency_label,
                "diesel_price_per_liter": diesel_price_per_liter,
                "fuel_efficiency_km_per_liter": fuel_efficiency_km_per_liter,
                "route_column": route_column,
                "address_column": address_column,
                "sequence_column": sequence_column or "",
                "bus_type_column": bus_type_column or "",
                "city_column": city_column or "",
                "country_column": country_column or "",
            },
            "route_results": route_records,
            "leg_results": leg_records,
        }
        _save_distance_checker_job(job)
        return {
            "job": {
                key: value
                for key, value in job.items()
                if key not in {"route_results", "leg_results"}
            },
            "summary": {
                "route_count": len(route_records),
                "leg_count": len(leg_records),
                "total_one_way_distance_km": round(total_distance, 3),
                "estimated_one_way_fuel_cost": round(total_cost, 2),
                "routes_with_unresolved_stops": unresolved_routes,
                "electric_routes_skipped": electric_routes,
                "currency_code": currency_code,
                "currency_label": currency_label,
            },
            "route_results": route_records,
            "leg_results": leg_records,
        }

    return _with_temp_workbook(payload, run_route_cost)


def _parse_rider_counts_payload(value: Any) -> list[int]:
    if isinstance(value, list):
        chunks = [str(item) for item in value]
    else:
        chunks = str(value or "").replace("\n", ",").split(",")
    rider_counts: list[int] = []
    for chunk in chunks:
        text = str(chunk or "").strip()
        if not text:
            continue
        try:
            rider_count = int(float(text))
        except ValueError as exc:
            raise ValueError(f"Invalid rider group value: {text!r}") from exc
        if rider_count <= 0:
            raise ValueError(f"Rider group values must be greater than zero: {text!r}")
        rider_counts.append(rider_count)
    if not rider_counts:
        raise ValueError("Enter at least one rider group or upload a demand workbook.")
    return rider_counts


def _demand_workbook_from_payload(payload: dict[str, Any]) -> tuple[Any | None, str]:
    if not str(payload.get("file_base64") or "").strip():
        return None, ""
    source_label, workbook_bytes = _decode_workbook_bytes(payload)
    demand_input = _client_module("demand_input")
    return demand_input.read_demand_workbook(io.BytesIO(workbook_bytes)), source_label


def _fleet_route_time_target(payload: dict[str, Any]) -> int | None:
    raw_value = payload.get("max_route_duration_minutes")
    if raw_value is None or raw_value == "":
        return None
    try:
        minutes = int(raw_value)
    except (TypeError, ValueError) as exc:
        raise ValueError("max_route_duration_minutes must be a whole number.") from exc
    if minutes < 5 or minutes > 240:
        raise ValueError("max_route_duration_minutes must be between 5 and 240.")
    return minutes


def _fleet_vehicle_catalog_payload(
    payload: dict[str, Any],
) -> list[dict[str, Any]] | None:
    raw_catalog = payload.get("vehicle_catalog")
    if raw_catalog is None:
        return None
    if not isinstance(raw_catalog, list):
        raise ValueError("vehicle_catalog must be a list.")
    return [dict(item) for item in raw_catalog if isinstance(item, dict)]


def _fleet_catalog_rows(catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "vehicle_type": vehicle.get("vehicle_type"),
            "vehicle": vehicle.get("display_name"),
            "display_name": vehicle.get("display_name"),
            "category": vehicle.get("category"),
            "propulsion": vehicle.get("propulsion"),
            "listed_seats": vehicle.get("listed_seats"),
            "monitor_seats": vehicle.get("monitor_seats"),
            "student_capacity": vehicle.get("student_capacity"),
            "available_count": vehicle.get("available_count"),
            "enabled": vehicle.get("enabled", True),
            "notes": vehicle.get("notes"),
        }
        for vehicle in catalog
    ]


def _handle_fleet_planner_vehicle_catalog(payload: dict[str, Any]) -> dict[str, Any]:
    vehicle_catalog = _client_module("vehicle_catalog")
    market = str(payload.get("market") or "KR").strip().upper()
    monitor_seats = int(payload.get("monitor_seats") or 0)
    catalog = vehicle_catalog.get_vehicle_catalog(market, monitor_seats=monitor_seats)
    return {
        "summary": {
            "market": market if market in {"CN", "KR"} else "KR",
            "monitor_seats": max(0, monitor_seats),
            "vehicle_count": len(catalog),
            "source": "default",
        },
        "catalog": _fleet_catalog_rows(catalog),
    }


def _handle_fleet_planner_preview(payload: dict[str, Any]) -> dict[str, Any]:
    demand_input = _client_module("demand_input")
    fleet_selector = _client_module("fleet_selector")
    planning_assumptions = _client_module("planning_assumptions")
    vehicle_catalog = _client_module("vehicle_catalog")

    market = str(payload.get("market") or "KR").strip().upper()
    mode = str(payload.get("mode") or "balanced").strip()
    monitor_seats = int(payload.get("monitor_seats") or 0)
    max_route_duration_minutes = _fleet_route_time_target(payload)
    custom_catalog = _fleet_vehicle_catalog_payload(payload)
    demand_workbook, source_label = _demand_workbook_from_payload(payload)
    workbook_payload: dict[str, Any] | None = None
    if demand_workbook is not None:
        rider_counts = [int(item["student_count"]) for item in demand_workbook.riders]
        workbook_payload = {
            "source_label": source_label,
            "school": dict(demand_workbook.school),
            "summary": dict(demand_workbook.summary),
            "warnings": list(demand_workbook.warnings),
            "riders": _dataframe_records(
                demand_input.demand_riders_to_dataframe(demand_workbook.riders)
            ),
        }
    else:
        rider_counts = _parse_rider_counts_payload(payload.get("rider_counts"))

    assumptions = planning_assumptions.get_planning_assumptions(
        market,
        mode=mode,
        monitor_seats=monitor_seats,
        max_route_duration_minutes=max_route_duration_minutes,
    )
    recommendations: list[dict[str, Any]] = []
    decision_details: list[dict[str, Any]] = []
    for rider_count in rider_counts:
        selection = fleet_selector.select_vehicle_for_group(
            int(rider_count),
            market=market,
            mode=mode,
            monitor_seats=monitor_seats,
            assumptions=assumptions,
            custom_catalog=custom_catalog,
        )
        selected = dict(selection.selected_vehicle or {})
        recommendations.append(
            {
                "riders": rider_count,
                "recommended_vehicle": selected.get(
                    "display_name", "No feasible vehicle"
                ),
                "student_capacity": selected.get("student_capacity", ""),
                "load_factor": selected.get("load_factor"),
                "empty_seats": selected.get("empty_seats", ""),
                "feasible_options": len(selection.feasible_options),
                "rejected_options": len(selection.rejected_options),
            }
        )
        decision_details.append(
            {
                "riders": rider_count,
                "selected_vehicle": selected,
                "feasible_options": selection.feasible_options[:8],
                "rejected_options": selection.rejected_options[:8],
            }
        )

    mix_summary = fleet_selector.estimate_vehicle_mix_for_groups(
        rider_counts,
        market=market,
        mode=mode,
        monitor_seats=monitor_seats,
        max_route_duration_minutes=max_route_duration_minutes,
        custom_catalog=custom_catalog,
    )
    catalog = vehicle_catalog.get_vehicle_catalog(
        assumptions.market,
        monitor_seats=assumptions.monitor_seats,
        custom_catalog=custom_catalog,
    )
    if custom_catalog is not None and not catalog:
        raise ValueError(
            "Custom vehicle catalog has no enabled vehicles with usable seats."
        )

    return {
        "summary": {
            "market": assumptions.market,
            "mode": assumptions.mode,
            "monitor_seats": assumptions.monitor_seats,
            "max_route_duration_minutes": assumptions.max_route_duration_minutes,
            "group_count": len(rider_counts),
            "total_riders": sum(rider_counts),
            "source": "demand_workbook"
            if demand_workbook is not None
            else "manual_rider_groups",
            "vehicle_catalog_source": "custom"
            if custom_catalog is not None
            else "default",
            "vehicle_catalog_count": len(catalog),
        },
        "assumptions": assumptions.to_dict(),
        "demand_workbook": workbook_payload,
        "recommendations": recommendations,
        "mix_summary": mix_summary,
        "decision_details": decision_details,
        "catalog": _fleet_catalog_rows(catalog),
    }


def _handle_fleet_planner_geocode(payload: dict[str, Any]) -> dict[str, Any]:
    demand_input = _client_module("demand_input")
    demand_workbook, source_label = _demand_workbook_from_payload(payload)
    if demand_workbook is None:
        raise ValueError("Upload a demand workbook before running geocode preview.")
    geocode_result = demand_input.geocode_demand_workbook(demand_workbook)
    return {
        "source_label": source_label,
        "summary": dict(geocode_result.get("summary") or {}),
        "school": dict(geocode_result.get("school") or {}),
        "demand_points": list(geocode_result.get("demand_points") or []),
        "rows": _dataframe_records(
            demand_input.demand_geocode_results_to_dataframe(geocode_result)
        ),
        "map_html": demand_input.build_demand_geocode_map_html(geocode_result),
    }


def _handle_fleet_planner_clusters(payload: dict[str, Any]) -> dict[str, Any]:
    demand_clustering = _client_module("demand_clustering")
    market = str(payload.get("market") or "KR").strip().upper()
    mode = str(payload.get("mode") or "balanced").strip()
    monitor_seats = int(payload.get("monitor_seats") or 0)
    max_route_duration_minutes = _fleet_route_time_target(payload)
    custom_catalog = _fleet_vehicle_catalog_payload(payload)
    sector_count = int(payload.get("sector_count") or 8)
    if sector_count not in {4, 8, 12}:
        raise ValueError("sector_count must be 4, 8, or 12.")
    geocode_result = dict(payload.get("geocode_result") or {})
    if not geocode_result:
        raise ValueError("Run demand geocode before building clusters.")
    cluster_result = demand_clustering.build_demand_clusters(
        geocode_result,
        market=market,
        mode=mode,
        monitor_seats=monitor_seats,
        max_route_duration_minutes=max_route_duration_minutes,
        custom_catalog=custom_catalog,
        sector_count=sector_count,
    )
    return {
        "summary": dict(cluster_result.get("summary") or {}),
        "school": dict(cluster_result.get("school") or {}),
        "clusters": list(cluster_result.get("clusters") or []),
        "failed_points": list(cluster_result.get("failed_points") or []),
        "rows": _dataframe_records(
            demand_clustering.demand_clusters_to_dataframe(cluster_result)
        ),
        "stop_rows": _dataframe_records(
            demand_clustering.cluster_points_to_dataframe(cluster_result)
        ),
        "map_html": demand_clustering.build_demand_cluster_map_html(cluster_result),
    }


def _fleet_service_direction_label(service_direction: str) -> str:
    normalized = str(service_direction or "").strip().lower().replace("-", "_").replace(" ", "_")
    return "To School" if normalized == "to_school" else "From School"


def _fleet_default_traffic_profile(service_direction: str) -> str:
    return "AM Peak" if _fleet_service_direction_label(service_direction) == "To School" else "PM Peak"


def _fleet_traffic_input_records(route_payload: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    candidate_items: list[dict[str, Any]] = []
    school = route_payload.get("school") or {}
    if isinstance(school, dict):
        candidate_items.append(dict(school))
    for point in list(route_payload.get("demand_points") or []):
        if isinstance(point, dict):
            candidate_items.append(dict(point))
    for cluster in list(route_payload.get("clusters") or []):
        if not isinstance(cluster, dict):
            continue
        for point in list(cluster.get("points") or []):
            if isinstance(point, dict):
                candidate_items.append(dict(point))

    for item in candidate_items:
        records.append(
            {
                "country": str(item.get("country") or "").strip(),
                "city": str(item.get("city") or "").strip(),
                "address": str(item.get("address") or item.get("formatted_address") or "").strip(),
            }
        )
    return records


def _fleet_traffic_context(
    geocode_result: dict[str, Any],
    *,
    service_direction: str,
    market: str | None = None,
    profile_name: str | None = None,
) -> dict[str, Any]:
    records = _fleet_traffic_input_records(geocode_result)
    selected_profile = str(profile_name or "").strip() or _fleet_default_traffic_profile(service_direction)
    traffic_profile_name, traffic_time_multiplier, traffic_profile_context = resolve_traffic_profile(
        selected_profile,
        records,
    )
    inferred_country, _inferred_city = infer_traffic_location(records)
    normalized_market = str(market or "").strip().upper()
    live_traffic_sample = None
    if normalized_market == "KR" or inferred_country == "SOUTH KOREA":
        live_traffic_sample = summarize_live_traffic_samples(
            service_direction=_fleet_service_direction_label(service_direction),
            input_records=records,
        )
    if live_traffic_sample:
        traffic_profile_name = str(live_traffic_sample["traffic_profile_name"])
        traffic_time_multiplier = float(live_traffic_sample["traffic_time_multiplier"])
        traffic_profile_context = str(live_traffic_sample["traffic_profile_context"])
    return {
        "traffic_profile_name": traffic_profile_name,
        "traffic_time_multiplier": float(traffic_time_multiplier),
        "traffic_profile_context": traffic_profile_context,
        "live_traffic_sample": live_traffic_sample,
    }


def _handle_fleet_planner_route_preview(payload: dict[str, Any]) -> dict[str, Any]:
    demand_clustering = _client_module("demand_clustering")
    demand_routing = _client_module("demand_routing")
    market = str(payload.get("market") or "KR").strip().upper()
    mode = str(payload.get("mode") or "balanced").strip()
    monitor_seats = int(payload.get("monitor_seats") or 0)
    service_direction = str(payload.get("service_direction") or "to_school").strip()
    if service_direction not in {"to_school", "from_school"}:
        raise ValueError("service_direction must be to_school or from_school.")
    max_route_duration_minutes = _fleet_route_time_target(payload)
    custom_catalog = _fleet_vehicle_catalog_payload(payload)
    cluster_result = dict(payload.get("cluster_result") or {})
    if not cluster_result:
        raise ValueError("Build demand clusters before route preview.")

    traffic_context = _fleet_traffic_context(
        dict(cluster_result),
        service_direction=service_direction,
        profile_name=payload.get("traffic_profile_name"),
        market=market,
    )
    route_preview = demand_routing.build_osrm_route_preview(
        cluster_result,
        service_direction=service_direction,
        max_route_duration_minutes=max_route_duration_minutes,
        traffic_time_multiplier=float(traffic_context["traffic_time_multiplier"]),
        traffic_profile_name=str(traffic_context["traffic_profile_name"]),
        traffic_profile_context=str(traffic_context["traffic_profile_context"]),
        live_traffic_sample=traffic_context.get("live_traffic_sample"),
    )
    overlong_route_ids = {
        str(row.get("cluster_id", "")).strip()
        for row in list(route_preview.get("route_rows") or [])
        if max_route_duration_minutes
        and float(row.get("duration_min", 0.0) or 0.0)
        > float(max_route_duration_minutes)
    }
    if overlong_route_ids:
        refined_cluster_result = demand_clustering.split_cluster_result_by_route_limit(
            cluster_result,
            overlong_route_ids,
            market=market,
            mode=mode,
            monitor_seats=monitor_seats,
            max_route_duration_minutes=max_route_duration_minutes,
            custom_catalog=custom_catalog,
        )
        route_preview = demand_routing.build_osrm_route_preview(
            refined_cluster_result,
            service_direction=service_direction,
            max_route_duration_minutes=max_route_duration_minutes,
            traffic_time_multiplier=float(traffic_context["traffic_time_multiplier"]),
            traffic_profile_name=str(traffic_context["traffic_profile_name"]),
            traffic_profile_context=str(traffic_context["traffic_profile_context"]),
            live_traffic_sample=traffic_context.get("live_traffic_sample"),
        )
        route_preview["refinement_note"] = (
            "One or more clusters exceeded the route-duration target and were split once by distance from school."
        )

    return _route_plan_response(
        route_preview, workbook_file_name="fleet_planner_generated_plan.xlsx"
    )


def _service_direction_label(service_direction: str) -> str:
    return (
        "To School"
        if str(service_direction).strip().lower() == "to_school"
        else "From School"
    )


def _route_plan_response(
    route_preview: dict[str, Any], *, workbook_file_name: str
) -> dict[str, Any]:
    demand_routing = _client_module("demand_routing")
    workbook_bytes = demand_routing.build_generated_plan_workbook_bytes(route_preview)
    map_data = demand_routing.build_route_preview_map_data(
        route_preview,
        scenario_key="optimized_plan",
        scenario_name="Optimized Plan",
    )
    return {
        "summary": dict(route_preview.get("summary") or {}),
        "school": dict(route_preview.get("school") or {}),
        "routes": list(route_preview.get("routes") or []),
        "rows": _dataframe_records(
            demand_routing.route_preview_to_dataframe(route_preview)
        ),
        "stop_rows": _dataframe_records(
            demand_routing.route_preview_stop_detail_to_dataframe(route_preview)
        ),
        "map_html": demand_routing.build_route_preview_map_html(route_preview),
        "map_data": map_data,
        "refinement_note": str(route_preview.get("refinement_note") or ""),
        "workbook_file_name": workbook_file_name,
        "workbook_base64": base64.b64encode(workbook_bytes).decode("ascii"),
    }


def _ensure_fleet_planner_map_data(
    route_preview_result: dict[str, Any],
) -> dict[str, Any]:
    result = deepcopy(route_preview_result or {})
    if result.get("map_data") or not result.get("routes"):
        return result
    try:
        demand_routing = _client_module("demand_routing")
        result["map_data"] = demand_routing.build_route_preview_map_data(
            result,
            scenario_key="optimized_plan",
            scenario_name="Optimized Plan",
        )
    except Exception as exc:
        result["map_data_error"] = str(exc)
    return result


def _hydrate_fleet_planner_history_record(record: dict[str, Any]) -> dict[str, Any]:
    hydrated = deepcopy(record or {})
    global_plan_result = hydrated.get("global_plan_result")
    if isinstance(global_plan_result, dict):
        hydrated["global_plan_result"] = _ensure_fleet_planner_map_data(
            global_plan_result
        )
    return hydrated


def _handle_fleet_planner_global_plan(payload: dict[str, Any]) -> dict[str, Any]:
    demand_global_optimizer = _client_module("demand_global_optimizer")
    market = str(payload.get("market") or "KR").strip().upper()
    mode = str(payload.get("mode") or "balanced").strip()
    monitor_seats = int(payload.get("monitor_seats") or 0)
    max_route_duration_minutes = _fleet_route_time_target(payload)
    custom_catalog = _fleet_vehicle_catalog_payload(payload)
    service_direction = str(payload.get("service_direction") or "to_school").strip()
    if service_direction not in {"to_school", "from_school"}:
        raise ValueError("service_direction must be to_school or from_school.")
    geocode_result = dict(payload.get("geocode_result") or {})
    if not geocode_result:
        raise ValueError("Run demand geocode before building a global plan.")
    traffic_context = _fleet_traffic_context(
        geocode_result,
        service_direction=service_direction,
        profile_name=payload.get("traffic_profile_name"),
        market=market,
    )
    global_plan = demand_global_optimizer.build_global_ortools_plan(
        geocode_result,
        market=market,
        mode=mode,
        monitor_seats=monitor_seats,
        max_route_duration_minutes=max_route_duration_minutes,
        custom_catalog=custom_catalog,
        service_direction=service_direction,
        traffic_time_multiplier=float(traffic_context["traffic_time_multiplier"]),
        traffic_profile_name=str(traffic_context["traffic_profile_name"]),
        traffic_profile_context=str(traffic_context["traffic_profile_context"]),
        live_traffic_sample=traffic_context.get("live_traffic_sample"),
    )
    return _route_plan_response(
        global_plan, workbook_file_name="fleet_planner_global_plan.xlsx"
    )


def _handle_fleet_planner_history_create(
    payload: dict[str, Any], user_email: str
) -> dict[str, Any]:
    preview_result = dict(payload.get("preview_result") or {})
    global_plan_result = _ensure_fleet_planner_map_data(
        dict(payload.get("global_plan_result") or {})
    )
    if not preview_result:
        raise ValueError("Run Fleet preview before saving history.")
    if not global_plan_result:
        raise ValueError("Build an optimized plan before saving history.")

    scenario = dict(payload.get("scenario") or {})
    plan_summary = dict(global_plan_result.get("summary") or {})
    preview_summary = dict(preview_result.get("summary") or {})
    history_payload = {
        "title": str(payload.get("title") or "").strip(),
        "scenario": scenario,
        "preview_result": preview_result,
        "geocode_result": dict(payload.get("geocode_result") or {}),
        "cluster_result": dict(payload.get("cluster_result") or {}),
        "route_preview_result": dict(payload.get("route_preview_result") or {}),
        "global_plan_result": global_plan_result,
        "summary": {
            "market": scenario.get("market") or preview_summary.get("market"),
            "mode": scenario.get("mode") or preview_summary.get("mode"),
            "monitor_seats": scenario.get("monitor_seats")
            or preview_summary.get("monitor_seats"),
            "max_route_duration_minutes": (
                scenario.get("max_route_duration_minutes")
                or plan_summary.get("max_route_duration_minutes")
                or preview_summary.get("max_route_duration_minutes")
            ),
            "vehicle_catalog_source": scenario.get("vehicle_catalog_source")
            or preview_summary.get("vehicle_catalog_source"),
            "vehicle_catalog_count": scenario.get("vehicle_catalog_count")
            or preview_summary.get("vehicle_catalog_count"),
            "service_direction": scenario.get("service_direction")
            or plan_summary.get("service_direction"),
            "routes": plan_summary.get("route_count"),
            "students": preview_summary.get("total_riders"),
            "total_distance_km": plan_summary.get("total_distance_km"),
            "total_duration_min": plan_summary.get("total_duration_min"),
        },
    }
    return {
        "job": FLEET_PLANNER_HISTORY_STORE.create(
            history_payload, owner_email=user_email
        )
    }


def _handle_distance_checker_history_create(
    payload: dict[str, Any], user_email: str
) -> dict[str, Any]:
    tool_mode = str(payload.get("tool_mode") or "").strip()
    reference_result = dict(payload.get("reference_result") or {})
    route_cost_result = dict(payload.get("route_cost_result") or {})
    if not tool_mode:
        tool_mode = "route_cost" if route_cost_result else "reference"
    if tool_mode not in {"reference", "route_cost"}:
        raise ValueError("tool_mode must be reference or route_cost.")
    if tool_mode == "reference" and not reference_result:
        raise ValueError("Run a reference distance check before saving history.")
    if tool_mode == "route_cost" and not route_cost_result:
        raise ValueError("Run a route cost calculation before saving history.")

    preview = dict(payload.get("preview") or {})
    scenario = dict(payload.get("scenario") or {})
    result = route_cost_result if tool_mode == "route_cost" else reference_result
    result_job = dict(result.get("job") or {})
    result_summary = dict(result.get("summary") or {})
    metadata = dict(result_job.get("metadata") or {})
    source_label = str(
        scenario.get("file_name")
        or preview.get("source_label")
        or metadata.get("source_label")
        or ""
    ).strip()
    selected_sheet = str(
        scenario.get("selected_sheet")
        or preview.get("selected_sheet")
        or metadata.get("selected_sheet")
        or ""
    ).strip()
    title = str(payload.get("title") or result_job.get("label") or "").strip()
    if not title:
        title = f"{'Route Cost' if tool_mode == 'route_cost' else 'Reference Distance'} - {datetime.now().strftime('%Y-%m-%d %H%M')}"

    summary: dict[str, Any] = {
        "tool_mode": tool_mode,
        "source_label": source_label,
        "selected_sheet": selected_sheet,
    }
    if tool_mode == "reference":
        summary.update(
            {
                "row_count": result_summary.get("row_count"),
                "resolved_count": result_summary.get("resolved_count"),
                "failed_count": result_summary.get("failed_count"),
                "blank_count": result_summary.get("blank_count"),
                "distance_mode": result_summary.get("distance_mode")
                or scenario.get("distance_mode"),
                "origin_address": (
                    dict(scenario.get("origin") or {}).get("address")
                    or metadata.get("origin_address")
                    or ""
                ),
            }
        )
    else:
        summary.update(
            {
                "route_count": result_summary.get("route_count"),
                "leg_count": result_summary.get("leg_count"),
                "total_one_way_distance_km": result_summary.get(
                    "total_one_way_distance_km"
                ),
                "estimated_one_way_fuel_cost": result_summary.get(
                    "estimated_one_way_fuel_cost"
                ),
                "currency_code": result_summary.get("currency_code"),
                "currency_label": result_summary.get("currency_label"),
                "routes_with_unresolved_stops": result_summary.get(
                    "routes_with_unresolved_stops"
                ),
                "electric_routes_skipped": result_summary.get(
                    "electric_routes_skipped"
                ),
            }
        )

    history_payload = {
        "title": title,
        "scenario": scenario,
        "preview": preview,
        "reference_result": reference_result,
        "route_cost_result": route_cost_result,
        "summary": summary,
    }
    return {
        "job": _distance_history_store_for_mode(tool_mode).create(
            history_payload, owner_email=user_email
        )
    }


def _infer_service_direction_from_label(source_label: str) -> str:
    label = str(source_label or "").lower().replace("_", " ").replace("-", " ")
    if "to school" in label or "morning" in label:
        return "To School"
    if "from school" in label or "afternoon" in label:
        return "From School"
    return ""


def _read_current_plan_upload(
    payload: dict[str, Any],
) -> tuple[Any, str, dict[str, Any]]:
    client_core = _client_core_module()
    source_label, workbook_bytes = _decode_workbook_bytes(payload)
    config_payload = dict(payload.get("config") or {})
    preferred_direction = _infer_service_direction_from_label(source_label) or str(
        config_payload.get("service_direction")
        or payload.get("service_direction")
        or "From School"
    )
    directions = [
        preferred_direction,
        "To School" if preferred_direction == "From School" else "From School",
    ]
    suffix = Path(source_label).suffix.lower()
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file.write(workbook_bytes)
            temp_path = temp_file.name
        last_error: Exception | None = None
        for service_direction in directions:
            try:
                current_plan = client_core.read_current_plan_from_excel(
                    temp_path, service_direction=service_direction
                )
                break
            except Exception as exc:
                last_error = exc
        else:
            assert last_error is not None
            raise last_error
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except OSError:
                pass
    return client_core, source_label, current_plan


def _find_subway_aggregation_block_reason(
    client_core: Any, records: list[dict[str, Any]]
) -> str | None:
    is_likely_english_korean_address = getattr(
        client_core.runtime, "is_likely_english_korean_address", None
    )
    if not is_likely_english_korean_address:
        return None
    for item in records:
        country = str(item.get("country", "")).strip()
        address = str(item.get("address", "")).strip()
        if is_likely_english_korean_address(country, address):
            return (
                "Subway aggregation is unavailable for South Korea rows that use English-only addresses, "
                "because those stops require Google geocoding."
            )
    return None


def _suggest_planner_config_from_current_plan(
    current_plan: dict[str, Any],
    config_payload: dict[str, Any],
) -> dict[str, Any]:
    suggested = _planner_config_payload(config_payload)
    normalized_fleet: list[dict[str, Any]] = []
    for item in list(current_plan.get("fleet") or []):
        bus_type = str(item.get("bus_type", "")).strip()
        if not bus_type:
            continue
        normalized_fleet.append(
            {
                "bus_type": bus_type,
                "seat_count": int(item.get("seat_count", 0) or 0),
                "vehicle_count": int(item.get("vehicle_count", 0) or 0),
            }
        )
    normalized_fleet.sort(
        key=lambda item: (
            -int(item["seat_count"]),
            -int(item["vehicle_count"]),
            str(item["bus_type"]).lower(),
        )
    )
    slot_defaults = [
        ("large", "Large Bus", 42),
        ("mid", "Mid Bus", 35),
        ("small", "Small Bus", 19),
    ]
    for index, (slot_key, default_name, default_capacity) in enumerate(slot_defaults):
        fleet_item = normalized_fleet[index] if index < len(normalized_fleet) else {}
        slot_name = (
            str(fleet_item.get("bus_type", default_name)).strip() or default_name
        )
        seat_count = int(
            fleet_item.get("seat_count", default_capacity) or default_capacity
        )
        vehicle_count = int(fleet_item.get("vehicle_count", 0) or 0)
        suggested[f"{slot_key}_bus_name"] = slot_name
        suggested[f"{slot_key}_bus_capacity"] = seat_count
        suggested[f"{slot_key}_bus_max_count"] = vehicle_count
        suggested[f"free_baseline_{slot_key}_bus_ratio"] = float(vehicle_count)
    suggested["service_direction"] = str(
        current_plan.get("service_direction")
        or suggested.get("service_direction")
        or "From School"
    )
    if suggested["service_direction"] == "To School":
        suggested["traffic_profile_name"] = "AM Peak"
        suggested["to_school_arrival_time"] = str(
            config_payload.get("to_school_arrival_time")
            or suggested.get("to_school_arrival_time")
            or "08:00"
        )
        suggested["time_window_start"] = str(config_payload.get("time_window_start") or "06:30")
        suggested["time_window_end"] = str(config_payload.get("time_window_end") or "08:00")
    else:
        suggested["traffic_profile_name"] = "PM Peak"
        departure_time = str(
            config_payload.get("from_school_departure_time")
            or suggested.get("from_school_departure_time")
            or "15:40"
        )
        suggested["from_school_departure_time"] = departure_time
        raw_window = (
            str(config_payload.get("time_window_start") or ""),
            str(config_payload.get("time_window_end") or ""),
        )
        if raw_window in {("", ""), ("06:30", "08:00")}:
            departure_label, departure_minutes = _parse_clock_payload(
                departure_time,
                "from_school_departure_time",
                "15:40",
            )
            window_end_minutes = departure_minutes + 120
            suggested["time_window_start"] = departure_label
            suggested["time_window_end"] = f"{(window_end_minutes // 60) % 24:02d}:{window_end_minutes % 60:02d}"
        else:
            suggested["time_window_start"] = raw_window[0]
            suggested["time_window_end"] = raw_window[1]
    if "include_subway_aggregation_scenario" not in config_payload:
        suggested["include_subway_aggregation_scenario"] = False
    if "include_nearby_aggregation_scenario" not in config_payload:
        suggested["include_nearby_aggregation_scenario"] = False
    return suggested


def _client_geocode_cache_key(
    client_core: Any,
    country: str,
    city: str,
    address: str,
) -> str:
    runtime = getattr(client_core, "runtime", None)
    key_func = getattr(runtime, "geocode_cache_key", None)
    if callable(key_func):
        try:
            return str(key_func(country, city, address))
        except Exception:
            pass
    return f"{country.strip()}|{city.strip()}|{address.strip()}"


def _client_geocode_cache_lookup_keys(
    client_core: Any,
    country: str,
    city: str,
    address: str,
) -> list[str]:
    runtime = getattr(client_core, "runtime", None)
    lookup_func = getattr(runtime, "geocode_cache_lookup_keys", None)
    if callable(lookup_func):
        try:
            return [str(item) for item in list(lookup_func(country, city, address) or []) if str(item)]
        except Exception:
            pass
    key = _client_geocode_cache_key(client_core, country, city, address)
    return [key] if key else []


def _address_review_record_key(client_core: Any, item: dict[str, Any]) -> str:
    return _client_geocode_cache_key(
        client_core,
        str(item.get("country") or item.get("requested_country") or "").strip(),
        str(item.get("city") or item.get("requested_city") or "").strip(),
        str(item.get("address") or item.get("requested_address") or item.get("display_address") or "").strip(),
    )


def _address_review_rows_label(rows: list[int]) -> str:
    values = sorted({int(row) for row in rows if int(row) > 0})
    return ", ".join(str(row) for row in values)


def _address_review_failure_warnings(
    input_records: list[dict[str, Any]],
    reason: str,
) -> list[dict[str, str]]:
    warnings: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for item in input_records:
        country = str(item.get("country") or "").strip()
        city = str(item.get("city") or "").strip()
        address = str(item.get("address") or "").strip()
        key = (country, city, address)
        if not address or key in seen:
            continue
        seen.add(key)
        warnings.append(
            {
                "country": country,
                "city": city,
                "address": address,
                "warning": reason,
                "suggestion": "Correct the workbook address or clear the cached geocode, then upload again.",
            }
        )
    return warnings


def _build_address_review(
    client_core: Any,
    input_records: list[dict[str, Any]],
    prepared_payload: dict[str, Any],
    geocode_warnings: list[dict[str, Any]],
) -> dict[str, Any]:
    records: dict[str, dict[str, Any]] = {}
    ordered_keys: list[str] = []
    address_to_keys: dict[str, list[str]] = {}
    for item in input_records:
        country = str(item.get("country") or "").strip()
        city = str(item.get("city") or "").strip()
        address = str(item.get("address") or "").strip()
        if not address:
            continue
        key = _client_geocode_cache_key(client_core, country, city, address)
        if key not in records:
            records[key] = {
                "country": country,
                "city": city,
                "address": address,
                "source_excel_rows": [],
            }
            ordered_keys.append(key)
            address_to_keys.setdefault(address.lower(), []).append(key)
        source_row = item.get("source_excel_row")
        if source_row not in (None, "", 0):
            records[key]["source_excel_rows"].append(int(source_row))

    points_by_key: dict[str, dict[str, Any]] = {}
    for point in list(prepared_payload.get("original_points") or []):
        point_dict = dict(point or {})
        key = _address_review_record_key(client_core, point_dict)
        if key and key not in points_by_key:
            points_by_key[key] = point_dict

    warnings_by_key: dict[str, dict[str, Any]] = {}
    for warning in geocode_warnings:
        warning_dict = dict(warning or {})
        key = _address_review_record_key(client_core, warning_dict)
        if key not in records:
            candidates = address_to_keys.get(str(warning_dict.get("address") or "").strip().lower(), [])
            if len(candidates) == 1:
                key = candidates[0]
        if key:
            warnings_by_key[key] = warning_dict

    items: list[dict[str, Any]] = []
    for key in ordered_keys:
        record = records[key]
        point = points_by_key.get(key)
        warning = warnings_by_key.get(key)
        warning_text = str((warning or {}).get("warning") or (warning or {}).get("reason") or "").strip()
        status = "ok"
        reason = ""
        suggestion = ""
        if point is None:
            status = "blocking"
            reason = warning_text or "This address could not be resolved to coordinates."
            suggestion = str((warning or {}).get("suggestion") or "Correct the workbook address or clear the cached geocode, then upload again.").strip()
        else:
            point_status = str(point.get("geocode_status") or point.get("validation_status") or "").strip().lower()
            warning_status = str((warning or {}).get("status") or "").strip().lower()
            if point_status == "needs_review" or warning_status == "needs_review":
                status = "needs_review"
            reason = warning_text or str(point.get("warning") or "").strip()
            suggestion = str((warning or {}).get("suggestion") or "").strip()

        cache_keys = _client_geocode_cache_lookup_keys(
            client_core,
            str(record.get("country") or ""),
            str(record.get("city") or ""),
            str(record.get("address") or ""),
        )
        item = {
            "id": key,
            "status": status,
            "severity": "error" if status == "blocking" else "warning" if status == "needs_review" else "ok",
            "country": str(record.get("country") or ""),
            "city": str(record.get("city") or ""),
            "address": str(record.get("address") or ""),
            "source_excel_rows": _address_review_rows_label(list(record.get("source_excel_rows") or [])),
            "provider": str((point or {}).get("provider") or ""),
            "formatted_address": str((point or warning or {}).get("formatted_address") or ""),
            "lat": (point or {}).get("lat"),
            "lng": (point or {}).get("lng"),
            "adcode": str((point or {}).get("adcode") or ""),
            "reason": reason,
            "suggestion": suggestion,
            "cache_keys": cache_keys,
        }
        if warning and warning.get("distance_to_school_km") is not None:
            item["distance_to_school_km"] = warning.get("distance_to_school_km")
        items.append(item)

    blocking_count = sum(1 for item in items if item["status"] == "blocking")
    review_count = sum(1 for item in items if item["status"] == "needs_review")
    ok_count = sum(1 for item in items if item["status"] == "ok")
    return {
        "status": "blocked" if blocking_count else "needs_review" if review_count else "ok",
        "blocking_count": blocking_count,
        "review_count": review_count,
        "ok_count": ok_count,
        "total_count": len(items),
        "requires_acknowledgement": review_count > 0,
        "items": items,
    }


def _address_review_block_message(address_review: dict[str, Any]) -> str:
    blocking_count = int(address_review.get("blocking_count") or 0)
    review_count = int(address_review.get("review_count") or 0)
    if blocking_count:
        return f"Address review has {blocking_count} unresolved address(es). Correct the workbook or clear cached geocodes before running audit."
    if review_count:
        return f"Address review has {review_count} warning(s). Review and acknowledge them before running audit."
    return ""


def _current_plan_preview_map(
    current_plan: dict[str, Any],
    prepared_payload: dict[str, Any],
    config_payload: dict[str, Any],
) -> tuple[dict[str, Any] | None, str | None]:
    points = [dict(item) for item in list(prepared_payload.get("original_points") or [])]
    if not points:
        return None, "No geocoded points are available for map preview."
    try:
        planner = load_legacy_planner()
        config = _build_planner_config(config_payload)
        assessment_time, assessment_distance = _build_assessment_metric_matrices(
            planner,
            points,
        )
        current_plan_assessment = assess_current_plan(
            planner,
            current_plan,
            points,
            config,
            solve_time=assessment_time,
            solve_distance=assessment_distance,
        )
        scenario = build_current_plan_map_scenario(
            planner,
            current_plan_assessment,
            points,
        )
        job_record = {
            "job_id": "workbook-preview",
            "result": {
                "service_direction": str(
                    getattr(config, "service_direction", "") or ""
                ).strip(),
                "traffic_profile_name": str(
                    getattr(config, "traffic_profile_name", "") or ""
                ).strip(),
                "structured_results": {"current_plan": scenario},
            },
        }
        return _build_job_map_payload(
            job_record,
            "current_plan",
            "current_plan",
            attach_impact=False,
        )
    except Exception as exc:
        return None, str(exc) or exc.__class__.__name__


def _workbook_preview_response(payload: dict[str, Any]) -> dict[str, Any]:
    client_core, source_label, current_plan = _read_current_plan_upload(payload)
    config_payload = dict(payload.get("config") or {})
    input_records = [
        dict(item) for item in list(current_plan.get("input_records") or [])
    ]
    block_reason = _find_subway_aggregation_block_reason(client_core, input_records)
    suggested_config = _suggest_planner_config_from_current_plan(
        current_plan, config_payload
    )
    if block_reason:
        suggested_config["include_subway_aggregation_scenario"] = False
    auto_route_budget: dict[str, Any] = {"status": "unavailable", "reason": "not_calculated"}
    address_review: dict[str, Any]
    current_plan_map: dict[str, Any] | None = None
    current_plan_map_error: str | None = None
    try:
        client_config = _build_client_planner_config(client_core, suggested_config)
        client_prep = client_core.prepare_client_payload(
            input_records,
            current_plan_data=current_plan,
            config=client_config,
        )
        prepared_payload = dict(client_prep["prepared_payload"])
        auto_route_budget = _auto_current_plan_route_budget_details(
            current_plan,
            prepared_payload,
        ) or {"status": "unavailable", "reason": "no_measurable_current_routes"}
        address_review = _build_address_review(
            client_core,
            input_records,
            prepared_payload,
            [dict(item) for item in list(client_prep.get("geocode_warnings") or [])],
        )
        current_plan_map, current_plan_map_error = _current_plan_preview_map(
            current_plan,
            prepared_payload,
            suggested_config,
        )
    except Exception as exc:
        auto_route_budget = {"status": "unavailable", "reason": exc.__class__.__name__}
        current_plan_map_error = str(exc) or exc.__class__.__name__
        address_review = _build_address_review(
            client_core,
            input_records,
            {},
            _address_review_failure_warnings(input_records, str(exc) or exc.__class__.__name__),
        )
    auto_route_budget_minutes = auto_route_budget.get("minutes")
    if auto_route_budget_minutes is not None:
        suggested_config["max_route_duration_minutes"] = int(auto_route_budget_minutes)
    return {
        "source_label": source_label,
        "selected_sheet": "current_plan_assignments",
        "job_default_name": _build_job_display_name(source_label),
        "summary": dict(current_plan.get("summary") or {}),
        "fleet": list(current_plan.get("fleet") or []),
        "input_record_count": _service_input_record_count(input_records),
        "subway_aggregation_block_reason": block_reason,
        "auto_route_budget": auto_route_budget,
        "address_review": address_review,
        "current_plan_map": current_plan_map,
        "current_plan_map_error": current_plan_map_error,
        "suggested_config": suggested_config,
    }


def _route_budget_address_key(item: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(item.get("country") or "").strip().lower(),
        str(item.get("city") or "").strip().lower(),
        str(item.get("address") or item.get("display_address") or item.get("requested_address") or "").strip().lower(),
    )


def _auto_route_budget_location(
    current_plan: dict[str, Any],
    points: list[dict[str, Any]],
) -> tuple[str, str]:
    country, city = infer_traffic_location(
        [dict(item) for item in list(current_plan.get("input_records") or [])]
    )
    if country:
        return str(country).upper(), str(city or "")
    for item in [*points, *list(current_plan.get("stops") or [])]:
        raw_country = str(dict(item or {}).get("country") or "").strip()
        raw_city = str(dict(item or {}).get("city") or "").strip()
        if raw_country.upper() in {"CN", "CHINA"} or raw_country in {"中国", "中國"}:
            return "CHINA", raw_city
    return "", ""


def _attach_current_plan_amap_budget_details(
    planner: Any,
    details: dict[str, Any],
    current_plan: dict[str, Any],
    points: list[dict[str, Any]],
    route_nodes_by_id: dict[str, list[int]],
) -> None:
    country, city = _auto_route_budget_location(current_plan, points)
    details["amap_route_country"] = country
    details["amap_route_city"] = city
    if country != "CHINA":
        details["amap_route_status"] = "not_applicable"
        details["amap_route_reason"] = "non_china_location"
        return
    if not str(getattr(planner, "AMAP_KEY", "") or "").strip():
        details["amap_route_status"] = "unavailable"
        details["amap_route_reason"] = "missing_amap_key"
        return
    measurable_routes = {
        route_id: nodes
        for route_id, nodes in dict(route_nodes_by_id or {}).items()
        if route_id and len(nodes) >= 2
    }
    if not measurable_routes:
        details["amap_route_status"] = "unavailable"
        details["amap_route_reason"] = "missing_route_nodes"
        return

    try:
        cache = load_json_object(FINAL_ROUTE_TRAFFIC_CACHE_PATH)
        state = {"api_calls": 0, "cache_hits": 0, "cache_changed": 0}
        max_route_id = ""
        max_duration_s = 0.0
        max_drive_duration_s = 0.0
        max_distance_m = 0.0
        max_point_count = 0
        measured_count = 0
        source = ""
        stop_service_s = max(0, int(getattr(planner, "STOP_SERVICE_SECONDS", 60) or 60))
        for route_id, route_nodes in measurable_routes.items():
            request_points = _route_amap_points(points, {"nodes": route_nodes})
            stats = _amap_route_stats(planner, request_points, cache, state)
            if not stats:
                continue
            drive_duration_s = float(stats.get("duration_s", 0.0) or 0.0)
            duration_s = drive_duration_s + sum(1 for node in route_nodes if int(node) != 0) * stop_service_s
            distance_m = float(stats.get("distance_m", 0.0) or 0.0)
            measured_count += 1
            if duration_s > max_duration_s:
                max_duration_s = duration_s
                max_drive_duration_s = drive_duration_s
                max_distance_m = distance_m
                max_route_id = str(route_id)
                max_point_count = len(request_points)
                source = str(stats.get("source") or "")
        if state.get("cache_changed"):
            save_json_object(FINAL_ROUTE_TRAFFIC_CACHE_PATH, cache, sort_keys=True)
    except Exception as exc:
        details["amap_route_status"] = "unavailable"
        details["amap_route_reason"] = exc.__class__.__name__
        return

    if max_duration_s <= 0:
        details["amap_route_status"] = "unavailable"
        details["amap_route_reason"] = "no_amap_route"
        return
    amap_budget_minutes = max(5, min(240, int(math.ceil(max_duration_s / 60.0))))
    if amap_budget_minutes > int(details.get("minutes", 0) or 0):
        details["minutes"] = amap_budget_minutes
        details["source"] = "max_current_plan_amap_route"
    details.update(
        {
            "amap_route_status": "ready",
            "amap_route_source": source,
            "amap_route_id": max_route_id,
            "amap_route_duration_minutes": round(max_duration_s / 60.0, 1),
            "amap_route_drive_duration_minutes": round(max_drive_duration_s / 60.0, 1),
            "amap_route_distance_km": round(max_distance_m / 1000.0, 1),
            "amap_route_api_calls": int(state.get("api_calls", 0) or 0),
            "amap_route_cache_hits": int(state.get("cache_hits", 0) or 0),
            "amap_route_point_count": max_point_count,
            "amap_route_measured_count": measured_count,
            "amap_budget_minutes": amap_budget_minutes,
        }
    )


def _auto_current_plan_route_budget_details(
    current_plan: dict[str, Any],
    prepared_payload: dict[str, Any],
) -> dict[str, Any] | None:
    points = [dict(item) for item in list(prepared_payload.get("original_points") or [])]
    if len(points) < 2:
        return None
    point_by_address = {
        _route_budget_address_key(point): int(point.get("node_id", index) or index)
        for index, point in enumerate(points)
    }
    stop_lookup = {
        str(stop.get("stop_id") or "").strip(): dict(stop)
        for stop in list(current_plan.get("stops") or [])
    }
    route_groups: dict[str, list[dict[str, Any]]] = {}
    for assignment in list(current_plan.get("assignments") or []):
        route_id = str(assignment.get("route_id") or "").strip()
        stop_id = str(assignment.get("stop_id") or "").strip()
        stop = stop_lookup.get(stop_id)
        if route_id and stop:
            route_groups.setdefault(route_id, []).append({**stop, **dict(assignment)})
    if not route_groups:
        for stop in list(current_plan.get("stops") or []):
            route_groups.setdefault(str(stop.get("route_id") or "").strip(), []).append(dict(stop))
    if not route_groups:
        return None

    planner = load_legacy_planner()
    planner.TRAFFIC_TIME_MULTIPLIER = 1.0
    previous_osrm_base_url = getattr(planner, "OSRM_BASE_URL", "")
    try:
        resolver = getattr(planner, "resolve_osrm_base_url", None)
        if callable(resolver):
            planner.OSRM_BASE_URL = resolver(points)
        time_matrix, _distance_matrix = planner.build_osrm_full_matrix(points)
    finally:
        if hasattr(planner, "OSRM_BASE_URL"):
            planner.OSRM_BASE_URL = previous_osrm_base_url
    raw_time_matrix = getattr(planner, "RAW_SOLVER_TIME_MATRIX", None) or time_matrix
    longest_s = 0
    longest_route_id = ""
    longest_route_nodes: list[int] = []
    route_nodes_by_id: dict[str, list[int]] = {}
    measured_route_count = 0
    service_direction = str(current_plan.get("service_direction") or "").strip()
    for route_id, route_stops in route_groups.items():
        ordered_stops = sorted(route_stops, key=lambda item: int(item.get("stop_sequence", 0) or 0))
        nodes = [
            point_by_address.get(_route_budget_address_key(stop))
            for stop in ordered_stops
        ]
        nodes = [node for node in nodes if node is not None]
        if nodes and 0 not in nodes:
            if service_direction == "To School":
                nodes = [*nodes, 0]
            else:
                nodes = [0, *nodes]
        if len(nodes) < 2:
            continue
        route_nodes_by_id[str(route_id)] = nodes
        route_s = sum(int(raw_time_matrix[a][b] or 0) for a, b in zip(nodes, nodes[1:]))
        measured_route_count += 1
        if route_s > longest_s:
            longest_s = route_s
            longest_route_id = route_id
            longest_route_nodes = nodes
    if longest_s <= 0:
        return None
    minutes = max(5, min(240, int(math.ceil(longest_s / 60.0))))
    details = {
        "status": "ready",
        "source": "longest_current_plan_osrm_route",
        "minutes": minutes,
        "longest_route_id": longest_route_id,
        "longest_route_duration_minutes": round(longest_s / 60.0, 1),
        "longest_route_node_count": len(longest_route_nodes),
        "measured_route_count": measured_route_count,
        "route_count": len(route_groups),
        "osrm_budget_minutes": minutes,
    }
    _attach_current_plan_amap_budget_details(planner, details, current_plan, points, route_nodes_by_id)
    return details



def _auto_current_plan_route_budget_minutes(
    current_plan: dict[str, Any],
    prepared_payload: dict[str, Any],
) -> int | None:
    details = _auto_current_plan_route_budget_details(current_plan, prepared_payload)
    if not details:
        return None
    minutes = details.get("minutes")
    return int(minutes) if minutes is not None else None


def _auto_route_budget_from_current_plan(
    client_core: Any,
    current_plan: dict[str, Any],
    config_payload: dict[str, Any],
) -> dict[str, Any]:
    input_records = [dict(item) for item in list(current_plan.get("input_records") or [])]
    if not input_records:
        return {"status": "unavailable", "reason": "no_input_records"}
    try:
        client_config = _build_client_planner_config(client_core, config_payload)
        client_prep = client_core.prepare_client_payload(
            input_records,
            current_plan_data=current_plan,
            config=client_config,
        )
        details = _auto_current_plan_route_budget_details(
            current_plan,
            dict(client_prep["prepared_payload"]),
        )
    except Exception as exc:
        return {"status": "unavailable", "reason": exc.__class__.__name__}
    return details or {"status": "unavailable", "reason": "no_measurable_current_routes"}


def _handle_workbook_preview(payload: dict[str, Any]) -> dict[str, Any]:
    return _workbook_preview_response(payload)


def _handle_workbook_submit(payload: dict[str, Any], user_email: str) -> dict[str, Any]:
    client_core, source_label, current_plan = _read_current_plan_upload(payload)
    config_payload = _planner_config_payload(dict(payload.get("config") or {}))
    input_records = [
        dict(item) for item in list(current_plan.get("input_records") or [])
    ]
    block_reason = _find_subway_aggregation_block_reason(client_core, input_records)
    if block_reason:
        config_payload["include_subway_aggregation_scenario"] = False
    client_config = _build_client_planner_config(client_core, config_payload)
    client_prep = client_core.prepare_client_payload(
        input_records,
        current_plan_data=current_plan,
        config=client_config,
    )
    address_review = _build_address_review(
        client_core,
        input_records,
        dict(client_prep["prepared_payload"]),
        [dict(item) for item in list(client_prep.get("geocode_warnings") or [])],
    )
    if int(address_review.get("blocking_count") or 0) > 0:
        raise ValueError(_address_review_block_message(address_review))
    if bool(address_review.get("requires_acknowledgement")) and not bool(payload.get("address_review_acknowledged")):
        raise ValueError(_address_review_block_message(address_review))
    auto_route_budget: dict[str, Any] = {"status": "unavailable", "reason": "not_calculated"}
    auto_route_budget_minutes: int | None = None
    try:
        auto_route_budget = _auto_current_plan_route_budget_details(
            current_plan,
            dict(client_prep["prepared_payload"]),
        ) or {"status": "unavailable", "reason": "no_measurable_current_routes"}
        if auto_route_budget.get("minutes") is not None:
            auto_route_budget_minutes = int(auto_route_budget["minutes"])
    except Exception:
        auto_route_budget = {"status": "unavailable", "reason": "calculation_failed"}
    if auto_route_budget_minutes is not None:
        config_payload["max_route_duration_minutes"] = auto_route_budget_minutes
    job_custom_name = str(payload.get("job_custom_name") or "").strip()
    job_default_name = _build_job_display_name(source_label)
    job_name = _build_job_display_name(source_label, job_custom_name)
    scheduled_requested = bool(payload.get("scheduled_job"))
    if scheduled_requested and not SCHEDULED_JOBS_ENABLED:
        raise ValueError("Scheduled jobs are not enabled for this deployment.")
    scheduled_start_at = None
    scheduled_trigger_label = None
    if scheduled_requested:
        scheduled_start_at, scheduled_trigger_label = _next_scheduled_job_trigger(
            config_payload,
            payload.get("scheduled_date"),
        )
    metadata = {
        "job_name": job_name,
        "job_default_name": job_default_name,
        "job_custom_name": job_custom_name,
        "source_label": source_label,
        "selected_sheet": "current_plan_assignments",
        "scheduled_job": scheduled_requested,
        "scheduled_date": payload.get("scheduled_date") if scheduled_requested else None,
        "scheduled_start_at": scheduled_start_at,
        "scheduled_trigger_label": scheduled_trigger_label,
        "planner_config": dict(config_payload),
        "client_prep": {
            "geocode_warnings": list(client_prep.get("geocode_warnings") or []),
            "excluded_stops": list(client_prep.get("excluded_stops") or []),
            "elapsed_seconds": float(client_prep.get("elapsed_seconds", 0.0) or 0.0),
            "logs": str(client_prep.get("logs", "") or ""),
            "auto_route_budget": auto_route_budget,
            "auto_route_budget_minutes": auto_route_budget_minutes,
            "address_review": address_review,
        },
    }
    summary = JOB_STORE.create_job(
        config_payload,
        dict(client_prep["prepared_payload"]),
        metadata=metadata,
        owner_email=user_email,
        status="scheduled" if scheduled_requested else "queued",
        scheduled_start_at=scheduled_start_at,
        scheduled_trigger_label=scheduled_trigger_label,
    )
    if not scheduled_requested:
        spawned = _spawn_job_worker(str(summary["job_id"]))
        if spawned:
            summary["worker_pid"] = spawned.get("worker_pid")
    return {
        "job": summary,
        "source_label": source_label,
        "selected_sheet": "current_plan_assignments",
        "summary": dict(current_plan.get("summary") or {}),
        "client_prep": metadata["client_prep"],
        "subway_aggregation_block_reason": block_reason,
        "address_review": address_review,
    }


def _normalized_cache_text(value: object) -> str:
    return " ".join(str(value or "").strip().split()).lower()


def _remove_geocode_cache_entries(
    cache: dict[str, Any],
    keys: list[str],
    *,
    address: str,
) -> list[str]:
    target_address = _normalized_cache_text(address)
    key_set = {str(key) for key in keys if str(key)}
    removed: list[str] = []
    for key in list(cache.keys()):
        entry = cache.get(key)
        key_address = str(key).split("|")[-1] if "|" in str(key) else ""
        entry_address = ""
        if isinstance(entry, dict):
            entry_address = str(
                entry.get("requested_address")
                or entry.get("address")
                or entry.get("display_address")
                or ""
            )
        if key in key_set or (target_address and _normalized_cache_text(key_address) == target_address) or (
            target_address and _normalized_cache_text(entry_address) == target_address
        ):
            cache.pop(key, None)
            removed.append(str(key))
    return removed


def _handle_geocode_cache_clear(payload: dict[str, Any]) -> dict[str, Any]:
    country = str(payload.get("country") or "").strip()
    city = str(payload.get("city") or "").strip()
    address = str(payload.get("address") or "").strip()
    if not address:
        raise ValueError("Address is required.")

    removed: dict[str, list[str]] = {"client": [], "backend": []}
    client_core = _client_core_module()
    runtime = getattr(client_core, "runtime", None)
    client_cache = getattr(runtime, "GEOCODE_CACHE", None)
    client_keys = _client_geocode_cache_lookup_keys(client_core, country, city, address)
    if isinstance(client_cache, dict):
        removed["client"] = _remove_geocode_cache_entries(
            client_cache,
            client_keys,
            address=address,
        )
        if removed["client"]:
            save_func = getattr(runtime, "save_json_cache", None)
            cache_path = getattr(runtime, "GEOCODE_CACHE_PATH", None)
            if callable(save_func) and cache_path is not None:
                save_func(cache_path, client_cache)

    planner = load_legacy_planner()
    backend_cache = getattr(planner, "GEOCODE_CACHE", None)
    backend_key_func = getattr(planner, "geocode_cache_key", None)
    backend_keys: list[str] = []
    if callable(backend_key_func):
        backend_keys.append(str(backend_key_func(country, city, address)))
    if isinstance(backend_cache, dict):
        removed["backend"] = _remove_geocode_cache_entries(
            backend_cache,
            backend_keys,
            address=address,
        )
        if removed["backend"]:
            save_func = getattr(planner, "save_json_cache", None)
            cache_path = getattr(planner, "GEOCODE_CACHE_PATH", None)
            if callable(save_func) and cache_path is not None:
                save_func(cache_path, backend_cache)

    return {
        "cleared": sum(len(items) for items in removed.values()),
        "removed": removed,
        "country": country,
        "city": city,
        "address": address,
    }


def _normalize_email(value: object) -> str:
    return str(value or "").strip().lower()


def _json_safe(value: Any) -> Any:
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    return value


def _is_admin_email(email: str) -> bool:
    normalized_email = _normalize_email(email)
    return bool(normalized_email and normalized_email in ADMIN_EMAILS)


def _auth_display_name() -> str:
    if AUTH_DISPLAY_NAME:
        return AUTH_DISPLAY_NAME
    if AUTH_PROVIDER in {"microsoft_sso_pending", "microsoft_oidc", "saml2"}:
        return "Microsoft SSO"
    if AUTH_PROVIDER in {"cloudflare", "cloudflare_header"}:
        return "Cloudflare Access"
    if AUTH_PROVIDER == "local":
        return "Local development"
    return AUTH_PROVIDER.replace("_", " ").title()


def _auth_login_url() -> str:
    if AUTH_LOGIN_URL:
        return AUTH_LOGIN_URL
    if AUTH_PROVIDER == "microsoft_sso_pending":
        return "/api/auth/login"
    return "/"


def _auth_logout_url() -> str:
    if AUTH_LOGOUT_URL:
        return AUTH_LOGOUT_URL
    if AUTH_PROVIDER in {"cloudflare", "cloudflare_header"}:
        return "/cdn-cgi/access/logout"
    return "/"


def _auth_config_payload() -> dict[str, Any]:
    return {
        "provider": AUTH_PROVIDER,
        "display_name": _auth_display_name(),
        "login_url": _auth_login_url(),
        "logout_url": _auth_logout_url(),
        "sso_ready": AUTH_PROVIDER not in {"microsoft_sso_pending"},
        "admin_source": "local_env",
    }


def _service_input_record_count(input_records: list[dict[str, Any]]) -> int:
    if not input_records:
        return 0
    count = 0
    for item in input_records:
        try:
            passenger_count = int(item.get("passenger_count", 0) or 0)
        except Exception:
            passenger_count = 0
        if passenger_count > 0:
            count += 1
    return count


def _summarize_prepared_payload(prepared_payload: dict[str, Any]) -> dict[str, Any]:
    input_records = list(prepared_payload.get("input_records") or [])
    current_plan = dict(prepared_payload.get("current_plan") or {})
    current_plan_summary = dict(current_plan.get("summary") or {})
    return {
        "input_record_count": _service_input_record_count(input_records),
        "input_point_count": len(input_records),
        "country_samples": sorted(
            {
                str(item.get("country", "")).strip()
                for item in input_records
                if str(item.get("country", "")).strip()
            }
        )[:10],
        "city_samples": sorted(
            {
                str(item.get("city", "")).strip()
                for item in input_records
                if str(item.get("city", "")).strip()
            }
        )[:10],
        "has_current_plan": bool(current_plan),
        "current_plan_route_count": int(
            current_plan_summary.get("route_count", 0) or 0
        ),
        "current_plan_assignment_count": int(
            current_plan_summary.get("assignment_count", 0) or 0
        ),
        "current_plan_service_stop_count": int(
            current_plan_summary.get(
                "service_stop_count", current_plan_summary.get("stop_count", 0)
            )
            or 0
        ),
        "current_plan_scheduled_assignment_count": int(
            current_plan_summary.get("scheduled_assignment_count", 0) or 0
        ),
    }


def _google_geocode_usage_month_key() -> str:
    return datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%Y-%m")


def _google_geocode_usage_payload() -> dict[str, Any]:
    if not GOOGLE_GEOCODE_USAGE_VISIBLE:
        return {"enabled": False}
    month_key = _google_geocode_usage_month_key()
    store = SqliteQuotaStore()
    store.migrate_flat_month_usage(
        source_path=GOOGLE_GEOCODE_USAGE_PATH,
        provider=GOOGLE_GEOCODE_USAGE_PROVIDER,
        counter=GOOGLE_GEOCODE_USAGE_COUNTER,
        provider_label=GOOGLE_GEOCODE_USAGE_PROVIDER_LABEL,
        sku_estimate=GOOGLE_GEOCODE_USAGE_SKU_ESTIMATE,
    )
    try:
        used = max(0, int(store.sum_usage(period_type="month", period_key=month_key, contains="google")))
    except Exception:
        used = 0
    return {
        "enabled": True,
        "month_key": month_key,
        "used": used,
        "limit": GOOGLE_GEOCODE_MONTHLY_LIMIT,
        "label": f"Google API usage this month: {used:,} / {GOOGLE_GEOCODE_MONTHLY_LIMIT:,}",
    }


def _deployment_features_payload() -> dict[str, Any]:
    root_text = str(BASE_DIR).replace("\\", "/").lower()
    if "/staging/" in root_text:
        available_languages = ["en", "ko", "zh"]
    elif "users/bus.eim/brp" in root_text:
        available_languages = ["en", "ko"]
    else:
        available_languages = ["en", "zh"]
    return {
        "language_switch_enabled": ENABLE_LANGUAGE_SWITCH,
        "available_languages": available_languages,
        "scheduled_jobs_enabled": SCHEDULED_JOBS_ENABLED,
        "default_traffic_coefficient_mode": DEFAULT_TRAFFIC_COEFFICIENT_MODE,
    }


def _osrm_manager_status_payload() -> dict[str, Any]:
    try:
        report = osrm_manager.manager_status()
    except Exception as exc:
        return {
            "status": "error",
            "error": str(exc),
        }
    regions = [row for row in report.get("regions", []) if isinstance(row, dict)]
    locks = [row for row in report.get("locks", []) if isinstance(row, dict)]
    running_regions = []
    idle_expired_regions = []
    for row in regions:
        container_status = row.get("container_status")
        if isinstance(container_status, dict) and container_status.get("running"):
            running_regions.append(row.get("region"))
        if row.get("idle_expired"):
            idle_expired_regions.append(row.get("region"))
    return {
        "status": "ok",
        "summary": {
            "region_count": len(regions),
            "running_region_count": len(running_regions),
            "running_regions": running_regions,
            "idle_expired_region_count": len(idle_expired_regions),
            "idle_expired_regions": idle_expired_regions,
            "lock_count": len(locks),
            "locked_lock_count": sum(1 for row in locks if row.get("locked")),
            "stale_lock_count": sum(1 for row in locks if row.get("stale")),
            "on_demand_enabled": bool(report.get("on_demand_enabled")),
            "lock_wait_seconds": report.get("lock_wait_seconds"),
            "max_running_regions": report.get("max_running_regions"),
            "running_managed_regions": report.get("running_managed_regions") or [],
            "available_memory_mb": report.get("available_memory_mb"),
        },
        "manager": report,
    }


_TRAFFIC_ROLLOUT_STATUS_MODULE: Any | None = None


def _load_traffic_rollout_status_module() -> Any:
    global _TRAFFIC_ROLLOUT_STATUS_MODULE
    if _TRAFFIC_ROLLOUT_STATUS_MODULE is not None:
        return _TRAFFIC_ROLLOUT_STATUS_MODULE
    script_dir = REPO_ROOT / "ops" / "scripts"
    script_path = script_dir / "report_traffic_rollout_status.py"
    if not script_path.exists():
        raise RuntimeError(f"missing traffic rollout status script: {script_path}")
    if str(script_dir) not in sys.path:
        sys.path.insert(0, str(script_dir))
    spec = importlib.util.spec_from_file_location(
        "brp_report_traffic_rollout_status",
        script_path,
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load traffic rollout status script: {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    _TRAFFIC_ROLLOUT_STATUS_MODULE = module
    return module


def _query_bool(query_params: dict[str, str], name: str, default: bool = True) -> bool:
    raw = str(query_params.get(name, "")).strip().lower()
    if not raw:
        return default
    if raw in {"0", "false", "no", "off"}:
        return False
    if raw in {"1", "true", "yes", "on"}:
        return True
    return default


def _traffic_rollout_status_payload(query_params: dict[str, str] | None = None) -> dict[str, Any]:
    query_params = dict(query_params or {})
    try:
        report_module = _load_traffic_rollout_status_module()
        readiness_module = report_module.report_traffic_rollout_readiness
        min_geo_ratio = float(query_params.get("min_geo_ratio") or 1.0)
        min_geo_ratio = min(1.0, max(0.0, min_geo_ratio))
        sample_dir = Path(
            query_params.get("sample_dir")
            or os.environ.get("BRP_LIVE_TRAFFIC_SAMPLE_DIR", "")
            or readiness_module.report_live_traffic_readiness.DEFAULT_SAMPLE_DIR
        )
        min_measured_at = (
            str(query_params.get("min_measured_at") or "").strip()
            or os.environ.get("BRP_TRAFFIC_ROLLOUT_MIN_MEASURED_AT", "")
            or readiness_module.DEFAULT_CUTOFF
        )
        local_timezone = str(
            query_params.get("local_timezone")
            or report_module.DEFAULT_LOCAL_TIMEZONE
        )
        profiles = (
            report_module.required_profiles_for_current_environment()
            if hasattr(report_module, "required_profiles_for_current_environment")
            else list(readiness_module.DEFAULT_PROFILES)
        )
        report = report_module.build_status(
            sample_dir=sample_dir,
            min_measured_at=min_measured_at,
            profiles=profiles,
            min_geo_ratio=min_geo_ratio,
            include_timers=_query_bool(query_params, "include_timers", True),
            include_osrm=_query_bool(query_params, "include_osrm", True),
            include_budget=_query_bool(query_params, "include_budget", True),
            include_remote=_query_bool(query_params, "include_remote", True),
            local_timezone=local_timezone,
        )
        report["endpoint"] = {
            "read_only": True,
            "provider_api_called": bool(
                dict(report.get("api_budget") or {}).get("provider_api_called")
            ),
            "osrm_started": bool(dict(report.get("api_budget") or {}).get("osrm_started")),
        }
        return report
    except Exception as exc:
        return {
            "status": "error",
            "error": str(exc),
            "endpoint": {
                "read_only": True,
                "provider_api_called": False,
                "osrm_started": False,
            },
        }


def _resolve_staleness_seconds(timestamp: datetime | None) -> float | None:
    if timestamp is None:
        return None
    if timestamp.tzinfo is None:
        timestamp = timestamp.replace(tzinfo=timezone.utc)
    return (
        datetime.now(timezone.utc) - timestamp.astimezone(timezone.utc)
    ).total_seconds()


def _normalize_ai_audit_language(language: Any) -> tuple[str, str]:
    raw = str(language or "").strip()
    normalized = raw.lower()
    if (
        normalized.startswith("ko")
        or normalized.startswith("kr")
        or "korean" in normalized
        or "한국" in raw
        or "한글" in raw
    ):
        return "ko", "Korean"
    return "en", "English"


def _ai_audit_language_key(language: Any) -> str:
    return _normalize_ai_audit_language(language)[0]


def _ai_audit_report_map(record: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not isinstance(record, dict):
        return {}
    reports_by_language: dict[str, dict[str, Any]] = {}
    reports = record.get("ai_audit_reports")
    if isinstance(reports, dict):
        for key, value in reports.items():
            if isinstance(value, dict) and value:
                reports_by_language[_ai_audit_language_key(key)] = dict(value)
    legacy_report = record.get("ai_audit_report")
    if isinstance(legacy_report, dict) and legacy_report:
        legacy_key = _ai_audit_language_key(legacy_report.get("language") or "English")
        reports_by_language.setdefault(legacy_key, dict(legacy_report))
    return reports_by_language


def _ai_audit_record_with_decision_context(job_record: dict[str, Any]) -> dict[str, Any]:
    enriched = deepcopy(job_record)
    result = dict(enriched.get("result") or {})
    if not result:
        return enriched

    time_constrained = dict(result.get("time_constrained_optimization") or {})
    existing_summary = (
        dict(time_constrained.get("time_impact") or {})
        or dict(dict(time_constrained.get("summary") or {}).get("time_impact") or {})
    )
    if existing_summary:
        return enriched

    payload, _error = _build_job_map_payload(
        enriched,
        "time_constrained",
        "time_constrained",
        attach_impact=True,
    )
    time_impact = dict(dict(payload or {}).get("summary") or {}).get("time_impact")
    if not isinstance(time_impact, dict) or not time_impact:
        return enriched

    time_constrained["time_impact"] = time_impact
    result["time_constrained_optimization"] = time_constrained
    structured = dict(result.get("structured_results") or {})
    structured_time_constrained = dict(structured.get("time_constrained") or {})
    structured_time_constrained["time_impact"] = time_impact
    structured["time_constrained"] = structured_time_constrained
    result["structured_results"] = structured
    enriched["result"] = result
    return enriched


def generate_ai_audit_report(
    job_record: dict[str, Any], *, force: bool = False, language: str | None = None
) -> dict[str, Any]:
    return _generate_ai_audit_report(
        _ai_audit_record_with_decision_context(job_record),
        force=force,
        language=language,
    )


def _select_ai_audit_report(
    reports_by_language: dict[str, dict[str, Any]], requested_key: str
) -> dict[str, Any]:
    return dict(
        reports_by_language.get(requested_key)
        or reports_by_language.get("en")
        or reports_by_language.get("ko")
        or {}
    )


def _flatten_location_marker_values(value: Any, depth: int = 0) -> list[str]:
    if depth > 4:
        return []
    if isinstance(value, str):
        return [value]
    if isinstance(value, (int, float, bool)):
        return [str(value)]
    if isinstance(value, list):
        markers: list[str] = []
        for item in value[:20]:
            markers.extend(_flatten_location_marker_values(item, depth + 1))
        return markers
    if isinstance(value, dict):
        markers: list[str] = []
        for item in value.values():
            markers.extend(_flatten_location_marker_values(item, depth + 1))
        return markers
    return []


def _collect_location_markers(value: Any, depth: int = 0) -> list[str]:
    if depth > 5:
        return []
    markers: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            key_text = str(key).lower()
            key_is_location = any(
                fragment in key_text
                for fragment in (
                    "city",
                    "country",
                    "locale",
                    "market",
                    "region",
                    "traffic_location",
                )
            )
            if key_is_location:
                markers.extend(_flatten_location_marker_values(item))
            if isinstance(item, (dict, list)):
                markers.extend(_collect_location_markers(item, depth + 1))
    elif isinstance(value, list):
        for item in value[:30]:
            markers.extend(_collect_location_markers(item, depth + 1))
    return markers


def _is_korean_ai_audit_job(record: dict[str, Any]) -> bool:
    markers: list[str] = []
    for section_key in ("prepared_payload_summary", "metadata", "config", "result"):
        markers.extend(_collect_location_markers(record.get(section_key)))
    marker_text = " | ".join(markers).lower()
    if not marker_text:
        return False
    return any(
        token in marker_text
        for token in ("south korea", "korea", "seoul", "kr", "대한민국", "한국")
    )


class JobStore:
    def __init__(self, jobs_dir: Path) -> None:
        self.jobs_dir = jobs_dir
        self.lock = threading.Lock()
        _runtime_sqlite_store().initialize()
        self.reconcile_running_jobs()

    def _load_job_unlocked(self, job_id: str) -> dict[str, Any] | None:
        try:
            return _runtime_sqlite_store().get_job(str(job_id or "").strip())
        except Exception:
            traceback.print_exc()
            return None

    def _list_jobs_unlocked(
        self, user_email: str = "", include_all: bool = False
    ) -> list[dict[str, Any]]:
        try:
            return _runtime_sqlite_store().list_jobs(
                user_email=user_email, include_all=include_all
            )
        except Exception:
            traceback.print_exc()
            return []

    def _save_job_unlocked(self, job_id: str, record: dict[str, Any]) -> None:
        record["job_id"] = str(record.get("job_id") or job_id).strip()
        _save_runtime_job(record)

    def create_job(
        self,
        config_payload: dict[str, Any],
        prepared_payload: dict[str, Any],
        metadata: dict[str, Any] | None = None,
        owner_email: str = "",
        status: str = "queued",
        scheduled_start_at: str | None = None,
        scheduled_trigger_label: str | None = None,
    ) -> dict[str, Any]:
        job_id = uuid4().hex[:12]
        created_at = utc_now_iso()
        normalized_owner_email = _normalize_email(owner_email)
        initial_status = str(status or "queued").strip().lower() or "queued"
        record = {
            "job_id": job_id,
            "owner_email": normalized_owner_email,
            "status": initial_status,
            "created_at": created_at,
            "started_at": None,
            "finished_at": None,
            "scheduled_start_at": scheduled_start_at,
            "scheduled_trigger_label": scheduled_trigger_label,
            "worker_pid": None,
            "job_slot_path": None,
            "config": deepcopy(config_payload or {}),
            "prepared_payload": deepcopy(prepared_payload or {}),
            "prepared_payload_summary": _summarize_prepared_payload(
                prepared_payload or {}
            ),
            "metadata": deepcopy(metadata or {}),
            "result": None,
            "error": None,
            "traceback": None,
        }
        with self.lock:
            self._save_job_unlocked(job_id, record)
        return {
            "job_id": job_id,
            "owner_email": normalized_owner_email,
            "status": initial_status,
            "created_at": created_at,
            "started_at": None,
            "finished_at": None,
            "scheduled_start_at": scheduled_start_at,
            "scheduled_trigger_label": scheduled_trigger_label,
            "metadata": deepcopy(metadata or {}),
            "prepared_payload_summary": record["prepared_payload_summary"],
            "error": None,
        }

    def update_job(self, job_id: str, **changes: Any) -> dict[str, Any] | None:
        with self.lock:
            record = self._load_job_unlocked(job_id)
            if not record:
                return None
            record.update(changes)
            self._save_job_unlocked(job_id, record)
            return deepcopy(record)

    def claim_queued_job(
        self,
        job_id: str,
        *,
        worker_pid: int | None = None,
        job_slot_path: str | None = None,
    ) -> dict[str, Any] | None:
        with self.lock:
            try:
                record = _runtime_sqlite_store().claim_queued_job(
                    job_id,
                    worker_pid=worker_pid,
                    job_slot_path=job_slot_path,
                )
            except Exception:
                traceback.print_exc()
                return None
            return deepcopy(record) if record else None

    def begin_ai_audit(
        self,
        job_id: str,
        *,
        force: bool = False,
        required_languages: list[str] | None = None,
    ) -> tuple[str, dict[str, Any] | None]:
        with self.lock:
            record = self._load_job_unlocked(job_id)
            if not record:
                return "missing", None
            required_keys = {
                _ai_audit_language_key(language)
                for language in (required_languages or ["English"])
            }
            existing_reports = _ai_audit_report_map(record)
            if (
                required_keys
                and all(key in existing_reports for key in required_keys)
                and not force
            ):
                return "cached", deepcopy(record)
            if str(record.get("ai_audit_status", "")).strip().lower() == "running":
                return "running", deepcopy(record)
            record["ai_audit_status"] = "running"
            record["ai_audit_started_at"] = utc_now_iso()
            record["ai_audit_finished_at"] = None
            record["ai_audit_error"] = None
            self._save_job_unlocked(job_id, record)
            return "started", deepcopy(record)

    def get_job(self, job_id: str) -> dict[str, Any] | None:
        with self.lock:
            record = self._load_job_unlocked(job_id)
            return deepcopy(record) if record else None

    def list_jobs(
        self, user_email: str = "", include_all: bool = False
    ) -> list[dict[str, Any]]:
        with self.lock:
            return deepcopy(
                self._list_jobs_unlocked(
                    user_email=user_email, include_all=include_all
                )
            )

    def list_queued_jobs(self) -> list[dict[str, Any]]:
        with self.lock:
            entries = self._list_jobs_unlocked(include_all=True)
            records: list[dict[str, Any]] = []
            for entry in entries:
                if str(entry.get("status", "")).strip().lower() != "queued":
                    continue
                job_id = str(entry.get("job_id", "")).strip()
                if not job_id:
                    continue
                record = self._load_job_unlocked(job_id)
                if record and str(record.get("status", "")).strip().lower() == "queued":
                    records.append(deepcopy(record))
            records.sort(key=lambda item: str(item.get("created_at") or ""))
            return records

    def release_due_scheduled_jobs(self) -> list[dict[str, Any]]:
        now = datetime.now(timezone.utc)
        released: list[dict[str, Any]] = []
        with self.lock:
            entries = self._list_jobs_unlocked(include_all=True)
            for entry in entries:
                if str(entry.get("status", "")).strip().lower() != "scheduled":
                    continue
                job_id = str(entry.get("job_id", "")).strip()
                if not job_id:
                    continue
                record = self._load_job_unlocked(job_id)
                if not record or str(record.get("status", "")).strip().lower() != "scheduled":
                    continue
                scheduled_at = _parse_iso_datetime(
                    record.get("scheduled_start_at")
                    or dict(record.get("metadata") or {}).get("scheduled_start_at")
                )
                if scheduled_at is None or scheduled_at > now:
                    continue
                metadata = dict(record.get("metadata") or {})
                metadata["scheduled_released_at"] = utc_now_iso()
                record["metadata"] = metadata
                record["status"] = "queued"
                record["worker_pid"] = None
                record["job_slot_path"] = None
                self._save_job_unlocked(job_id, record)
                released.append(deepcopy(record))
        return released

    def release_scheduled_job(self, job_id: str, *, release_mode: str = "manual") -> dict[str, Any] | None:
        normalized_job_id = str(job_id or "").strip()
        if not normalized_job_id:
            return None
        with self.lock:
            record = self._load_job_unlocked(normalized_job_id)
            if not record:
                return None
            if str(record.get("status", "")).strip().lower() != "scheduled":
                return deepcopy(record)
            metadata = dict(record.get("metadata") or {})
            metadata["scheduled_released_at"] = utc_now_iso()
            metadata["scheduled_release_mode"] = str(release_mode or "manual").strip() or "manual"
            record["metadata"] = metadata
            record["status"] = "queued"
            record["worker_pid"] = None
            record["job_slot_path"] = None
            self._save_job_unlocked(normalized_job_id, record)
            return deepcopy(record)

    def delete_job(self, job_id: str) -> bool:
        with self.lock:
            return _delete_runtime_job(job_id)

    def reconcile_running_jobs(self) -> None:
        with self.lock:
            index_entries = self._list_jobs_unlocked(include_all=True)
            for entry in index_entries:
                job_id = str(entry.get("job_id", "")).strip()
                if not job_id:
                    continue
                record = self._load_job_unlocked(job_id)
                if not record:
                    continue
                status = str(record.get("status", "")).strip().lower()
                if status == "queued":
                    record["worker_pid"] = None
                    record["job_slot_path"] = None
                    self._save_job_unlocked(job_id, record)
                    continue
                if status == "running":
                    record["status"] = "failed"
                    record["finished_at"] = utc_now_iso()
                    record["error"] = (
                        "Job was interrupted because the backend service restarted."
                    )
                    record["traceback"] = None
                    record["worker_pid"] = None
                    record["job_slot_path"] = None
                    self._save_job_unlocked(job_id, record)


class SideToolHistoryStore:
    def __init__(self, root_dir: Path, tool_key: str) -> None:
        self.tool_key = tool_key
        self.lock = threading.Lock()
        _runtime_sqlite_store().initialize()

    def _load_record_unlocked(self, run_id: str) -> dict[str, Any] | None:
        try:
            return _runtime_sqlite_store().get_side_tool_run(
                self.tool_key, str(run_id or "").strip()
            )
        except Exception:
            traceback.print_exc()
            return None

    def _list_summaries_unlocked(
        self, user_email: str = "", include_all: bool = False
    ) -> list[dict[str, Any]]:
        try:
            return _runtime_sqlite_store().list_side_tool_runs(
                self.tool_key, user_email=user_email, include_all=include_all
            )
        except Exception:
            traceback.print_exc()
            return []

    def _summary_for_record(self, record: dict[str, Any]) -> dict[str, Any]:
        return {
            "run_id": str(record.get("run_id") or ""),
            "tool_key": self.tool_key,
            "owner_email": _normalize_email(record.get("owner_email")),
            "title": str(record.get("title") or ""),
            "created_at": record.get("created_at"),
            "shared_with_all": bool(record.get("shared_with_all")),
            "summary": deepcopy(record.get("summary") or {}),
        }

    def create(self, payload: dict[str, Any], owner_email: str) -> dict[str, Any]:
        run_id = uuid4().hex[:12]
        created_at = utc_now_iso()
        normalized_owner_email = _normalize_email(owner_email)
        title = (
            str(payload.get("title") or "").strip()
            or f"Fleet Planner Run - {datetime.now().strftime('%Y-%m-%d %H%M')}"
        )
        record = {
            "run_id": run_id,
            "tool_key": self.tool_key,
            "owner_email": normalized_owner_email,
            "title": title,
            "created_at": created_at,
            "scenario": deepcopy(payload.get("scenario") or {}),
            "shared_with_all": bool(payload.get("shared_with_all")),
            "preview_result": deepcopy(payload.get("preview_result") or {}),
            "geocode_result": deepcopy(payload.get("geocode_result") or {}),
            "cluster_result": deepcopy(payload.get("cluster_result") or {}),
            "route_preview_result": deepcopy(payload.get("route_preview_result") or {}),
            "global_plan_result": deepcopy(payload.get("global_plan_result") or {}),
            "preview": deepcopy(payload.get("preview") or {}),
            "reference_result": deepcopy(payload.get("reference_result") or {}),
            "route_cost_result": deepcopy(payload.get("route_cost_result") or {}),
            "summary": deepcopy(payload.get("summary") or {}),
        }
        summary = self._summary_for_record(record)
        with self.lock:
            _save_runtime_side_tool(self.tool_key, record)
        return summary

    def get(self, run_id: str) -> dict[str, Any] | None:
        with self.lock:
            record = self._load_record_unlocked(run_id)
            return deepcopy(record) if record else None

    def list(
        self, user_email: str = "", include_all: bool = False
    ) -> list[dict[str, Any]]:
        with self.lock:
            return deepcopy(
                self._list_summaries_unlocked(
                    user_email=user_email, include_all=include_all
                )
            )

    def delete(self, run_id: str) -> bool:
        with self.lock:
            return _delete_runtime_side_tool(self.tool_key, run_id)


JOB_STORE = JobStore(JOBS_DIR)
DISTANCE_CHECKER_HISTORY_STORE = SideToolHistoryStore(
    SIDE_TOOLS_DIR, "distance_checker"
)
REFERENCE_DISTANCE_HISTORY_STORE = SideToolHistoryStore(
    SIDE_TOOLS_DIR, "reference_distance"
)
ROUTE_COST_HISTORY_STORE = SideToolHistoryStore(SIDE_TOOLS_DIR, "route_cost")
FLEET_PLANNER_HISTORY_STORE = SideToolHistoryStore(SIDE_TOOLS_DIR, "fleet_planner")
JOB_QUEUE = JobQueueManager(
    job_store=JOB_STORE,
    runner_path=JOB_RUNNER_PATH,
    base_dir=BASE_DIR,
    python_executable=sys.executable,
    max_concurrent_jobs=MAX_CONCURRENT_JOBS,
    concurrency_dir=JOB_CONCURRENCY_DIR,
    poll_seconds=JOB_QUEUE_POLL_SECONDS,
    slot_attach_stale_seconds=JOB_SLOT_ATTACH_STALE_SECONDS,
)
JOB_GATE = JOB_QUEUE.gate


def _normalize_distance_history_mode(value: Any) -> str:
    return (
        "route_cost"
        if str(value or "").strip().lower() == "route_cost"
        else "reference"
    )


def _distance_history_store_for_mode(tool_mode: str) -> SideToolHistoryStore:
    return (
        ROUTE_COST_HISTORY_STORE
        if _normalize_distance_history_mode(tool_mode) == "route_cost"
        else REFERENCE_DISTANCE_HISTORY_STORE
    )


def _distance_history_mode_for_summary(entry: dict[str, Any]) -> str:
    summary = dict(entry.get("summary") or {})
    return _normalize_distance_history_mode(summary.get("tool_mode"))


def _distance_history_mode_for_record(record: dict[str, Any]) -> str:
    summary = dict(record.get("summary") or {})
    scenario = dict(record.get("scenario") or {})
    mode_value = summary.get("tool_mode") or scenario.get("tool_mode")
    if not mode_value and record.get("route_cost_result"):
        mode_value = "route_cost"
    return _normalize_distance_history_mode(mode_value)


def _list_distance_history(
    tool_mode: str, *, user_email: str, include_all: bool
) -> list[dict[str, Any]]:
    mode = _normalize_distance_history_mode(tool_mode) if tool_mode else ""
    stores = (
        [REFERENCE_DISTANCE_HISTORY_STORE, ROUTE_COST_HISTORY_STORE]
        if not mode
        else [_distance_history_store_for_mode(mode)]
    )
    entries: list[dict[str, Any]] = []
    for store in stores:
        entries.extend(store.list(user_email=user_email, include_all=include_all))
    legacy_entries = [
        entry
        for entry in DISTANCE_CHECKER_HISTORY_STORE.list(
            user_email=user_email, include_all=include_all
        )
        if not mode or _distance_history_mode_for_summary(entry) == mode
    ]
    entries.extend(legacy_entries)

    seen: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for entry in sorted(
        entries, key=lambda item: str(item.get("created_at") or ""), reverse=True
    ):
        run_id = str(entry.get("run_id") or "").strip()
        if not run_id or run_id in seen:
            continue
        seen.add(run_id)
        deduped.append(entry)
    return deduped


def _get_distance_history_record(
    run_id: str, tool_mode: str = ""
) -> tuple[dict[str, Any] | None, SideToolHistoryStore | None]:
    mode = _normalize_distance_history_mode(tool_mode) if tool_mode else ""
    stores = (
        [REFERENCE_DISTANCE_HISTORY_STORE, ROUTE_COST_HISTORY_STORE]
        if not mode
        else [_distance_history_store_for_mode(mode)]
    )
    stores = [*stores, DISTANCE_CHECKER_HISTORY_STORE]
    for store in stores:
        record = store.get(run_id)
        if not record:
            continue
        if mode and _distance_history_mode_for_record(record) != mode:
            continue
        return record, store
    return None, None


def _process_is_alive(pid: int | None) -> bool:
    return JOB_QUEUE.process_is_alive(pid)


def _worker_creation_flags() -> int:
    return worker_creation_flags()


def _spawn_job_worker(job_id: str) -> dict[str, Any] | None:
    return JOB_QUEUE.spawn_job_worker(job_id)


def _schedule_queued_jobs() -> None:
    JOB_QUEUE.schedule_queued_jobs()


def _release_scheduled_job(job_id: str) -> dict[str, Any] | None:
    released = JOB_STORE.release_scheduled_job(job_id, release_mode="manual")
    if not released:
        return None
    if str(released.get("status", "")).strip().lower() == "queued":
        JOB_QUEUE.schedule_queued_jobs()
        return JOB_STORE.get_job(job_id) or released
    return released


def _start_job_scheduler() -> None:
    JOB_QUEUE.start()


def _terminate_worker_process(pid: int) -> None:
    terminate_worker_process(pid)


def _cancel_job(job_id: str) -> dict[str, Any] | None:
    return JOB_QUEUE.cancel_job(job_id)


def _can_access_job(
    job_record: dict[str, Any], user_email: str, include_all: bool = False
) -> bool:
    if include_all:
        return True
    if bool(job_record.get("shared_with_all")):
        return True
    return _normalize_email(job_record.get("owner_email")) == _normalize_email(
        user_email
    )


def _strip_api_prefix(path: str) -> str:
    if path == "/api":
        return "/"
    if path.startswith("/api/"):
        return path[4:]
    return path


def _path_is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def _job_result_scenario(result: dict[str, Any], scenario_key: str) -> dict[str, Any]:
    structured = dict(result.get("structured_results") or {})
    scenario = dict(structured.get(scenario_key) or {})
    if scenario_key != "original":
        return scenario

    recovered = (
        dict(result.get("free_optimization_baseline") or {})
        or dict(structured.get("free_optimization_baseline") or {})
    )
    if not recovered:
        return scenario

    merged = dict(scenario)
    merged.update(recovered)
    if not merged.get("output_html") and scenario.get("output_html"):
        merged["output_html"] = scenario.get("output_html")
    return merged


def _format_time_impact_limit_minutes(value: Any) -> str:
    try:
        numeric = max(0.0, float(value))
    except (TypeError, ValueError):
        numeric = float(TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.1f}".rstrip("0").rstrip(".")


def _job_time_impact_limit_minutes(
    job_record: dict[str, Any],
    result: dict[str, Any] | None = None,
    scenario_key: str = "",
) -> float:
    result = dict(result if result is not None else job_record.get("result") or {})
    scenario = _job_result_scenario(result, scenario_key) if scenario_key else {}
    time_constraint = dict(scenario.get("time_constraint") or {})
    config = _job_planner_config_payload(job_record)
    for value in (
        scenario.get("time_impact_limit_minutes"),
        time_constraint.get("time_impact_limit_minutes"),
        time_constraint.get("threshold_minutes"),
        config.get("time_impact_limit_minutes"),
    ):
        try:
            return max(0.0, float(value))
        except (TypeError, ValueError):
            continue
    return float(TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)


def _job_map_scenario_label(
    job_record: dict[str, Any],
    result: dict[str, Any],
    scenario_key: str,
) -> str:
    scenario = _job_result_scenario(result, scenario_key)
    for key in ("display_name", "scenario_label"):
        label = str(scenario.get(key) or "").strip()
        if label:
            return label
    if scenario_key == "time_constrained":
        return f"{_format_time_impact_limit_minutes(_job_time_impact_limit_minutes(job_record, result, scenario_key))}-Minute Constrained"
    if scenario_key == "ep15min":
        return f"EP {_format_time_impact_limit_minutes(_job_time_impact_limit_minutes(job_record, result, scenario_key))}-Minute"
    return MAP_SCENARIO_LABELS.get(scenario_key, scenario_key)


def _resolve_job_map_artifact(
    job_record: dict[str, Any], artifact_key: str
) -> tuple[Path | None, str | None]:
    scenario_key = MAP_ARTIFACT_KEYS.get(artifact_key.strip().lower())
    if not scenario_key:
        return None, f"Unknown artifact: {artifact_key}"

    result = dict(job_record.get("result") or {})
    structured = dict(result.get("structured_results") or {})
    scenario = _job_result_scenario(result, scenario_key)
    output_paths = dict(structured.get("output_paths") or {})
    candidate_paths = [
        scenario.get("output_html"),
        output_paths.get(scenario_key),
        result.get(MAP_ARTIFACT_TOP_LEVEL_KEYS.get(scenario_key, "")),
    ]
    outputs_root = (BASE_DIR / "outputs").resolve()

    for raw_path in candidate_paths:
        if not raw_path:
            continue
        artifact_path = Path(str(raw_path)).expanduser().resolve()
        if artifact_path.suffix.lower() != ".html":
            return None, "Map artifact is not an HTML file."
        if not _path_is_relative_to(artifact_path, outputs_root):
            return None, "Map artifact is outside the backend outputs directory."
        if not artifact_path.exists():
            return None, f"Map artifact file is missing: {artifact_path.name}"
        return artifact_path, None

    return None, f"Map artifact is not available: {artifact_key}"


def _float_or_none(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isfinite(number):
        return number
    return None


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _map_point_coordinates(point: dict[str, Any]) -> tuple[float, float] | None:
    lat = _float_or_none(
        point.get("plot_lat") if point.get("plot_lat") is not None else point.get("lat")
    )
    lng = _float_or_none(
        point.get("plot_lng") if point.get("plot_lng") is not None else point.get("lng")
    )
    if lat is None or lng is None:
        return None
    return lat, lng


def _map_bounds_from_coordinates(
    coordinates: list[tuple[float, float]],
) -> dict[str, float] | None:
    if not coordinates:
        return None
    lats = [lat for lat, _lng in coordinates]
    lngs = [lng for _lat, lng in coordinates]
    return {
        "min_lng": min(lngs),
        "min_lat": min(lats),
        "max_lng": max(lngs),
        "max_lat": max(lats),
    }


def _route_geometry_coordinates(route: dict[str, Any]) -> list[list[float]]:
    coordinates: list[list[float]] = []
    for leg_detail in list(route.get("leg_details") or []):
        geometry = list(leg_detail.get("geometry") or [])
        for index, raw_pair in enumerate(geometry):
            if not isinstance(raw_pair, (list, tuple)) or len(raw_pair) < 2:
                continue
            lat = _float_or_none(raw_pair[0])
            lng = _float_or_none(raw_pair[1])
            if lat is None or lng is None:
                continue
            if coordinates and index == 0 and coordinates[-1] == [lng, lat]:
                continue
            coordinates.append([lng, lat])
    return coordinates


def _route_connector_coordinates(
    route: dict[str, Any],
    route_id: str,
    route_index: int,
) -> list[dict[str, Any]]:
    connectors: list[dict[str, Any]] = []
    for leg_index, leg_detail in enumerate(list(route.get("leg_details") or [])):
        for connector_index, raw_connector in enumerate(list(dict(leg_detail).get("snap_connectors") or [])):
            connector = dict(raw_connector or {})
            geometry: list[list[float]] = []
            for raw_pair in list(connector.get("geometry") or []):
                if not isinstance(raw_pair, (list, tuple)) or len(raw_pair) < 2:
                    continue
                lat = _float_or_none(raw_pair[0])
                lng = _float_or_none(raw_pair[1])
                if lat is None or lng is None:
                    continue
                geometry.append([lng, lat])
            if len(geometry) < 2:
                continue
            connectors.append(
                {
                    "id": f"{route_id}:connector:{leg_index}:{connector_index}",
                    "route_id": route_id,
                    "route_index": route_index,
                    "from_node": _int_or_none(dict(leg_detail).get("from_node")),
                    "to_node": _int_or_none(dict(leg_detail).get("to_node")),
                    "connector_type": str(connector.get("type") or "snap").strip() or "snap",
                    "distance_m": float(connector.get("distance_m", 0.0) or 0.0),
                    "geometry": geometry,
                }
            )
    return connectors


def _china_coordinate_out_of_bounds(lat: float, lng: float) -> bool:
    return not (73.66 < lng < 135.05 and 3.86 < lat < 53.55)


def _transform_lat(x: float, y: float) -> float:
    ret = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
    ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(y * math.pi) + 40.0 * math.sin(y / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(y / 12.0 * math.pi) + 320 * math.sin(y * math.pi / 30.0)) * 2.0 / 3.0
    return ret


def _transform_lng(x: float, y: float) -> float:
    ret = 300.0 + x + 2.0 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * math.sqrt(abs(x))
    ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(x * math.pi) + 40.0 * math.sin(x / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(x / 12.0 * math.pi) + 300.0 * math.sin(x / 30.0 * math.pi)) * 2.0 / 3.0
    return ret


def _gcj02_to_wgs84(lat: float, lng: float) -> tuple[float, float]:
    if _china_coordinate_out_of_bounds(lat, lng):
        return lat, lng
    a = 6378245.0
    ee = 0.00669342162296594323
    dlat = _transform_lat(lng - 105.0, lat - 35.0)
    dlng = _transform_lng(lng - 105.0, lat - 35.0)
    radlat = math.radians(lat)
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrt_magic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrt_magic) * math.pi)
    dlng = (dlng * 180.0) / (a / sqrt_magic * math.cos(radlat) * math.pi)
    mg_lat = lat + dlat
    mg_lng = lng + dlng
    return lat * 2 - mg_lat, lng * 2 - mg_lng


def _wgs84_to_gcj02(lat: float, lng: float) -> tuple[float, float]:
    if _china_coordinate_out_of_bounds(lat, lng):
        return lat, lng
    a = 6378245.0
    ee = 0.00669342162296594323
    dlat = _transform_lat(lng - 105.0, lat - 35.0)
    dlng = _transform_lng(lng - 105.0, lat - 35.0)
    radlat = math.radians(lat)
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrt_magic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrt_magic) * math.pi)
    dlng = (dlng * 180.0) / (a / sqrt_magic * math.cos(radlat) * math.pi)
    return lat + dlat, lng + dlng


def _compact_location_text(value: Any) -> str:
    return "".join(str(value or "").strip().lower().split())


def _mapping_text_for_keys(mapping: dict[str, Any], keys: tuple[str, ...]) -> list[str]:
    values: list[str] = []
    for key in keys:
        value = mapping.get(key)
        if value is not None and str(value).strip():
            values.append(str(value).strip())
    return values


def _point_indicates_china(point: dict[str, Any]) -> bool:
    provider = str(point.get("provider") or point.get("geocode_provider") or "").strip().lower()
    if provider == "amap":
        return True
    adcode = str(point.get("adcode") or "").strip()
    if adcode and adcode[:2].isdigit():
        return True
    marker_text = _compact_location_text(
        " ".join(
            str(point.get(key) or "")
            for key in (
                "country",
                "city",
                "district",
                "address",
                "display_address",
                "requested_address",
            )
        )
    )
    return any(
        marker in marker_text
        for marker in ("china", "中国", "上海", "北京", "苏州", "西安", "shanghai", "beijing", "suzhou")
    )


def _should_use_amap_display_geometry(
    job_record: dict[str, Any],
    result: dict[str, Any],
    structured: dict[str, Any],
    points: list[Any],
) -> bool:
    if not AMAP_DISPLAY_GEOMETRY_ENABLED:
        return False
    if not _amap_display_api_key():
        return False

    config = _job_planner_config_payload(job_record)
    country_keys = (
        "country",
        "market_country",
        "school_country",
        "traffic_country",
        "default_country",
    )
    city_keys = (
        "city",
        "market_city",
        "school_city",
        "traffic_city",
        "default_city",
    )
    country_texts: list[str] = []
    city_texts: list[str] = []
    for mapping in (config, result, structured):
        country_texts.extend(_mapping_text_for_keys(dict(mapping or {}), country_keys))
        city_texts.extend(_mapping_text_for_keys(dict(mapping or {}), city_keys))

    country_blob = _compact_location_text(" ".join(country_texts))
    if any(marker in country_blob for marker in ("southkorea", "korea", "kr", "대한민국", "한국", "thailand", "thai", "th", "태국")):
        return False
    if any(marker in country_blob for marker in ("china", "cn", "中国", "中华人民共和国")):
        return True

    city_blob = _compact_location_text(" ".join(city_texts))
    if any(marker in city_blob for marker in ("shanghai", "上海", "beijing", "北京", "suzhou", "苏州", "xian", "西安")):
        return True

    return any(_point_indicates_china(dict(point or {})) for point in points)


def _amap_request_coordinates_for_point(point: dict[str, Any]) -> tuple[float, float] | None:
    provider = str(point.get("provider") or point.get("geocode_provider") or "").strip().lower()
    raw_lat = _float_or_none(point.get("lat"))
    raw_lng = _float_or_none(point.get("lng"))
    if raw_lat is not None and raw_lng is not None:
        if provider == "amap" or str(point.get("adcode") or "").strip():
            return raw_lat, raw_lng
    plot_coords = _map_point_coordinates(point)
    if plot_coords:
        return _wgs84_to_gcj02(plot_coords[0], plot_coords[1])
    if raw_lat is not None and raw_lng is not None:
        return _wgs84_to_gcj02(raw_lat, raw_lng)
    return None


def _load_amap_display_cache_unlocked() -> dict[str, Any]:
    return load_json_object(AMAP_DISPLAY_GEOMETRY_CACHE_PATH)


def _save_amap_display_cache_unlocked(cache: dict[str, Any]) -> None:
    save_json_object(AMAP_DISPLAY_GEOMETRY_CACHE_PATH, cache, indent=None)


def _dedupe_line_coordinates(coordinates: list[list[float]]) -> list[list[float]]:
    deduped: list[list[float]] = []
    for coordinate in coordinates:
        if len(coordinate) < 2:
            continue
        lng = _float_or_none(coordinate[0])
        lat = _float_or_none(coordinate[1])
        if lng is None or lat is None:
            continue
        candidate = [round(lng, 7), round(lat, 7)]
        if not deduped or deduped[-1] != candidate:
            deduped.append(candidate)
    return deduped


def _amap_display_cache_key(
    request_points: list[tuple[float, float]]
) -> str:
    encoded = "|".join(
        f"{lng:.6f},{lat:.6f}" for lat, lng in request_points
    )
    return f"{AMAP_DISPLAY_GEOMETRY_VERSION}|{encoded}"


def _decode_amap_polyline(polyline: str) -> list[list[float]]:
    coordinates: list[list[float]] = []
    for chunk in str(polyline or "").split(";"):
        text = chunk.strip()
        if not text or "," not in text:
            continue
        lng_text, lat_text = text.split(",", 1)
        lat = _float_or_none(lat_text)
        lng = _float_or_none(lng_text)
        if lat is None or lng is None:
            continue
        wgs_lat, wgs_lng = _gcj02_to_wgs84(lat, lng)
        coordinates.append([wgs_lng, wgs_lat])
    return _dedupe_line_coordinates(coordinates)


def _throttle_amap_display_request() -> None:
    global _AMAP_DISPLAY_LAST_REQUEST_AT
    if AMAP_DISPLAY_GEOMETRY_REQUEST_INTERVAL_S <= 0:
        return
    with _AMAP_DISPLAY_REQUEST_LOCK:
        now = time.monotonic()
        wait_s = AMAP_DISPLAY_GEOMETRY_REQUEST_INTERVAL_S - (
            now - _AMAP_DISPLAY_LAST_REQUEST_AT
        )
        if wait_s > 0:
            time.sleep(wait_s)
        _AMAP_DISPLAY_LAST_REQUEST_AT = time.monotonic()


def _fetch_amap_display_segment(
    request_points: list[tuple[float, float]]
) -> dict[str, Any]:
    if len(request_points) < 2:
        return {"geometry": [], "duration_s": None, "distance_m": None}
    amap_key = _amap_display_api_key()
    if not amap_key:
        return {"geometry": [], "duration_s": None, "distance_m": None}
    origin_lat, origin_lng = request_points[0]
    dest_lat, dest_lng = request_points[-1]
    params: dict[str, str] = {
        "key": amap_key,
        "origin": f"{origin_lng:.6f},{origin_lat:.6f}",
        "destination": f"{dest_lng:.6f},{dest_lat:.6f}",
        "extensions": "base",
        "output": "json",
    }
    waypoint_values = [
        f"{lng:.6f},{lat:.6f}" for lat, lng in request_points[1:-1]
    ]
    if waypoint_values:
        params["waypoints"] = ";".join(waypoint_values)

    _throttle_amap_display_request()
    url = f"https://restapi.amap.com/v3/direction/driving?{urlencode(params)}"
    request = Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "BRP Bus Route Planner display-geometry",
        },
    )
    with urlopen(request, timeout=12) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if str(payload.get("status") or "") != "1":
        message = str(payload.get("info") or payload.get("infocode") or "unknown error")
        raise RuntimeError(f"AMap display route failed: {message}")
    paths = list(dict(payload.get("route") or {}).get("paths") or [])
    if not paths:
        return {"geometry": [], "duration_s": None, "distance_m": None}
    path = dict(paths[0] or {})
    coordinates: list[list[float]] = []
    for step in list(path.get("steps") or []):
        coordinates.extend(_decode_amap_polyline(str(dict(step or {}).get("polyline") or "")))
    if not coordinates:
        coordinates = _decode_amap_polyline(str(path.get("polyline") or ""))
    duration_s = _float_or_none(path.get("duration"))
    distance_m = _float_or_none(path.get("distance"))
    return {
        "geometry": _dedupe_line_coordinates(coordinates),
        "duration_s": duration_s,
        "distance_m": distance_m,
    }


def _fetch_amap_display_geometry(
    request_points: list[tuple[float, float]]
) -> dict[str, Any]:
    max_points_per_request = max(2, AMAP_DISPLAY_GEOMETRY_MAX_WAYPOINTS + 2)
    if len(request_points) <= max_points_per_request:
        return _fetch_amap_display_segment(request_points)

    geometry: list[list[float]] = []
    duration_s = 0.0
    distance_m = 0.0
    has_duration = False
    has_distance = False
    start_index = 0
    while start_index < len(request_points) - 1:
        end_index = min(len(request_points), start_index + max_points_per_request)
        segment = _fetch_amap_display_segment(request_points[start_index:end_index])
        geometry.extend(list(segment.get("geometry") or []))
        if segment.get("duration_s") is not None:
            duration_s += float(segment.get("duration_s") or 0.0)
            has_duration = True
        if segment.get("distance_m") is not None:
            distance_m += float(segment.get("distance_m") or 0.0)
            has_distance = True
        start_index = end_index - 1
    return {
        "geometry": _dedupe_line_coordinates(geometry),
        "duration_s": duration_s if has_duration else None,
        "distance_m": distance_m if has_distance else None,
    }


def _amap_display_geometry_for_route(
    points: list[Any],
    nodes: list[Any],
) -> tuple[list[list[float]] | None, str, str, float | None, float | None]:
    route_points: list[dict[str, Any]] = []
    for node in nodes:
        node_index = _int_or_none(node)
        if node_index is None or node_index < 0 or node_index >= len(points):
            continue
        point = dict(points[node_index] or {})
        if _amap_request_coordinates_for_point(point):
            route_points.append(point)
    request_points = [
        coords
        for point in route_points
        if (coords := _amap_request_coordinates_for_point(point)) is not None
    ]
    if len(request_points) < 2:
        return None, "osrm", "", None, None

    cache_key = _amap_display_cache_key(request_points)
    stale_cached_geometry: list[list[float]] | None = None
    with _AMAP_DISPLAY_CACHE_LOCK:
        cache = _load_amap_display_cache_unlocked()
        cached = dict(cache.get(cache_key) or {})
        cached_geometry = cached.get("geometry")
        if isinstance(cached_geometry, list) and len(cached_geometry) >= 2:
            cached_duration_s = _float_or_none(cached.get("duration_s"))
            cached_distance_m = _float_or_none(cached.get("distance_m"))
            if cached_duration_s is None:
                stale_cached_geometry = cached_geometry
            else:
                return (
                    cached_geometry,
                    "amap_cn_cache",
                    "",
                    cached_duration_s,
                    cached_distance_m,
                )

    def stale_cache_result(
        message: str,
    ) -> tuple[list[list[float]], str, str, float | None, float | None] | None:
        if not stale_cached_geometry:
            return None
        return stale_cached_geometry, "amap_cn_cache_stale", message, None, None

    try:
        response = _fetch_amap_display_geometry(request_points)
    except (HTTPError, URLError, TimeoutError, OSError, RuntimeError, json.JSONDecodeError) as exc:
        fallback = stale_cache_result(str(exc))
        if fallback:
            return fallback
        return None, "osrm", str(exc), None, None
    geometry = list(response.get("geometry") or [])
    if len(geometry) < 2:
        fallback = stale_cache_result("AMap display route returned no geometry")
        if fallback:
            return fallback
        return None, "osrm", "AMap display route returned no geometry", None, None

    with _AMAP_DISPLAY_CACHE_LOCK:
        cache = _load_amap_display_cache_unlocked()
        cache[cache_key] = {
            "created_at": utc_now_iso(),
            "point_count": len(request_points),
            "geometry": geometry,
            "duration_s": response.get("duration_s"),
            "distance_m": response.get("distance_m"),
        }
        _save_amap_display_cache_unlocked(cache)
    return (
        geometry,
        "amap_cn",
        "",
        _float_or_none(response.get("duration_s")),
        _float_or_none(response.get("distance_m")),
    )


DEFAULT_TO_SCHOOL_ARRIVAL_MINUTES = 8 * 60
DEFAULT_FROM_SCHOOL_DEPARTURE_MINUTES = 15 * 60 + 40
AM_EARLIEST_DEPARTURE_MINUTES = 6 * 60
AM_LATEST_ARRIVAL_MINUTES = 8 * 60
AM_ARRIVAL_GATE_GRACE_MINUTES = float(
    os.environ.get("BRP_AM_ARRIVAL_GATE_GRACE_MINUTES", "0") or 0
)


def _job_planner_config_payload(job_record: dict[str, Any]) -> dict[str, Any]:
    result = dict(job_record.get("result") or {})
    metadata = dict(job_record.get("metadata") or {})
    return dict(result.get("planner_config") or job_record.get("config") or metadata.get("planner_config") or {})


def _parse_clock_minutes(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    parts = text.split(":")
    if len(parts) < 2:
        return None
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
    except ValueError:
        return None
    if hours < 0 or minutes < 0 or minutes > 59:
        return None
    return (hours * 60 + minutes) % (24 * 60)


def _format_clock_minutes(minutes: float | int | None) -> str:
    if minutes is None:
        return ""
    total_minutes = int(round(float(minutes))) % (24 * 60)
    hours = total_minutes // 60
    minute = total_minutes % 60
    return f"{hours:02d}:{minute:02d}"


def _schedule_anchor_minutes(
    job_record: dict[str, Any], service_direction: str
) -> tuple[int, str, str]:
    config = _job_planner_config_payload(job_record)
    if service_direction == "To School":
        candidate_keys = (
            "to_school_arrival_time",
            "school_arrival_time",
            "target_arrival_time",
            "arrival_time",
        )
        default_minutes = DEFAULT_TO_SCHOOL_ARRIVAL_MINUTES
        label = "School arrival"
    else:
        candidate_keys = (
            "from_school_departure_time",
            "school_departure_time",
            "target_departure_time",
            "departure_time",
        )
        default_minutes = DEFAULT_FROM_SCHOOL_DEPARTURE_MINUTES
        label = "School departure"
    for key in candidate_keys:
        parsed = _parse_clock_minutes(config.get(key))
        if parsed is not None:
            return parsed, _format_clock_minutes(parsed), label
    return default_minutes, _format_clock_minutes(default_minutes), label


def _to_school_time_window_minutes(job_record: dict[str, Any]) -> tuple[int, int]:
    target_minutes, _target_label, _target_kind = _schedule_anchor_minutes(job_record, "To School")
    return AM_EARLIEST_DEPARTURE_MINUTES, min(target_minutes, AM_LATEST_ARRIVAL_MINUTES)


def _map_route_duration_scale(route: dict[str, Any], stops: list[dict[str, Any]]) -> tuple[float, float]:
    raw_duration_s = float(route.get("raw_duration_s", 0.0) or 0.0)
    display_duration_s = float(route.get("duration_s", 0.0) or 0.0)
    max_cumulative_s = max(
        [float(stop.get("cumulative_duration_s", 0.0) or 0.0) for stop in stops] or [0.0]
    )
    base_duration_s = raw_duration_s if raw_duration_s > 0 else max_cumulative_s
    if display_duration_s <= 0:
        display_duration_s = base_duration_s
    scale = (display_duration_s / base_duration_s) if base_duration_s > 0 else 1.0
    return max(display_duration_s, max_cumulative_s * scale), scale


def _apply_schedule_times(payload: dict[str, Any], job_record: dict[str, Any]) -> None:
    service_direction = (
        "To School"
        if str(payload.get("service_direction") or "").strip() == "To School"
        else "From School"
    )
    anchor_minutes, anchor_label, anchor_kind = _schedule_anchor_minutes(
        job_record, service_direction
    )
    config = _job_planner_config_payload(job_record)
    try:
        dwell_seconds = max(0.0, float(config.get("stop_service_minutes", 1) or 1) * 60.0)
    except (TypeError, ValueError):
        dwell_seconds = 60.0

    routes_by_id = {str(route.get("id") or ""): route for route in list(payload.get("routes") or [])}
    stops_by_route: dict[str, list[dict[str, Any]]] = {}
    for stop in list(payload.get("stops") or []):
        stops_by_route.setdefault(str(stop.get("route_id") or ""), []).append(stop)

    for route_id, route_stops in stops_by_route.items():
        route_stops.sort(key=lambda item: int(item.get("order", 0) or 0))
        route = routes_by_id.get(route_id, {})
        route_duration_s, scale = _map_route_duration_scale(route, route_stops)
        service_orders = sorted(
            int(stop.get("order", 0) or 0)
            for stop in route_stops
            if not bool(stop.get("is_depot"))
        )
        service_order_count = len(service_orders)
        for stop in route_stops:
            order = int(stop.get("order", 0) or 0)
            drive_elapsed_s = float(stop.get("cumulative_duration_s", 0.0) or 0.0) * scale
            if service_direction == "To School":
                if bool(stop.get("is_depot")):
                    offset_s = 0.0
                else:
                    downstream_dwell_count = len([item for item in service_orders if item >= order])
                    remaining_drive_s = max(0.0, route_duration_s - drive_elapsed_s)
                    offset_s = -(remaining_drive_s + downstream_dwell_count * dwell_seconds)
            else:
                prior_dwell_count = len([item for item in service_orders if item < order])
                if service_order_count == 0 or bool(stop.get("is_depot")):
                    prior_dwell_count = 0
                offset_s = drive_elapsed_s + prior_dwell_count * dwell_seconds
            scheduled_minutes = anchor_minutes + (offset_s / 60.0)
            stop["schedule_anchor_label"] = anchor_label
            stop["schedule_anchor_kind"] = anchor_kind
            stop["scheduled_offset_s"] = offset_s
            stop["scheduled_time_minutes"] = scheduled_minutes
            stop["scheduled_time_label"] = _format_clock_minutes(scheduled_minutes)


def _attach_am_arrival_gate(payload: dict[str, Any], job_record: dict[str, Any]) -> None:
    if str(payload.get("service_direction") or "").strip() != "To School":
        return
    target_minutes, target_label, _target_kind = _schedule_anchor_minutes(job_record, "To School")
    earliest_departure_minutes, latest_arrival_minutes = _to_school_time_window_minutes(job_record)
    grace_s = max(0.0, AM_ARRIVAL_GATE_GRACE_MINUTES * 60.0)
    checked = 0
    failed = 0
    unavailable = 0
    max_overrun_s = 0.0
    for route in list(payload.get("routes") or []):
        planned_drive_s = _float_or_none(route.get("duration_s"))
        verified_drive_s = planned_drive_s
        stop_service_s = _float_or_none(route.get("stop_service_time_s")) or 0.0
        source = str(route.get("traffic_time_source") or "route_timing")
        gate: dict[str, Any] = {
            "target_arrival_minutes": target_minutes,
            "target_arrival_label": target_label,
            "earliest_departure_minutes": earliest_departure_minutes,
            "earliest_departure_label": _format_clock_minutes(earliest_departure_minutes),
            "latest_arrival_minutes": latest_arrival_minutes,
            "latest_arrival_label": _format_clock_minutes(latest_arrival_minutes),
            "planned_drive_duration_s": planned_drive_s,
            "verified_drive_duration_s": verified_drive_s,
            "verified_stop_service_s": stop_service_s,
            "verified_source": source,
            "grace_minutes": AM_ARRIVAL_GATE_GRACE_MINUTES,
        }
        if planned_drive_s is None or verified_drive_s is None:
            unavailable += 1
            gate.update({"status": "unavailable", "passes": None})
        else:
            checked += 1
            verified_total_s = verified_drive_s + stop_service_s
            latest_departure_minutes = latest_arrival_minutes - (verified_total_s / 60.0)
            overrun_s = max(0.0, (earliest_departure_minutes - latest_departure_minutes) * 60.0)
            scheduled_departure_minutes = max(earliest_departure_minutes, latest_departure_minutes)
            scheduled_arrival_minutes = scheduled_departure_minutes + (verified_total_s / 60.0)
            max_overrun_s = max(max_overrun_s, overrun_s)
            passes = overrun_s <= grace_s
            if not passes:
                failed += 1
            gate.update(
                {
                    "status": "passed" if passes else "failed",
                    "passes": passes,
                    "verified_total_duration_s": verified_total_s,
                    "estimated_arrival_delay_s": overrun_s,
                    "estimated_arrival_delay_minutes": overrun_s / 60.0,
                    "time_window_overrun_s": overrun_s,
                    "time_window_overrun_minutes": overrun_s / 60.0,
                    "verified_departure_minutes": scheduled_departure_minutes,
                    "verified_departure_label": _format_clock_minutes(scheduled_departure_minutes),
                    "verified_arrival_minutes": scheduled_arrival_minutes,
                    "verified_arrival_label": _format_clock_minutes(scheduled_arrival_minutes),
                }
            )
        route["am_arrival_gate"] = gate

    status = "unavailable"
    if failed:
        status = "failed"
    elif unavailable:
        status = "unavailable"
    elif checked:
        status = "passed"
    payload.setdefault("summary", {})["am_arrival_gate"] = {
        "enabled": True,
        "status": status,
        "target_arrival_label": target_label,
        "checked_route_count": checked,
        "failed_route_count": failed,
        "unavailable_route_count": unavailable,
        "max_estimated_arrival_delay_minutes": max_overrun_s / 60.0,
        "max_time_window_overrun_minutes": max_overrun_s / 60.0,
        "grace_minutes": AM_ARRIVAL_GATE_GRACE_MINUTES,
    }


def _normalize_stop_match_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _stop_match_keys(stop: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    node_index = _int_or_none(stop.get("node_index"))
    if node_index is not None:
        keys.append(f"node:{node_index}")
    demand_batch_index = _int_or_none(stop.get("demand_batch_index"))
    if demand_batch_index is not None:
        keys.append(f"batch:{demand_batch_index}")
    for field in ("requested_address", "address"):
        normalized = _normalize_stop_match_text(stop.get(field))
        if normalized:
            keys.append(f"addr:{normalized}")
    return keys


def _time_impact_level(adverse_delta_minutes: float) -> str:
    if adverse_delta_minutes <= 0.5:
        return "better"
    if adverse_delta_minutes <= 10:
        return "acceptable"
    if adverse_delta_minutes <= 20:
        return "notice"
    if adverse_delta_minutes <= 30:
        return "elevated"
    if adverse_delta_minutes <= 45:
        return "severe"
    return "critical"


def _time_impact_direction(delta_minutes: float, adverse_delta_minutes: float) -> str:
    if adverse_delta_minutes > 0.5:
        return "worse"
    if abs(delta_minutes) > 0.5:
        return "better"
    return "neutral"


def _time_impact_change_direction(delta_minutes: float) -> str:
    if delta_minutes < -0.5:
        return "earlier"
    if delta_minutes > 0.5:
        return "later"
    return "same"


def _time_impact_passenger_count(stop: dict[str, Any]) -> int:
    impact = dict(stop.get("time_impact") or {})
    raw_count = impact.get("affected_rider_count", stop.get("passenger_count", 0))
    try:
        return max(0, int(raw_count or 0))
    except (TypeError, ValueError):
        return 0


def _time_impact_top_stops(
    compared: list[dict[str, Any]], *, limit: int = 5
) -> list[dict[str, Any]]:
    ranked = sorted(
        compared,
        key=lambda stop: (
            float(dict(stop.get("time_impact") or {}).get("adverse_delta_minutes", 0.0) or 0.0),
            _time_impact_passenger_count(stop),
            abs(float(dict(stop.get("time_impact") or {}).get("delta_minutes", 0.0) or 0.0)),
        ),
        reverse=True,
    )
    items: list[dict[str, Any]] = []
    for stop in ranked[:limit]:
        impact = dict(stop.get("time_impact") or {})
        items.append(
            {
                "stop_id": str(stop.get("id") or ""),
                "address": str(stop.get("address") or ""),
                "route_id": str(stop.get("route_id") or ""),
                "current_route_id": str(impact.get("current_route_id") or ""),
                "new_route_id": str(impact.get("new_route_id") or ""),
                "current_time_label": str(impact.get("current_time_label") or ""),
                "new_time_label": str(impact.get("new_time_label") or ""),
                "delta_minutes": float(impact.get("delta_minutes", 0.0) or 0.0),
                "adverse_delta_minutes": float(
                    impact.get("adverse_delta_minutes", 0.0) or 0.0
                ),
                "absolute_delta_minutes": float(
                    impact.get("absolute_delta_minutes", 0.0) or 0.0
                ),
                "acceptance_threshold_minutes": float(
                    impact.get("acceptance_threshold_minutes", TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)
                    or TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES
                ),
                "within_acceptance": bool(impact.get("within_acceptance")),
                "acceptance_status": str(impact.get("acceptance_status") or ""),
                "over_acceptance_minutes": float(
                    impact.get("over_acceptance_minutes", 0.0) or 0.0
                ),
                "affected_rider_count": _time_impact_passenger_count(stop),
                "level": str(impact.get("level") or ""),
                "impact_direction": str(impact.get("impact_direction") or ""),
                "route_changed": bool(impact.get("route_changed")),
            }
        )
    return items


def _time_impact_summary(
    stops: list[dict[str, Any]],
    acceptance_threshold: float | None = None,
) -> dict[str, Any]:
    service_stops = [stop for stop in stops if not bool(stop.get("is_depot"))]
    compared = [
        stop
        for stop in service_stops
        if bool(dict(stop.get("time_impact") or {}).get("comparison_available"))
    ]
    unavailable = [
        stop
        for stop in service_stops
        if dict(stop.get("time_impact") or {}).get("comparison_available") is False
    ]
    compared_rider_count = sum(_time_impact_passenger_count(stop) for stop in compared)
    adverse_values = [
        float(dict(stop.get("time_impact") or {}).get("adverse_delta_minutes", 0.0) or 0.0)
        for stop in compared
    ]
    absolute_values = [
        float(dict(stop.get("time_impact") or {}).get("absolute_delta_minutes", 0.0) or 0.0)
        for stop in compared
    ]
    signed_values = [
        float(dict(stop.get("time_impact") or {}).get("delta_minutes", 0.0) or 0.0)
        for stop in compared
    ]
    adverse_values.sort()
    p90 = 0.0
    if adverse_values:
        p90_index = min(len(adverse_values) - 1, max(0, math.ceil(len(adverse_values) * 0.9) - 1))
        p90 = adverse_values[p90_index]
    levels = [
        str(dict(stop.get("time_impact") or {}).get("level") or "")
        for stop in compared
    ]
    impact_directions = [
        str(dict(stop.get("time_impact") or {}).get("impact_direction") or "")
        for stop in compared
    ]
    total_adverse_rider_minutes = sum(
        float(dict(stop.get("time_impact") or {}).get("adverse_rider_minutes", 0.0) or 0.0)
        for stop in compared
    )
    total_absolute_rider_minutes = sum(
        float(dict(stop.get("time_impact") or {}).get("absolute_rider_minutes", 0.0) or 0.0)
        for stop in compared
    )
    total_benefit_rider_minutes = sum(
        float(dict(stop.get("time_impact") or {}).get("benefit_rider_minutes", 0.0) or 0.0)
        for stop in compared
    )
    high_risk_levels = {"severe", "critical"}
    acceptance_threshold = (
        float(acceptance_threshold)
        if acceptance_threshold is not None
        else float(TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)
    )
    within_acceptance = [
        stop
        for stop in compared
        if float(dict(stop.get("time_impact") or {}).get("adverse_delta_minutes", 0.0) or 0.0)
        <= acceptance_threshold
    ]
    over_acceptance = [
        stop
        for stop in compared
        if float(dict(stop.get("time_impact") or {}).get("adverse_delta_minutes", 0.0) or 0.0)
        > acceptance_threshold
    ]
    within_acceptance_rider_count = sum(
        _time_impact_passenger_count(stop) for stop in within_acceptance
    )
    over_acceptance_rider_count = sum(
        _time_impact_passenger_count(stop) for stop in over_acceptance
    )
    over_acceptance_values = [
        max(
            0.0,
            float(dict(stop.get("time_impact") or {}).get("adverse_delta_minutes", 0.0) or 0.0)
            - acceptance_threshold,
        )
        for stop in over_acceptance
    ]
    return {
        "available": bool(compared),
        "acceptance_threshold_minutes": acceptance_threshold,
        "service_stop_count": len(service_stops),
        "compared_stop_count": len(compared),
        "unavailable_stop_count": len(unavailable),
        "compared_rider_count": compared_rider_count,
        "within_acceptance_stop_count": len(within_acceptance),
        "within_acceptance_rider_count": within_acceptance_rider_count,
        "over_acceptance_stop_count": len(over_acceptance),
        "over_acceptance_rider_count": over_acceptance_rider_count,
        "acceptance_stop_ratio": (
            len(within_acceptance) / len(compared) if compared else 0.0
        ),
        "acceptance_rider_ratio": (
            within_acceptance_rider_count / compared_rider_count
            if compared_rider_count
            else 0.0
        ),
        "max_over_acceptance_delta_minutes": (
            max(over_acceptance_values) if over_acceptance_values else 0.0
        ),
        "avg_adverse_delta_minutes": (
            sum(adverse_values) / len(adverse_values) if adverse_values else 0.0
        ),
        "avg_absolute_delta_minutes": (
            sum(absolute_values) / len(absolute_values) if absolute_values else 0.0
        ),
        "avg_signed_delta_minutes": (
            sum(signed_values) / len(signed_values) if signed_values else 0.0
        ),
        "weighted_avg_adverse_delta_minutes": (
            total_adverse_rider_minutes / compared_rider_count
            if compared_rider_count
            else 0.0
        ),
        "weighted_avg_absolute_delta_minutes": (
            total_absolute_rider_minutes / compared_rider_count
            if compared_rider_count
            else 0.0
        ),
        "p90_adverse_delta_minutes": p90,
        "max_adverse_delta_minutes": adverse_values[-1] if adverse_values else 0.0,
        "max_absolute_delta_minutes": max(absolute_values) if absolute_values else 0.0,
        "notice_stop_count": len([level for level in levels if level == "notice"]),
        "elevated_stop_count": len([level for level in levels if level == "elevated"]),
        "severe_stop_count": len([level for level in levels if level == "severe"]),
        "critical_stop_count": len([level for level in levels if level == "critical"]),
        "high_risk_stop_count": len([level for level in levels if level in high_risk_levels]),
        "worse_stop_count": len(
            [direction for direction in impact_directions if direction == "worse"]
        ),
        "better_stop_count": len(
            [direction for direction in impact_directions if direction == "better"]
        ),
        "neutral_stop_count": len(
            [direction for direction in impact_directions if direction == "neutral"]
        ),
        "worse_rider_count": sum(
            _time_impact_passenger_count(stop)
            for stop in compared
            if str(dict(stop.get("time_impact") or {}).get("impact_direction") or "") == "worse"
        ),
        "better_rider_count": sum(
            _time_impact_passenger_count(stop)
            for stop in compared
            if str(dict(stop.get("time_impact") or {}).get("impact_direction") or "") == "better"
        ),
        "neutral_rider_count": sum(
            _time_impact_passenger_count(stop)
            for stop in compared
            if str(dict(stop.get("time_impact") or {}).get("impact_direction") or "") == "neutral"
        ),
        "high_risk_rider_count": sum(
            _time_impact_passenger_count(stop)
            for stop in compared
            if str(dict(stop.get("time_impact") or {}).get("level") or "") in high_risk_levels
        ),
        "total_adverse_rider_minutes": total_adverse_rider_minutes,
        "total_absolute_rider_minutes": total_absolute_rider_minutes,
        "total_benefit_rider_minutes": total_benefit_rider_minutes,
        "route_changed_stop_count": len(
            [
                stop
                for stop in compared
                if bool(dict(stop.get("time_impact") or {}).get("route_changed"))
            ]
        ),
        "route_changed_rider_count": sum(
            _time_impact_passenger_count(stop)
            for stop in compared
            if bool(dict(stop.get("time_impact") or {}).get("route_changed"))
        ),
        "top_impacted_stops": _time_impact_top_stops(compared),
    }


def _attach_schedule_impact(
    payload: dict[str, Any],
    current_payload: dict[str, Any] | None,
    *,
    acceptance_threshold: float | None = None,
) -> None:
    stops = list(payload.get("stops") or [])
    if not current_payload:
        payload.setdefault("summary", {})["time_impact"] = _time_impact_summary(
            stops,
            acceptance_threshold,
        )
        return

    current_lookup: dict[str, dict[str, Any]] = {}
    for current_stop in list(current_payload.get("stops") or []):
        if bool(current_stop.get("is_depot")):
            continue
        if not str(current_stop.get("scheduled_time_label") or "").strip():
            continue
        for key in _stop_match_keys(current_stop):
            current_lookup.setdefault(key, current_stop)

    service_direction = (
        "To School"
        if str(payload.get("service_direction") or "").strip() == "To School"
        else "From School"
    )
    time_role = "pickup" if service_direction == "To School" else "dropoff"
    for stop in stops:
        if bool(stop.get("is_depot")):
            continue
        current_stop = None
        matched_key = ""
        for key in _stop_match_keys(stop):
            current_stop = current_lookup.get(key)
            if current_stop:
                matched_key = key
                break
        if not current_stop:
            stop["time_impact"] = {
                "comparison_available": False,
                "comparison_status": "current_stop_not_found",
                "time_role": time_role,
                "new_route_id": str(stop.get("route_id") or ""),
                "new_route_index": _int_or_none(stop.get("route_index")),
                "new_stop_order": _int_or_none(stop.get("order")),
                "new_time_minutes": _float_or_none(stop.get("scheduled_time_minutes")),
                "new_time_label": str(stop.get("scheduled_time_label") or ""),
                "affected_rider_count": int(stop.get("passenger_count", 0) or 0),
            }
            continue
        current_minutes = _float_or_none(current_stop.get("scheduled_time_minutes"))
        new_minutes = _float_or_none(stop.get("scheduled_time_minutes"))
        if current_minutes is None or new_minutes is None:
            stop["time_impact"] = {
                "comparison_available": False,
                "comparison_status": "schedule_time_missing",
                "time_role": time_role,
                "current_route_id": str(current_stop.get("route_id") or ""),
                "new_route_id": str(stop.get("route_id") or ""),
                "current_time_label": str(current_stop.get("scheduled_time_label") or ""),
                "new_time_label": str(stop.get("scheduled_time_label") or ""),
                "affected_rider_count": int(stop.get("passenger_count", 0) or 0),
                "matched_key": matched_key,
            }
            continue
        delta_minutes = new_minutes - current_minutes
        if service_direction == "To School":
            adverse_delta_minutes = max(0.0, -delta_minutes)
            adverse_direction = "earlier_pickup"
        else:
            adverse_delta_minutes = max(0.0, delta_minutes)
            adverse_direction = "later_dropoff"
        route_changed = str(current_stop.get("route_id") or "") != str(stop.get("route_id") or "")
        affected_rider_count = max(
            int(current_stop.get("passenger_count", 0) or 0),
            int(stop.get("passenger_count", 0) or 0),
        )
        absolute_delta_minutes = abs(delta_minutes)
        impact_direction = _time_impact_direction(delta_minutes, adverse_delta_minutes)
        benefit_delta_minutes = (
            absolute_delta_minutes if impact_direction == "better" else 0.0
        )
        stop_acceptance_threshold = (
            float(acceptance_threshold)
            if acceptance_threshold is not None
            else float(TIME_IMPACT_ACCEPTANCE_THRESHOLD_MINUTES)
        )
        within_acceptance = adverse_delta_minutes <= stop_acceptance_threshold
        stop["time_impact"] = {
            "comparison_available": True,
            "comparison_status": "matched",
            "matched_key": matched_key,
            "time_role": time_role,
            "current_route_id": str(current_stop.get("route_id") or ""),
            "new_route_id": str(stop.get("route_id") or ""),
            "current_route_index": _int_or_none(current_stop.get("route_index")),
            "new_route_index": _int_or_none(stop.get("route_index")),
            "current_stop_order": _int_or_none(current_stop.get("order")),
            "new_stop_order": _int_or_none(stop.get("order")),
            "current_time_minutes": current_minutes,
            "new_time_minutes": new_minutes,
            "current_time_label": str(current_stop.get("scheduled_time_label") or ""),
            "new_time_label": str(stop.get("scheduled_time_label") or ""),
            "current_offset_s": _float_or_none(current_stop.get("scheduled_offset_s")),
            "new_offset_s": _float_or_none(stop.get("scheduled_offset_s")),
            "delta_minutes": delta_minutes,
            "absolute_delta_minutes": absolute_delta_minutes,
            "adverse_delta_minutes": adverse_delta_minutes,
            "acceptance_threshold_minutes": stop_acceptance_threshold,
            "within_acceptance": within_acceptance,
            "acceptance_status": "within" if within_acceptance else "over",
            "over_acceptance_minutes": max(0.0, adverse_delta_minutes - stop_acceptance_threshold),
            "benefit_delta_minutes": benefit_delta_minutes,
            "adverse_direction": adverse_direction,
            "change_direction": _time_impact_change_direction(delta_minutes),
            "impact_direction": impact_direction,
            "affected_rider_count": affected_rider_count,
            "adverse_rider_minutes": adverse_delta_minutes * affected_rider_count,
            "absolute_rider_minutes": absolute_delta_minutes * affected_rider_count,
            "benefit_rider_minutes": benefit_delta_minutes * affected_rider_count,
            "level": _time_impact_level(adverse_delta_minutes),
            "route_changed": route_changed,
        }

    summary = _time_impact_summary(stops, acceptance_threshold)
    payload.setdefault("summary", {})["time_impact"] = summary

    stops_by_route: dict[str, list[dict[str, Any]]] = {}
    for stop in stops:
        stops_by_route.setdefault(str(stop.get("route_id") or ""), []).append(stop)
    for route in list(payload.get("routes") or []):
        route["time_impact"] = _time_impact_summary(
            stops_by_route.get(str(route.get("id") or ""), []),
            acceptance_threshold,
        )


def _build_job_map_data(
    job_record: dict[str, Any], artifact_key: str
) -> tuple[dict[str, Any] | None, str | None]:
    scenario_key = MAP_ARTIFACT_KEYS.get(artifact_key.strip().lower())
    if not scenario_key:
        return None, f"Unknown map scenario: {artifact_key}"

    return _build_job_map_payload(
        job_record,
        scenario_key,
        artifact_key,
        attach_impact=True,
    )


def _build_job_map_payload(
    job_record: dict[str, Any],
    scenario_key: str,
    artifact_key: str,
    *,
    attach_impact: bool,
) -> tuple[dict[str, Any] | None, str | None]:
    result = dict(job_record.get("result") or {})
    structured = dict(result.get("structured_results") or {})
    scenario = _job_result_scenario(result, scenario_key)
    points = list(scenario.get("points") or [])
    routes = list(scenario.get("routes") or [])
    if not points or not routes:
        return None, f"Map data is not available: {artifact_key}"

    route_payloads: list[dict[str, Any]] = []
    stop_payloads: list[dict[str, Any]] = []
    route_connectors: list[dict[str, Any]] = []
    all_coordinates: list[tuple[float, float]] = []
    use_amap_display_geometry = _should_use_amap_display_geometry(
        job_record, result, structured, points
    )

    for route_index, route in enumerate(routes):
        route_id = str(
            route.get("route_id") or f"Bus {route.get('vehicle_id', route_index + 1)}"
        )
        geometry = _route_geometry_coordinates(dict(route))
        nodes = list(route.get("nodes") or [])
        display_geometry: list[list[float]] | None = None
        display_geometry_source = "osrm"
        display_geometry_message = ""
        display_duration_s: float | None = None
        display_distance_m: float | None = None
        if use_amap_display_geometry:
            (
                display_geometry,
                display_geometry_source,
                display_geometry_message,
                display_duration_s,
                display_distance_m,
            ) = _amap_display_geometry_for_route(points, nodes)

        visible_geometry = display_geometry if display_geometry else geometry
        for lng, lat in visible_geometry:
            all_coordinates.append((lat, lng))

        if not display_geometry:
            connectors = _route_connector_coordinates(dict(route), route_id, route_index)
            route_connectors.extend(connectors)
            for connector in connectors:
                for lng, lat in list(connector.get("geometry") or []):
                    all_coordinates.append((lat, lng))
        traffic_gate = dict(route.get("final_route_traffic_gate") or {})
        arrival_check = dict(route.get("arrival_reverse_check") or {})
        verified_drive_duration_s = _float_or_none(arrival_check.get("verified_drive_duration_s"))
        if verified_drive_duration_s is None:
            verified_drive_duration_s = _float_or_none(traffic_gate.get("verified_drive_duration_s"))
        verified_total_duration_s = _float_or_none(arrival_check.get("verified_total_duration_s"))
        if verified_total_duration_s is None:
            verified_total_duration_s = _float_or_none(traffic_gate.get("verified_total_duration_s"))
        route_duration_s = (
            verified_drive_duration_s
            if verified_drive_duration_s is not None
            else _float_or_none(route.get("traffic_api_duration_s"))
        )
        if route_duration_s is None:
            route_duration_s = (
                _float_or_none(route.get("traffic_adjusted_drive_time_s"))
                or display_duration_s
                or _float_or_none(route.get("time_s"))
                or 0.0
            )
        leg_details = list(route.get("leg_details") or [])
        cumulative_duration_s = 0.0
        cumulative_distance_m = 0.0
        route_stop_ids: list[str] = []
        for order, node in enumerate(nodes):
            node_index = _int_or_none(node)
            if node_index is None or node_index < 0 or node_index >= len(points):
                continue
            point = dict(points[node_index] or {})
            coords = _map_point_coordinates(point)
            if not coords:
                continue
            lat, lng = coords
            if order > 0 and order - 1 < len(leg_details):
                leg = dict(leg_details[order - 1] or {})
                cumulative_duration_s += float(leg.get("duration_s", 0.0) or 0.0)
                cumulative_distance_m += float(leg.get("distance_m", 0.0) or 0.0)
            stop_id = f"{route_id}:{order}:{node_index}"
            route_stop_ids.append(stop_id)
            all_coordinates.append((lat, lng))
            stop_payloads.append(
                {
                    "id": stop_id,
                    "route_id": route_id,
                    "route_index": route_index,
                    "order": order,
                    "node_index": node_index,
                    "address": str(
                        point.get("display_address") or point.get("address") or ""
                    ).strip(),
                    "requested_address": str(
                        point.get("requested_address") or ""
                    ).strip(),
                    "passenger_count": int(point.get("passenger_count", 0) or 0),
                    "is_depot": bool(point.get("is_depot")),
                    "lat": lat,
                    "lng": lng,
                    "cumulative_duration_s": cumulative_duration_s,
                    "cumulative_distance_m": cumulative_distance_m,
                    "demand_batch_index": _int_or_none(point.get("demand_batch_index")),
                    "demand_batch_count": _int_or_none(point.get("demand_batch_count")),
                }
            )
        route_payloads.append(
            {
                "id": route_id,
                "route_index": route_index,
                "vehicle_id": route.get("vehicle_id"),
                "bus_type_name": str(route.get("bus_type_name") or "").strip(),
                "load": int(route.get("load", 0) or 0),
                "bus_capacity": _int_or_none(route.get("bus_capacity")),
                "comfort_capacity": _int_or_none(route.get("comfort_capacity")),
                "stop_count": _int_or_none(route.get("stop_count"))
                or max(0, len(nodes) - 1),
                "max_stops": _int_or_none(route.get("max_stops")),
                "distance_m": float(route.get("distance_m", 0.0) or 0.0),
                "duration_s": float(route_duration_s),
                "raw_duration_s": float(route.get("time_s", 0.0) or 0.0),
                "stop_service_time_s": float(route.get("stop_service_time_s", 0.0) or 0.0),
                "verified_drive_duration_s": verified_drive_duration_s,
                "verified_total_duration_s": verified_total_duration_s,
                "traffic_time_source": str(
                    route.get("traffic_time_source") or ""
                ).strip(),
                "geometry": geometry,
                "display_geometry": display_geometry,
                "display_geometry_source": display_geometry_source,
                "display_geometry_message": display_geometry_message,
                "display_duration_s": display_duration_s,
                "display_distance_m": display_distance_m,
                "stop_ids": route_stop_ids,
            }
        )

    point_by_address = {
        str(point.get("address") or point.get("display_address") or "").strip(): dict(
            point
        )
        for point in points
        if str(point.get("address") or point.get("display_address") or "").strip()
    }
    private_links: list[dict[str, Any]] = []
    for index, item in enumerate(
        list(scenario.get("outlying_private_access_rows") or [])
    ):
        row = dict(item or {})
        geometry: list[list[float]] = []
        for raw_pair in list(row.get("private_drive_geometry") or []):
            if not isinstance(raw_pair, (list, tuple)) or len(raw_pair) < 2:
                continue
            lat = _float_or_none(raw_pair[0])
            lng = _float_or_none(raw_pair[1])
            if lat is None or lng is None:
                continue
            geometry.append([lng, lat])
            all_coordinates.append((lat, lng))
        if len(geometry) < 2:
            stop_point = point_by_address.get(str(row.get("address") or "").strip())
            pickup_point = point_by_address.get(
                str(row.get("pickup_address") or "").strip()
            )
            stop_coords = _map_point_coordinates(stop_point or row)
            pickup_coords = _map_point_coordinates(
                pickup_point
                or {
                    "plot_lat": row.get("pickup_plot_lat"),
                    "plot_lng": row.get("pickup_plot_lng"),
                }
            )
            if stop_coords and pickup_coords:
                geometry = [
                    [stop_coords[1], stop_coords[0]],
                    [pickup_coords[1], pickup_coords[0]],
                ]
                all_coordinates.extend([stop_coords, pickup_coords])
        if geometry:
            private_links.append(
                {
                    "id": f"private-link-{index}",
                    "access_type": str(
                        row.get("private_access_type") or "clustered_rider"
                    ).strip()
                    or "clustered_rider",
                    "address": str(row.get("address") or "").strip(),
                    "pickup_address": str(row.get("pickup_address") or "").strip(),
                    "pickup_route_id": str(row.get("pickup_route_id") or "").strip(),
                    "drive_time_s": float(row.get("private_drive_time_s", 0.0) or 0.0),
                    "drive_distance_m": float(
                        row.get("private_drive_distance_m", 0.0) or 0.0
                    ),
                    "geometry": geometry,
                }
            )

    payload = {
        "job_id": str(job_record.get("job_id") or ""),
        "scenario_key": scenario_key,
        "scenario_name": _job_map_scenario_label(job_record, result, scenario_key),
        "service_direction": str(
            result.get("service_direction") or structured.get("service_direction") or ""
        ).strip(),
        "traffic_profile_name": str(
            result.get("traffic_profile_name")
            or structured.get("traffic_profile_name")
            or ""
        ).strip(),
        "bounds": _map_bounds_from_coordinates(all_coordinates),
        "routes": route_payloads,
        "stops": stop_payloads,
        "route_connectors": route_connectors,
        "private_links": private_links,
        "summary": {
            "route_count": len(route_payloads),
            "stop_count": len(
                [item for item in stop_payloads if not item.get("is_depot")]
            ),
            "passenger_count": sum(int(route.get("load", 0) or 0) for route in routes),
            "distance_m": sum(
                float(route.get("distance_m", 0.0) or 0.0) for route in routes
            ),
            "duration_s": max(
                [float(route.get("time_s", 0.0) or 0.0) for route in routes] or [0.0]
            ),
        },
    }
    _apply_schedule_times(payload, job_record)
    _attach_am_arrival_gate(payload, job_record)
    acceptance_threshold = _job_time_impact_limit_minutes(job_record, result, scenario_key)
    if attach_impact and scenario_key != "current_plan":
        current_payload, _current_error = _build_job_map_payload(
            job_record,
            "current_plan",
            "current_plan",
            attach_impact=False,
        )
        if current_payload:
            _attach_schedule_impact(
                payload,
                current_payload,
                acceptance_threshold=acceptance_threshold,
            )
    else:
        payload.setdefault("summary", {})["time_impact"] = _time_impact_summary(
            list(payload.get("stops") or []),
            acceptance_threshold,
        )
    return payload, None


def _traffic_as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _traffic_as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _traffic_method_counts(estimates: list[Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in estimates:
        method = str(_traffic_as_dict(item).get("method") or "unknown")
        counts[method] = counts.get(method, 0) + 1
    return counts


def _traffic_quality_counts(estimates: list[Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in estimates:
        reason = str(_traffic_as_dict(item).get("quality_reason") or "unknown")
        counts[reason] = counts.get(reason, 0) + 1
    return counts


def _traffic_non_geo_routes(estimates: list[Any], *, limit: int = 12) -> list[dict[str, Any]]:
    routes: list[dict[str, Any]] = []
    for item in estimates:
        estimate = _traffic_as_dict(item)
        method = str(estimate.get("method") or "unknown")
        if method == "geo_route_similarity":
            continue
        routes.append(
            {
                "route_id": str(estimate.get("route_id") or ""),
                "method": method,
                "quality_reason": str(estimate.get("quality_reason") or "unknown"),
                "reason": str(estimate.get("reason") or ""),
                "factor": float(estimate.get("factor") or 0.0),
                "avg_similarity": float(estimate.get("avg_similarity") or 0.0),
                "matched_sample_count": int(estimate.get("matched_sample_count") or 0),
                "geo_candidate_count": int(estimate.get("geo_candidate_count") or 0),
                "usable_geo_candidate_count": int(estimate.get("usable_geo_candidate_count") or 0),
            }
        )
    return routes[:limit]


def _traffic_route_evidence(item: Any, *, include_top_matches: bool = False) -> dict[str, Any]:
    estimate = _traffic_as_dict(item)
    evidence: dict[str, Any] = {
        "route_id": str(estimate.get("route_id") or ""),
        "scenario": str(estimate.get("scenario") or ""),
        "vehicle_id": str(estimate.get("vehicle_id") or ""),
        "bus_type_name": str(estimate.get("bus_type_name") or ""),
        "method": str(estimate.get("method") or "unknown"),
        "quality_reason": str(estimate.get("quality_reason") or "unknown"),
        "factor": float(estimate.get("factor") or 0.0),
        "avg_similarity": float(estimate.get("avg_similarity") or 0.0),
        "matched_sample_count": int(estimate.get("matched_sample_count") or 0),
        "candidate_count": int(estimate.get("candidate_count") or 0),
        "geo_candidate_count": int(estimate.get("geo_candidate_count") or 0),
        "usable_geo_candidate_count": int(estimate.get("usable_geo_candidate_count") or 0),
        "osrm_duration_min": round(float(estimate.get("osrm_duration_s") or 0.0) / 60.0, 2),
        "stop_count": int(estimate.get("stop_count") or 0),
        "fallback": bool(estimate.get("fallback")),
        "reason": str(estimate.get("reason") or ""),
    }
    top_matches = _traffic_as_list(estimate.get("top_matches"))
    if include_top_matches:
        evidence["top_matches"] = [
            {
                "route_id": str(match.get("route_id") or ""),
                "source_id": str(match.get("source_id") or ""),
                "factor": float(match.get("factor") or 0.0),
                "similarity_score": float(match.get("similarity_score") or 0.0),
                "similarity_method": str(match.get("similarity_method") or "unknown"),
                "geo_similarity_score": float(match.get("geo_similarity_score") or 0.0),
                "corridor_overlap": float(match.get("corridor_overlap") or 0.0),
                "center_distance_km": float(match.get("center_distance_km") or 0.0),
                "bearing_score": float(match.get("bearing_score") or 0.0),
                "duration_score": float(match.get("duration_score") or 0.0),
                "stop_score": float(match.get("stop_score") or 0.0),
                "scale_score": float(match.get("scale_score") or 0.0),
            }
            for match in (_traffic_as_dict(match) for match in top_matches)
        ]
    else:
        evidence["top_match_count"] = len(top_matches)
    return evidence


def _traffic_scenario_summary(
    name: str,
    payload: dict[str, Any],
    *,
    include_route_evidence: bool = False,
    include_top_matches: bool = False,
) -> dict[str, Any]:
    estimates = _traffic_as_list(payload.get("route_estimates"))
    method_counts = _traffic_as_dict(payload.get("method_counts")) or _traffic_method_counts(estimates)
    quality_counts = _traffic_as_dict(payload.get("quality_reason_counts")) or _traffic_quality_counts(estimates)
    route_count = int(payload.get("route_count") or len(estimates) or 0)
    geo_count = int(payload.get("geo_attributed_route_count") or method_counts.get("geo_route_similarity", 0) or 0)
    summary: dict[str, Any] = {
        "scenario": name,
        "present": True,
        "route_estimate_count": route_count,
        "geo_attributed_route_count": geo_count,
        "route_similarity_route_count": int(
            payload.get("route_similarity_route_count") or method_counts.get("route_similarity", 0) or 0
        ),
        "fallback_route_count": int(payload.get("fallback_route_count") or method_counts.get("fallback", 0) or 0),
        "non_geo_route_count": max(0, route_count - geo_count),
        "non_geo_routes": _traffic_non_geo_routes(estimates),
        "geo_attributed_route_ratio": (geo_count / route_count) if route_count else 0.0,
        "observed_route_sample_count": int(payload.get("observed_route_sample_count") or 0),
        "geo_route_sample_count": int(payload.get("geo_route_sample_count") or 0),
        "scale_only_route_sample_count": int(payload.get("scale_only_route_sample_count") or 0),
        "geo_route_sample_ratio": float(payload.get("geo_route_sample_ratio") or 0.0),
        "geo_ready": bool(payload.get("geo_ready")),
        "method_counts": method_counts,
        "quality_reason_counts": quality_counts,
    }
    if include_route_evidence:
        summary["route_evidence"] = [
            _traffic_route_evidence(item, include_top_matches=include_top_matches)
            for item in estimates
        ]
    return summary


def _job_traffic_attribution_payload(
    job_record: dict[str, Any],
    *,
    include_route_evidence: bool = False,
    include_top_matches: bool = False,
) -> dict[str, Any]:
    result = _traffic_as_dict(job_record.get("result"))
    structured = _traffic_as_dict(result.get("structured_results"))
    traffic = _traffic_as_dict(structured.get("traffic_attribution") or result.get("traffic_attribution"))
    scenario_estimates = _traffic_as_dict(traffic.get("scenario_route_estimates"))
    scenarios: list[dict[str, Any]] = []
    for name, payload in sorted(scenario_estimates.items()):
        if isinstance(payload, dict):
            scenarios.append(
                _traffic_scenario_summary(
                    str(name),
                    payload,
                    include_route_evidence=include_route_evidence,
                    include_top_matches=include_top_matches,
                )
            )
    return {
        "job_id": str(job_record.get("job_id") or ""),
        "status": str(job_record.get("status") or ""),
        "service_direction": str(structured.get("service_direction") or result.get("service_direction") or ""),
        "traffic_profile_name": str(structured.get("traffic_profile_name") or result.get("traffic_profile_name") or ""),
        "traffic_time_multiplier": float(
            structured.get("traffic_time_multiplier") or result.get("traffic_time_multiplier") or 0.0
        ),
        "traffic_coefficient_mode": str(structured.get("traffic_coefficient_mode") or ""),
        "has_traffic_attribution": bool(traffic),
        "attribution_enabled": bool(traffic.get("enabled")),
        "attribution_succeeded": bool(traffic.get("succeeded")),
        "attribution_mode": str(traffic.get("mode") or ""),
        "attribution_method": str(traffic.get("method") or ""),
        "attribution_reason": str(traffic.get("reason") or ""),
        "attribution_confidence": str(traffic.get("confidence") or ""),
        "route_level_applied": bool(traffic.get("route_level_applied")),
        "observed_route_sample_count": int(traffic.get("observed_route_sample_count") or 0),
        "geo_route_sample_count": int(traffic.get("geo_route_sample_count") or 0),
        "scale_only_route_sample_count": int(traffic.get("scale_only_route_sample_count") or 0),
        "geo_route_sample_ratio": float(traffic.get("geo_route_sample_ratio") or 0.0),
        "scenario_count": len(scenarios),
        "scenarios": scenarios,
    }


def _infer_output_directory_name(result: dict[str, Any]) -> str:
    structured = dict(result.get("structured_results") or {})
    outputs_root = (BASE_DIR / "outputs").resolve()
    for scenario_key in MAP_ARTIFACT_TOP_LEVEL_KEYS:
        scenario = dict(structured.get(scenario_key) or {})
        raw_path = scenario.get("output_html") or result.get(
            MAP_ARTIFACT_TOP_LEVEL_KEYS.get(scenario_key, "")
        )
        if not raw_path:
            continue
        artifact_path = Path(str(raw_path)).expanduser().resolve()
        if _path_is_relative_to(artifact_path, outputs_root):
            return artifact_path.parent.name
    return ""


def _build_rerender_config_for_job(job_record: dict[str, Any]) -> PlannerConfig:
    metadata = dict(job_record.get("metadata") or {})
    config_payload = dict(
        job_record.get("config") or metadata.get("planner_config") or {}
    )
    config = _build_planner_config(config_payload)
    if not config.output_directory_name:
        result = dict(job_record.get("result") or {})
        config.output_directory_name = _infer_output_directory_name(result) or str(
            job_record.get("job_id") or uuid4().hex
        )
    return config


def _rerender_job_map_artifacts(
    job_id: str, job_record: dict[str, Any]
) -> dict[str, Any]:
    result = dict(job_record.get("result") or {})
    structured = dict(result.get("structured_results") or {})
    if not structured:
        return job_record
    hydrated = rerender_html_from_structured_results(
        structured, _build_rerender_config_for_job(job_record)
    )
    updated_result = dict(result)
    updated_result["structured_results"] = hydrated
    output_paths = dict(hydrated.get("output_paths") or {})
    for scenario_key, top_level_key in MAP_ARTIFACT_TOP_LEVEL_KEYS.items():
        if output_paths.get(scenario_key):
            updated_result[top_level_key] = output_paths[scenario_key]
    updated = JOB_STORE.update_job(job_id, result=updated_result)
    if updated:
        return updated
    rerendered = dict(job_record)
    rerendered["result"] = updated_result
    return rerendered


def _build_scenario_template_export(
    job_record: dict[str, Any],
    scenario_key: str,
) -> tuple[bytes | None, str | None]:
    normalized_key = str(scenario_key or "").strip().lower() or "original"
    scenario_key = MAP_ARTIFACT_KEYS.get(normalized_key, normalized_key)
    result = dict(job_record.get("result") or {})
    scenario_label = _job_map_scenario_label(job_record, result, scenario_key)
    scenario = _job_result_scenario(result, scenario_key)
    if not list(scenario.get("routes") or []) or not list(scenario.get("points") or []):
        return None, f"{scenario_label} has no route table to export."
    planner_config = dict(
        result.get("planner_config") or job_record.get("config") or {}
    )
    service_direction = str(
        result.get("service_direction")
        or planner_config.get("service_direction")
        or "From School"
    )
    impact_payload, _impact_error = _build_job_map_payload(
        job_record,
        scenario_key,
        scenario_key,
        attach_impact=True,
    )
    if impact_payload:
        route_gate_by_id = {
            str(route.get("id") or "").strip(): dict(route.get("am_arrival_gate") or {})
            for route in list(impact_payload.get("routes") or [])
        }
        points_with_impact = [
            dict(point) for point in list(scenario.get("points") or [])
        ]
        for stop in list(impact_payload.get("stops") or []):
            node_index = _int_or_none(stop.get("node_index"))
            if node_index is None or node_index < 0 or node_index >= len(points_with_impact):
                continue
            time_impact = dict(stop.get("time_impact") or {})
            if not time_impact:
                time_impact = {"new_time_label": str(stop.get("scheduled_time_label") or "")}
            elif not str(time_impact.get("new_time_label") or "").strip():
                time_impact["new_time_label"] = str(stop.get("scheduled_time_label") or "")
            point = dict(points_with_impact[node_index] or {})
            point["time_impact"] = time_impact
            points_with_impact[node_index] = point
        routes_with_gate = []
        for route_index, route in enumerate(list(scenario.get("routes") or [])):
            route = dict(route or {})
            route_id = str(
                route.get("route_id") or f"Bus {route.get('vehicle_id', route_index + 1)}"
            ).strip()
            route_gate = route_gate_by_id.get(route_id)
            if route_gate:
                route["am_arrival_gate"] = route_gate
            routes_with_gate.append(route)
        scenario = dict(scenario)
        scenario["points"] = points_with_impact
        scenario["routes"] = routes_with_gate
    try:
        return build_baseline_template_workbook_bytes(
            scenario,
            service_direction=service_direction,
            source_label=scenario_label,
        ), None
    except Exception as exc:
        return None, str(exc)


def _build_free_baseline_template_export(
    job_record: dict[str, Any],
) -> tuple[bytes | None, str | None]:
    return _build_scenario_template_export(job_record, "original")


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return json.dumps(value, ensure_ascii=False)


def _append_excel_table(
    sheet: Any,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_safe_value(value) for value in row])
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        header_cell = column_cells[0]
        max_length = max(
            [len(str(cell.value or "")) for cell in column_cells] or [0]
        )
        sheet.column_dimensions[header_cell.column_letter].width = min(
            max(max_length + 2, 10), 42
        )


def _build_time_impact_workbook_export(
    job_record: dict[str, Any],
    scenario_key: str,
) -> tuple[bytes | None, str | None]:
    normalized_key = scenario_key.strip().lower() or "original"
    scenario_key = MAP_ARTIFACT_KEYS.get(normalized_key, normalized_key)
    if scenario_key == "current_plan" or scenario_key not in MAP_SCENARIO_LABELS:
        return None, f"Unknown time impact scenario: {normalized_key}"

    payload, payload_error = _build_job_map_payload(
        job_record,
        scenario_key,
        scenario_key,
        attach_impact=True,
    )
    if payload_error or not payload:
        return None, payload_error or "Time impact data is not available."

    payload_summary = dict(payload.get("summary") or {})
    summary = dict(payload_summary.get("time_impact") or {})
    am_gate = dict(payload_summary.get("am_arrival_gate") or {})
    if not bool(summary.get("available")):
        return None, "Time impact comparison is not available for this scenario."

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_rows = [
        ["Job ID", payload.get("job_id")],
        ["Scenario", payload.get("scenario_name")],
        ["Service direction", payload.get("service_direction")],
        ["Acceptance threshold minutes", summary.get("acceptance_threshold_minutes")],
        ["Compared stops", summary.get("compared_stop_count")],
        ["Compared riders", summary.get("compared_rider_count")],
        ["Within-threshold stops", summary.get("within_acceptance_stop_count")],
        ["Within-threshold riders", summary.get("within_acceptance_rider_count")],
        ["Over-threshold stops", summary.get("over_acceptance_stop_count")],
        ["Over-threshold riders", summary.get("over_acceptance_rider_count")],
        ["Acceptance rider ratio", summary.get("acceptance_rider_ratio")],
        ["Worse stops", summary.get("worse_stop_count")],
        ["Worse riders", summary.get("worse_rider_count")],
        ["High-risk stops", summary.get("high_risk_stop_count")],
        ["High-risk riders", summary.get("high_risk_rider_count")],
        ["Route-changed stops", summary.get("route_changed_stop_count")],
        ["Route-changed riders", summary.get("route_changed_rider_count")],
        ["Weighted average adverse minutes", summary.get("weighted_avg_adverse_delta_minutes")],
        ["P90 adverse minutes", summary.get("p90_adverse_delta_minutes")],
        ["Max adverse minutes", summary.get("max_adverse_delta_minutes")],
        ["Total adverse rider-minutes", summary.get("total_adverse_rider_minutes")],
        ["Total benefit rider-minutes", summary.get("total_benefit_rider_minutes")],
        ["AM window status", am_gate.get("status")],
        ["AM target arrival", am_gate.get("target_arrival_label")],
        ["AM checked routes", am_gate.get("checked_route_count")],
        ["AM failed routes", am_gate.get("failed_route_count")],
        ["AM unavailable routes", am_gate.get("unavailable_route_count")],
        ["AM max overrun minutes", am_gate.get("max_time_window_overrun_minutes")],
    ]
    _append_excel_table(summary_sheet, ["Metric", "Value"], summary_rows)

    route_sheet = wb.create_sheet("Routes")
    route_headers = [
        "Scenario",
        "Route",
        "Bus type",
        "Riders",
        "Stops",
        "Over-threshold riders",
        "Over-threshold stops",
        "Worse riders",
        "High-risk stops",
        "Weighted adverse minutes",
        "Max adverse minutes",
        "Route-changed riders",
        "AM window status",
        "AM departure",
        "AM arrival",
        "AM overrun minutes",
    ]
    route_rows = []
    for route in list(payload.get("routes") or []):
        route_impact = dict(dict(route).get("time_impact") or {})
        route_gate = dict(dict(route).get("am_arrival_gate") or {})
        route_rows.append(
            [
                payload.get("scenario_name"),
                route.get("id"),
                route.get("bus_type_name"),
                route.get("load"),
                route.get("stop_count"),
                route_impact.get("over_acceptance_rider_count"),
                route_impact.get("over_acceptance_stop_count"),
                route_impact.get("worse_rider_count"),
                route_impact.get("high_risk_stop_count"),
                route_impact.get("weighted_avg_adverse_delta_minutes"),
                route_impact.get("max_adverse_delta_minutes"),
                route_impact.get("route_changed_rider_count"),
                route_gate.get("status"),
                route_gate.get("verified_departure_label"),
                route_gate.get("verified_arrival_label"),
                route_gate.get("time_window_overrun_minutes"),
            ]
        )
    _append_excel_table(route_sheet, route_headers, route_rows)

    stop_sheet = wb.create_sheet("Stops")
    stop_headers = [
        "Scenario",
        "Optimized route",
        "Stop order",
        "Address",
        "Riders",
        "Current route",
        "Current time",
        "Optimized time",
        "Delta minutes",
        "Adverse minutes",
        "Acceptance status",
        "Over-threshold minutes",
        "Absolute minutes",
        "Impact direction",
        "Level",
        "Route changed",
        "Comparison status",
        "Matched key",
    ]
    stop_rows = []
    for stop in list(payload.get("stops") or []):
        stop = dict(stop or {})
        if bool(stop.get("is_depot")):
            continue
        impact = dict(stop.get("time_impact") or {})
        stop_rows.append(
            [
                payload.get("scenario_name"),
                stop.get("route_id"),
                stop.get("order"),
                stop.get("address") or stop.get("requested_address"),
                impact.get("affected_rider_count", stop.get("passenger_count")),
                impact.get("current_route_id"),
                impact.get("current_time_label"),
                impact.get("new_time_label") or stop.get("scheduled_time_label"),
                impact.get("delta_minutes"),
                impact.get("adverse_delta_minutes"),
                impact.get("acceptance_status"),
                impact.get("over_acceptance_minutes"),
                impact.get("absolute_delta_minutes"),
                impact.get("impact_direction"),
                impact.get("level"),
                "yes" if impact.get("route_changed") else "no",
                impact.get("comparison_status"),
                impact.get("matched_key"),
            ]
        )
    _append_excel_table(stop_sheet, stop_headers, stop_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), None


def _map_tile_cache_path(z: int, x: int, y: int) -> Path:
    return MAP_TILE_CACHE_DIR / str(z) / str(x) / f"{y}.png"


def _parse_map_tile_path(path: str) -> tuple[int, int, int] | None:
    parts = [part for part in path.split("/") if part]
    if len(parts) != 4 or parts[0] != "map-tiles":
        return None
    raw_y = parts[3]
    if not raw_y.endswith(".png"):
        return None
    try:
        z = int(parts[1])
        x = int(parts[2])
        y = int(raw_y[:-4])
    except ValueError:
        return None
    if z < 0 or z > 22:
        return None
    max_tile = 2**z
    if x < 0 or y < 0 or x >= max_tile or y >= max_tile:
        return None
    return z, x, y


def _load_or_fetch_map_tile(z: int, x: int, y: int) -> tuple[bytes, bool]:
    tile_path = _map_tile_cache_path(z, x, y)
    if tile_path.exists():
        try:
            return tile_path.read_bytes(), True
        except OSError:
            pass

    upstream_url = MAP_TILE_UPSTREAM_TEMPLATE.format(z=z, x=x, y=y)
    request = Request(
        upstream_url,
        headers={"User-Agent": "BRP route planner tile proxy/1.0"},
    )
    try:
        with urlopen(request, timeout=8) as response:
            if getattr(response, "status", 200) != 200:
                raise HTTPError(
                    upstream_url,
                    getattr(response, "status", 502),
                    "Tile upstream returned non-200",
                    response.headers,
                    None,
                )
            body = response.read()
            if body:
                tile_path.parent.mkdir(parents=True, exist_ok=True)
                try:
                    tile_path.write_bytes(body)
                except OSError:
                    pass
                return body, False
    except (HTTPError, URLError, TimeoutError, OSError) as exc:
        print(f"[WARN] Map tile fetch failed z={z} x={x} y={y}: {exc}")
    return MAP_TILE_FALLBACK_BYTES, False


if __name__ == "__main__":
    raise SystemExit(
        "backend_service.py no longer starts an HTTP server. "
        "Use ops/scripts/run_backend.sh or ops/scripts/run_backend.ps1 to start FastAPI."
    )
