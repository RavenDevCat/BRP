from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from datetime import datetime, timedelta, timezone
import hashlib
import json
import math
from pathlib import Path
import statistics
import sys
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
CLIENT_DIR = BASE_DIR.parent / "client"
if str(CLIENT_DIR) not in sys.path:
    sys.path.insert(0, str(CLIENT_DIR))

try:
    from . import planner_core
except ImportError:  # pragma: no cover - direct worker execution
    import planner_core  # type: ignore

import distance_tool  # type: ignore


DEFAULT_ANALYSIS_CONFIG: dict[str, Any] = {
    "service_direction": "To School",
    "stop_service_minutes": 1.0,
    "time_window_start": "06:30",
    "time_window_end": "08:00",
    "from_school_departure_time": "15:40",
    "far_duration_minutes": 45.0,
    "provider_call_limit": 500,
}

ADDITIONAL_REMOVAL_STRATEGY = "direct_duration_desc_then_route_saving"


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return default
    return parsed if math.isfinite(parsed) else default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _normalized_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _address_key(item: dict[str, Any]) -> str:
    parts = (
        _normalized_text(item.get("country")),
        _normalized_text(item.get("city")),
        _normalized_text(item.get("address") or item.get("display_address")),
    )
    return "|".join(parts)


def _stop_key(item: dict[str, Any]) -> str:
    return hashlib.sha1(_address_key(item).encode("utf-8")).hexdigest()[:16]


def _point_coordinates(point: dict[str, Any]) -> tuple[float, float] | None:
    lat = _safe_float(point.get("plot_lat", point.get("lat")), math.nan)
    lng = _safe_float(point.get("plot_lng", point.get("lng")), math.nan)
    if not math.isfinite(lat) or not math.isfinite(lng):
        return None
    return lat, lng


def _haversine_km(a: tuple[float, float], b: tuple[float, float]) -> float:
    lat1, lng1 = map(math.radians, a)
    lat2, lng2 = map(math.radians, b)
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    return 6371.0088 * 2 * math.atan2(math.sqrt(h), math.sqrt(max(0.0, 1 - h)))


def _clock_minutes(value: Any, default: int) -> int:
    parts = str(value or "").strip().split(":")
    if len(parts) < 2:
        return default
    try:
        return (int(parts[0]) * 60 + int(parts[1])) % (24 * 60)
    except ValueError:
        return default


def _clock_label(minutes: float) -> str:
    total = int(round(minutes)) % (24 * 60)
    return f"{total // 60:02d}:{total % 60:02d}"


def _window_duration_minutes(config: dict[str, Any]) -> float:
    start = _clock_minutes(config.get("time_window_start"), 6 * 60 + 30)
    end = _clock_minutes(config.get("time_window_end"), 8 * 60)
    duration = (end - start) % (24 * 60)
    return float(duration or 90)


def _measure_active_route(
    provider: "FreshRouteProvider",
    runtime: dict[str, Any],
    active_indexes: list[int],
    stop_service_minutes: float,
) -> dict[str, Any]:
    ordered = list(runtime["ordered"])
    points = list(runtime["points"])
    active_points = [points[index] for index in active_indexes]
    service_indexes = [index for index in active_indexes if not bool(ordered[index].get("is_depot"))]
    if len(active_points) >= 2:
        live = provider.route(active_points)
        drive_s = _safe_float(live.get("duration_s"))
        distance_m = _safe_float(live.get("distance_m"))
        called_at = live.get("called_at")
    else:
        drive_s = 0.0
        distance_m = 0.0
        called_at = None
    dwell_s = len(service_indexes) * stop_service_minutes * 60.0
    return {
        "total_duration_min": round((drive_s + dwell_s) / 60.0, 2),
        "provider_duration_min": round(drive_s / 60.0, 2),
        "provider_distance_km": round(distance_m / 1000.0, 3),
        "stop_count": len(service_indexes),
        "riders": sum(max(0, _safe_int(ordered[index].get("passenger_count"))) for index in service_indexes),
        "provider_called_at": called_at,
    }


def _estimated_removal_saving_min(
    runtime: dict[str, Any],
    active_indexes: list[int],
    remove_index: int,
    stop_service_minutes: float,
    osrm_cache: dict[str, dict[str, Any]],
) -> float:
    position = active_indexes.index(remove_index)
    points = list(runtime["points"])
    scale = max(0.0, _safe_float(runtime.get("scale"), 1.0))
    if len(active_indexes) <= 1:
        drive_s = 0.0
    elif position == 0:
        drive_s = _safe_float(_osrm_leg(points[remove_index], points[active_indexes[1]], osrm_cache).get("duration_s"))
    elif position == len(active_indexes) - 1:
        drive_s = _safe_float(_osrm_leg(points[active_indexes[-2]], points[remove_index], osrm_cache).get("duration_s"))
    else:
        previous_index = active_indexes[position - 1]
        next_index = active_indexes[position + 1]
        through_s = _safe_float(_osrm_leg(points[previous_index], points[remove_index], osrm_cache).get("duration_s"))
        through_s += _safe_float(_osrm_leg(points[remove_index], points[next_index], osrm_cache).get("duration_s"))
        bypass_s = _safe_float(_osrm_leg(points[previous_index], points[next_index], osrm_cache).get("duration_s"))
        drive_s = max(0.0, through_s - bypass_s)
    return round(max(0.0, drive_s * scale / 60.0) + stop_service_minutes, 2)


def _additional_removal_rank(
    occurrence: dict[str, Any],
    *,
    estimated_saving_min: float,
    riders: int,
    stop_sequence: int,
) -> tuple[float, float, int, int]:
    """Prefer the longest direct trip, with deterministic operational tie-breakers."""
    return (
        _safe_float(occurrence.get("direct_duration_min")),
        _safe_float(estimated_saving_min),
        -max(0, riders),
        -max(0, stop_sequence),
    )


def _percentile(values: list[float], quantile: float) -> float | None:
    ordered = sorted(value for value in values if math.isfinite(value))
    if not ordered:
        return None
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * min(1.0, max(0.0, quantile))
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def _analysis_config(value: dict[str, Any] | None) -> dict[str, Any]:
    config = {**DEFAULT_ANALYSIS_CONFIG, **dict(value or {})}
    for legacy_key in (
        "far_distance_km",
        "burden_minutes",
        "bypass_candidate_limit",
        "candidate_cluster_radius_km",
    ):
        config.pop(legacy_key, None)
    config["service_direction"] = (
        "To School" if str(config.get("service_direction")) == "To School" else "From School"
    )
    for key in (
        "stop_service_minutes",
        "far_duration_minutes",
    ):
        config[key] = max(0.0, _safe_float(config.get(key), _safe_float(DEFAULT_ANALYSIS_CONFIG[key])))
    config["provider_call_limit"] = max(1, min(2000, _safe_int(config.get("provider_call_limit"), 500)))
    return config


def _point_lookup(
    input_records: list[dict[str, Any]],
    points: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for point in points:
        key_candidates = {
            _address_key(point),
            _normalized_text(point.get("display_address")),
        }
        for member in list(point.get("original_members") or []):
            key_candidates.add(_normalized_text(member))
        for key in key_candidates:
            if key:
                lookup.setdefault(key, dict(point))
    if len(input_records) == len(points):
        for record, point in zip(input_records, points):
            lookup.setdefault(_address_key(record), dict(point))
            lookup.setdefault(_normalized_text(record.get("address")), dict(point))
    return lookup


def _lookup_point(item: dict[str, Any], lookup: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    point = lookup.get(_address_key(item)) or lookup.get(_normalized_text(item.get("address")))
    return dict(point) if point else None


def _provider_for_records(input_records: list[dict[str, Any]]) -> str:
    country, _city = planner_core.infer_traffic_location(input_records)
    normalized = str(country or "").strip().upper()
    if normalized == "CHINA":
        return "amap"
    if normalized == "SOUTH KOREA":
        return "kakao_navi"
    return "none"


class FreshRouteProvider:
    def __init__(
        self,
        provider: str,
        *,
        departure_time: datetime | None,
        api_call_limit: int,
    ) -> None:
        self.provider = provider
        self.departure_time = departure_time
        self.planner = planner_core.load_legacy_planner()
        self.cache: dict[str, Any] = {}
        self.state: dict[str, int] = {
            "api_calls": 0,
            "api_call_limit": api_call_limit,
            "cache_hits": 0,
        }
        if provider == "amap" and not str(getattr(self.planner, "AMAP_KEY", "") or "").strip():
            raise RuntimeError("AMap route API is not configured for this deployment.")
        if provider == "kakao_navi" and not planner_core.KAKAO_NAVI_API_KEY:
            raise RuntimeError("Kakao Navi route API is not configured for this deployment.")
        if provider == "none":
            raise RuntimeError("No live route provider is configured for this workbook market.")

    def route(self, points: list[dict[str, Any]]) -> dict[str, Any]:
        request_points = [coords for point in points if (coords := _point_coordinates(point))]
        if len(request_points) != len(points) or len(request_points) < 2:
            raise RuntimeError("Route contains unresolved coordinates.")
        before_calls = int(self.state.get("api_calls", 0))
        called_at = utc_now_iso()
        if self.provider == "amap":
            result = planner_core._amap_route_stats(
                self.planner, request_points, self.cache, self.state
            )
        else:
            departure = self.departure_time or datetime.now(timezone.utc) + timedelta(minutes=10)
            result = planner_core._kakao_route_stats(
                request_points,
                self.cache,
                self.state,
                departure_time=departure,
            )
        if not result:
            raise RuntimeError("Live route provider returned no usable route.")
        return {
            **dict(result),
            "provider": self.provider,
            "called_at": called_at,
            "api_calls": int(self.state.get("api_calls", 0)) - before_calls,
            "in_run_reuse": int(self.state.get("api_calls", 0)) == before_calls,
        }


def _osrm_leg(
    origin: dict[str, Any],
    destination: dict[str, Any],
    cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    origin_coords = _point_coordinates(origin)
    destination_coords = _point_coordinates(destination)
    if not origin_coords or not destination_coords:
        raise RuntimeError("OSRM route contains unresolved coordinates.")
    key = "|".join(
        f"{value:.6f}" for value in (*origin_coords, *destination_coords)
    )
    if key not in cache:
        # Audit maps route on plot coordinates (WGS84), while China geocoders also
        # retain raw GCJ-02 coordinates in lat/lng. Normalize the request before
        # calling the shared distance helper so the route and map markers use the
        # same coordinate system.
        route_origin = {
            **origin,
            "lat": origin_coords[0],
            "lng": origin_coords[1],
        }
        route_destination = {
            **destination,
            "lat": destination_coords[0],
            "lng": destination_coords[1],
        }
        detail = distance_tool.compute_osrm_route_leg_details(
            [route_origin, route_destination]
        )[0]
        geometry = [
            [float(lng), float(lat)]
            for lat, lng in list(detail.get("geometry") or [])
        ]
        snap_connectors: list[dict[str, Any]] = []
        for connector in list(detail.get("snap_connectors") or []):
            connector_geometry = [
                [float(lng), float(lat)]
                for lat, lng in list(dict(connector or {}).get("geometry") or [])
            ]
            if len(connector_geometry) < 2:
                continue
            snap_connectors.append(
                {
                    "type": str(dict(connector or {}).get("type") or "snap"),
                    "distance_m": _safe_float(
                        dict(connector or {}).get("distance_m")
                    ),
                    "geometry": connector_geometry,
                }
            )
        cache[key] = {
            "duration_s": detail.get("duration_s"),
            "distance_m": detail.get("distance_m"),
            "geometry": geometry,
            "snap_connectors": snap_connectors,
            "coordinate_source": "plot_wgs84",
        }
    return deepcopy(cache[key])


def _rotate(items: list[Any], seed: str) -> list[Any]:
    if len(items) < 2:
        return list(items)
    offset = int(hashlib.sha1(seed.encode("utf-8")).hexdigest()[:8], 16) % len(items)
    return [*items[offset:], *items[:offset]]


def _base_result(
    *,
    config: dict[str, Any],
    provider: str,
    school: dict[str, Any],
    rows: list[dict[str, Any]],
    route_count: int,
    logical_call_estimate: int,
    progress: dict[str, Any],
    errors: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "analysis_version": 4,
        "analysis_type": "direct_school",
        "status": "running",
        "generated_at": utc_now_iso(),
        "provider": provider,
        "service_direction": config["service_direction"],
        "school": school,
        "parameters": config,
        "summary": {
            "address_count": len(rows),
            "route_count": route_count,
            "resolved_count": sum(1 for row in rows if row.get("provider_status") == "resolved"),
            "failed_count": sum(1 for row in rows if row.get("provider_status") == "failed"),
            "logical_call_estimate": logical_call_estimate,
        },
        "progress": progress,
        "stops": rows,
        "routes": [],
        "errors": errors,
    }


def run_direct_school_analysis(
    prepared_payload: dict[str, Any],
    analysis_config: dict[str, Any] | None = None,
    *,
    scheduled_start_at: str | None = None,
    run_seed: str = "",
    checkpoint: Callable[[dict[str, Any]], None] | None = None,
    resume_result: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = _analysis_config(analysis_config)
    input_records = [dict(item) for item in list(prepared_payload.get("input_records") or [])]
    points = [dict(item) for item in list(prepared_payload.get("original_points") or [])]
    current_plan = dict(prepared_payload.get("current_plan") or {})
    stops = [dict(item) for item in list(current_plan.get("stops") or [])]
    if not input_records or not points or not stops:
        raise ValueError("Direct-to-school analysis requires a validated Audit workbook.")

    provider_name = _provider_for_records(input_records)
    departure_time = None
    if scheduled_start_at:
        try:
            departure_time = datetime.fromisoformat(str(scheduled_start_at).replace("Z", "+00:00"))
        except ValueError:
            departure_time = None
    provider = FreshRouteProvider(
        provider_name,
        departure_time=departure_time,
        api_call_limit=int(config["provider_call_limit"]),
    )
    lookup = _point_lookup(input_records, points)
    school_record = input_records[0]
    school_point = _lookup_point(school_record, lookup)
    if not school_point:
        raise ValueError("The shared school address could not be geocoded.")
    school_coords = _point_coordinates(school_point)
    if not school_coords:
        raise ValueError("The shared school address could not be geocoded.")
    school = {
        "country": school_record.get("country"),
        "city": school_record.get("city"),
        "address": school_record.get("address"),
        "lat": school_coords[0],
        "lng": school_coords[1],
    }

    route_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    unique_stops: dict[str, dict[str, Any]] = {}
    for stop in stops:
        route_id = str(stop.get("route_id") or "").strip()
        route_groups[route_id].append(stop)
        if bool(stop.get("is_depot")):
            continue
        key = _address_key(stop)
        item = unique_stops.setdefault(
            key,
            {
                "stop_key": _stop_key(stop),
                "country": stop.get("country"),
                "city": stop.get("city"),
                "address": stop.get("address"),
                "riders": 0,
                "route_ids": [],
                "occurrences": [],
                "point": _lookup_point(stop, lookup),
            },
        )
        item["riders"] += max(0, _safe_int(stop.get("passenger_count")))
        if route_id and route_id not in item["route_ids"]:
            item["route_ids"].append(route_id)
        item["occurrences"].append(
            {
                "route_id": route_id,
                "stop_sequence": _safe_int(stop.get("stop_sequence")),
                "passenger_count": max(0, _safe_int(stop.get("passenger_count"))),
            }
        )

    prior_rows = {
        str(row.get("stop_key") or ""): dict(row)
        for row in list(dict(resume_result or {}).get("stops") or [])
        if str(row.get("stop_key") or "")
    }
    rows: list[dict[str, Any]] = []
    for item in unique_stops.values():
        point = dict(item.pop("point") or {})
        coords = _point_coordinates(point)
        row = {
            **item,
            "lat": coords[0] if coords else None,
            "lng": coords[1] if coords else None,
            "provider_status": "pending" if coords else "failed",
            "quality_status": "ready" if coords else "geocode_failed",
            "operational_category": "data_review" if not coords else "pending",
            "reasons": ["Address could not be geocoded."] if not coords else [],
        }
        previous = prior_rows.get(str(row["stop_key"]))
        if previous and previous.get("provider_status") == "resolved":
            row.update(previous)
        row["_point"] = point
        rows.append(row)

    service_occurrence_count = sum(
        1
        for route_stops in route_groups.values()
        for stop in route_stops
        if not bool(stop.get("is_depot"))
    )
    logical_call_estimate = (
        len(rows)
        + len(route_groups)
        + len(route_groups)
        + service_occurrence_count
    )
    errors: list[dict[str, Any]] = list(dict(resume_result or {}).get("errors") or [])
    progress = {
        "phase": "direct_routes",
        "completed": sum(1 for row in rows if row.get("provider_status") == "resolved"),
        "total": logical_call_estimate,
        "provider_api_calls": int(provider.state.get("api_calls", 0)),
        "in_run_reuse_count": int(provider.state.get("cache_hits", 0)),
    }
    osrm_cache: dict[str, dict[str, Any]] = {}

    def save_checkpoint() -> None:
        if not checkpoint:
            return
        public_rows = [{key: value for key, value in row.items() if key != "_point"} for row in rows]
        checkpoint(
            _base_result(
                config=config,
                provider=provider_name,
                school=school,
                rows=public_rows,
                route_count=len(route_groups),
                logical_call_estimate=logical_call_estimate,
                progress=deepcopy(progress),
                errors=deepcopy(errors),
            )
        )

    direct_order = _rotate(rows, run_seed or scheduled_start_at or utc_now_iso())
    for row in direct_order:
        if row.get("provider_status") == "resolved":
            continue
        point = dict(row.get("_point") or {})
        if not point:
            progress["completed"] += 1
            save_checkpoint()
            continue
        request_points = [point, school_point]
        if config["service_direction"] == "From School":
            request_points.reverse()
        try:
            osrm = _osrm_leg(request_points[0], request_points[1], osrm_cache)
            live = provider.route(request_points)
            straight_km = _haversine_km(_point_coordinates(point), school_coords)  # type: ignore[arg-type]
            direct_distance_m = _safe_float(live.get("distance_m"))
            direct_duration_s = _safe_float(live.get("duration_s"))
            osrm_distance_m = _safe_float(osrm.get("distance_m"))
            osrm_duration_s = _safe_float(osrm.get("duration_s"))
            row.update(
                {
                    "provider_status": "resolved",
                    "quality_status": "ready",
                    "provider": provider_name,
                    "provider_called_at": live.get("called_at"),
                    "direct_distance_km": round(direct_distance_m / 1000.0, 3),
                    "direct_duration_min": round(direct_duration_s / 60.0, 2),
                    "direct_service_duration_min": round(
                        direct_duration_s / 60.0 + float(config["stop_service_minutes"]), 2
                    ),
                    "osrm_distance_km": round(osrm_distance_m / 1000.0, 3),
                    "osrm_duration_min": round(osrm_duration_s / 60.0, 2),
                    "straight_distance_km": round(straight_km, 3),
                    "congestion_increment_min": round((direct_duration_s - osrm_duration_s) / 60.0, 2),
                    "live_to_osrm_ratio": round(direct_duration_s / osrm_duration_s, 3) if osrm_duration_s > 0 else None,
                    "road_to_straight_ratio": round((direct_distance_m / 1000.0) / straight_km, 3) if straight_km > 0 else None,
                    "direct_geometry": osrm.get("geometry") or [],
                    "direct_snap_connectors": osrm.get("snap_connectors") or [],
                    "direct_geometry_source": osrm.get("coordinate_source") or "plot_wgs84",
                }
            )
            if config["service_direction"] == "To School":
                latest_arrival = _clock_minutes(config.get("time_window_end"), 8 * 60)
                row["latest_direct_departure"] = _clock_label(latest_arrival - direct_duration_s / 60.0)
            else:
                departure = _clock_minutes(config.get("from_school_departure_time"), 15 * 60 + 40)
                row["estimated_direct_arrival"] = _clock_label(departure + direct_duration_s / 60.0)
        except Exception as exc:
            row["provider_status"] = "failed"
            row["quality_status"] = "provider_failed"
            row["operational_category"] = "data_review"
            row["reasons"] = [str(exc)]
            errors.append({"scope": "direct_route", "stop_key": row["stop_key"], "address": row["address"], "error": str(exc)})
        progress["completed"] += 1
        progress["provider_api_calls"] = int(provider.state.get("api_calls", 0))
        progress["in_run_reuse_count"] = int(provider.state.get("cache_hits", 0))
        save_checkpoint()

    progress["phase"] = "current_routes"
    route_results: list[dict[str, Any]] = []
    route_contexts_by_stop: dict[str, list[dict[str, Any]]] = defaultdict(list)
    route_runtime: dict[str, dict[str, Any]] = {}
    route_order = _rotate(sorted(route_groups), f"route|{run_seed or scheduled_start_at or ''}")
    for route_id in route_order:
        ordered = sorted(route_groups[route_id], key=lambda item: _safe_int(item.get("stop_sequence")))
        route_points = [_lookup_point(stop, lookup) for stop in ordered]
        if any(point is None for point in route_points):
            error = "One or more route stops could not be geocoded."
            route_results.append({"route_id": route_id, "status": "partial", "error": error})
            errors.append({"scope": "current_route", "route_id": route_id, "error": error})
            progress["completed"] += 1
            save_checkpoint()
            continue
        resolved_points = [dict(point) for point in route_points if point]
        try:
            leg_details = [
                _osrm_leg(origin, destination, osrm_cache)
                for origin, destination in zip(resolved_points[:-1], resolved_points[1:])
            ]
            osrm_drive_s = sum(_safe_float(leg.get("duration_s")) for leg in leg_details)
            osrm_distance_m = sum(_safe_float(leg.get("distance_m")) for leg in leg_details)
            live = provider.route(resolved_points)
            live_drive_s = _safe_float(live.get("duration_s"))
            live_distance_m = _safe_float(live.get("distance_m"))
            service_stop_count = sum(1 for stop in ordered if not bool(stop.get("is_depot")))
            dwell_s = service_stop_count * float(config["stop_service_minutes"]) * 60.0
            scale = live_drive_s / osrm_drive_s if osrm_drive_s > 0 else 1.0
            geometry = [coord for leg in leg_details for coord in list(leg.get("geometry") or [])]
            route_result = {
                "route_id": route_id,
                "status": "resolved",
                "stop_count": service_stop_count,
                "riders": sum(max(0, _safe_int(stop.get("passenger_count"))) for stop in ordered),
                "provider_duration_min": round(live_drive_s / 60.0, 2),
                "total_duration_min": round((live_drive_s + dwell_s) / 60.0, 2),
                "provider_distance_km": round(live_distance_m / 1000.0, 3),
                "osrm_duration_min": round(osrm_drive_s / 60.0, 2),
                "osrm_distance_km": round(osrm_distance_m / 1000.0, 3),
                "congestion_ratio": round(scale, 3),
                "provider_called_at": live.get("called_at"),
                "geometry": geometry,
            }
            route_results.append(route_result)
            route_runtime[route_id] = {
                "ordered": ordered,
                "points": resolved_points,
                "legs": leg_details,
                "live_drive_s": live_drive_s,
                "dwell_s": dwell_s,
                "scale": scale,
                "full_total_s": live_drive_s + dwell_s,
                "full_distance_m": live_distance_m,
            }
            service_indexes = [index for index, stop in enumerate(ordered) if not bool(stop.get("is_depot"))]
            for index in service_indexes:
                stop = ordered[index]
                key = _address_key(stop)
                if config["service_direction"] == "To School":
                    relevant_legs = leg_details[index:]
                    dwell_count = sum(1 for service_index in service_indexes if service_index >= index)
                else:
                    relevant_legs = leg_details[:index]
                    dwell_count = sum(1 for service_index in service_indexes if service_index <= index)
                ride_s = sum(_safe_float(leg.get("duration_s")) for leg in relevant_legs) * scale
                ride_s += dwell_count * float(config["stop_service_minutes"]) * 60.0
                route_contexts_by_stop[key].append(
                    {
                        "route_id": route_id,
                        "stop_sequence": _safe_int(stop.get("stop_sequence")),
                        "riders": max(0, _safe_int(stop.get("passenger_count"))),
                        "estimated_current_ride_min": round(ride_s / 60.0, 2),
                        "route_total_min": route_result["total_duration_min"],
                    }
                )
        except Exception as exc:
            route_results.append({"route_id": route_id, "status": "failed", "error": str(exc)})
            errors.append({"scope": "current_route", "route_id": route_id, "error": str(exc)})
        progress["completed"] += 1
        progress["provider_api_calls"] = int(provider.state.get("api_calls", 0))
        progress["in_run_reuse_count"] = int(provider.state.get("cache_hits", 0))
        save_checkpoint()

    for row in rows:
        contexts = route_contexts_by_stop.get(_address_key(row), [])
        row["route_contexts"] = contexts
        if contexts:
            worst = max(contexts, key=lambda item: _safe_float(item.get("estimated_current_ride_min")))
            row["primary_route_id"] = worst.get("route_id")
            row["estimated_current_ride_min"] = worst.get("estimated_current_ride_min")
            if row.get("direct_duration_min") is not None:
                row["rider_detour_min"] = round(
                    _safe_float(worst.get("estimated_current_ride_min")) - _safe_float(row.get("direct_duration_min")),
                    2,
                )

    duration_limit = float(config["far_duration_minutes"])
    primary_removals_by_route: dict[str, set[int]] = defaultdict(set)
    occurrence_lookup: dict[tuple[str, int], dict[str, Any]] = {}
    for row in rows:
        direct_over = (
            row.get("provider_status") == "resolved"
            and _safe_float(row.get("direct_duration_min")) >= duration_limit
        )
        row_categories: set[str] = set()
        for context in list(row.get("route_contexts") or []):
            route_id = str(context.get("route_id") or "")
            sequence = _safe_int(context.get("stop_sequence"))
            current_ride = _safe_float(context.get("estimated_current_ride_min"))
            if direct_over:
                category = "direct_over_limit"
                excess = _safe_float(row.get("direct_duration_min")) - duration_limit
            elif current_ride >= duration_limit:
                category = "route_only_over_limit"
                excess = current_ride - duration_limit
            else:
                category = "within_limit"
                excess = 0.0
            context["operational_category"] = category
            context["over_limit_min"] = round(max(0.0, excess), 2)
            row_categories.add(category)
            occurrence_lookup[(route_id, sequence)] = {
                "stop_key": row.get("stop_key"),
                "address": row.get("address"),
                "riders": max(0, _safe_int(context.get("riders"))),
                "direct_duration_min": row.get("direct_duration_min"),
                "current_ride_min": context.get("estimated_current_ride_min"),
                "operational_category": category,
                "over_limit_min": context.get("over_limit_min"),
            }
            if category in {"direct_over_limit", "route_only_over_limit"}:
                primary_removals_by_route[route_id].add(sequence)
        if direct_over:
            row["operational_category"] = "direct_over_limit"
        elif "route_only_over_limit" in row_categories:
            row["operational_category"] = "route_only_over_limit"
        elif row.get("provider_status") == "resolved":
            row["operational_category"] = "within_limit"
        else:
            row["operational_category"] = "data_review"

    progress["phase"] = "route_window_recovery"
    route_window_min = _window_duration_minutes(config)
    route_window_analysis: list[dict[str, Any]] = []
    additional_occurrences: list[dict[str, Any]] = []
    route_results_by_id = {str(item.get("route_id") or ""): item for item in route_results}
    for route_id in sorted(route_groups):
        runtime = route_runtime.get(route_id)
        original = route_results_by_id.get(route_id, {})
        if not runtime or original.get("status") != "resolved":
            route_window_analysis.append(
                {
                    "route_id": route_id,
                    "status": "data_review",
                    "window_limit_min": route_window_min,
                    "error": original.get("error") or "Current route could not be measured.",
                }
            )
            continue
        ordered = list(runtime["ordered"])
        active_indexes = list(range(len(ordered)))
        primary_sequences = primary_removals_by_route.get(route_id, set())
        primary_indexes = [
            index
            for index, stop in enumerate(ordered)
            if not bool(stop.get("is_depot")) and _safe_int(stop.get("stop_sequence")) in primary_sequences
        ]
        primary_entries = [
            occurrence_lookup.get((route_id, _safe_int(ordered[index].get("stop_sequence"))), {})
            for index in primary_indexes
        ]
        active_indexes = [index for index in active_indexes if index not in primary_indexes]
        try:
            if primary_indexes:
                post_primary = _measure_active_route(
                    provider,
                    runtime,
                    active_indexes,
                    float(config["stop_service_minutes"]),
                )
            else:
                post_primary = {
                    "total_duration_min": original.get("total_duration_min"),
                    "provider_duration_min": original.get("provider_duration_min"),
                    "provider_distance_km": original.get("provider_distance_km"),
                    "stop_count": original.get("stop_count"),
                    "riders": original.get("riders"),
                    "provider_called_at": original.get("provider_called_at"),
                }
            additional_entries: list[dict[str, Any]] = []
            final_measurement = dict(post_primary)
            while _safe_float(final_measurement.get("total_duration_min")) > route_window_min:
                candidates: list[dict[str, Any]] = []
                for index in active_indexes:
                    stop = ordered[index]
                    if bool(stop.get("is_depot")):
                        continue
                    riders = max(0, _safe_int(stop.get("passenger_count")))
                    sequence = _safe_int(stop.get("stop_sequence"))
                    occurrence = occurrence_lookup.get((route_id, sequence), {})
                    saving = _estimated_removal_saving_min(
                        runtime,
                        active_indexes,
                        index,
                        float(config["stop_service_minutes"]),
                        osrm_cache,
                    )
                    candidates.append(
                        {
                            "index": index,
                            "occurrence": occurrence,
                            "estimated_saving_min": saving,
                            "rank": _additional_removal_rank(
                                occurrence,
                                estimated_saving_min=saving,
                                riders=riders,
                                stop_sequence=sequence,
                            ),
                        }
                    )
                if not candidates:
                    break
                candidate = max(candidates, key=lambda item: item["rank"])
                remove_index = _safe_int(candidate.get("index"))
                estimated_saving = _safe_float(candidate.get("estimated_saving_min"))
                stop = ordered[remove_index]
                sequence = _safe_int(stop.get("stop_sequence"))
                entry = {
                    **dict(candidate.get("occurrence") or {}),
                    "route_id": route_id,
                    "stop_sequence": sequence,
                    "riders": max(0, _safe_int(stop.get("passenger_count"))),
                    "estimated_route_saving_min": estimated_saving,
                    "selection_rank": len(additional_entries) + 1,
                    "selection_basis": ADDITIONAL_REMOVAL_STRATEGY,
                    "operational_category": "additional_window_candidate",
                }
                additional_entries.append(entry)
                active_indexes.remove(remove_index)
                final_measurement = _measure_active_route(
                    provider,
                    runtime,
                    active_indexes,
                    float(config["stop_service_minutes"]),
                )
            final_duration = _safe_float(final_measurement.get("total_duration_min"))
            final_status = "within_window" if final_duration <= route_window_min else "still_over_window"
            route_window_analysis.append(
                {
                    "route_id": route_id,
                    "status": final_status,
                    "window_limit_min": route_window_min,
                    "original_duration_min": original.get("total_duration_min"),
                    "original_stop_count": original.get("stop_count"),
                    "original_riders": original.get("riders"),
                    "primary_removed_addresses": len(primary_entries),
                    "primary_removed_riders": sum(_safe_int(item.get("riders")) for item in primary_entries),
                    "primary_removals": primary_entries,
                    "post_primary_duration_min": post_primary.get("total_duration_min"),
                    "post_primary_stop_count": post_primary.get("stop_count"),
                    "post_primary_riders": post_primary.get("riders"),
                    "additional_removed_addresses": len(additional_entries),
                    "additional_removed_riders": sum(_safe_int(item.get("riders")) for item in additional_entries),
                    "additional_removals": additional_entries,
                    "additional_removal_strategy": ADDITIONAL_REMOVAL_STRATEGY,
                    "final_duration_min": final_measurement.get("total_duration_min"),
                    "final_stop_count": final_measurement.get("stop_count"),
                    "final_riders": final_measurement.get("riders"),
                    "provider_called_at": final_measurement.get("provider_called_at"),
                }
            )
            additional_occurrences.extend(additional_entries)
        except Exception as exc:
            route_window_analysis.append(
                {
                    "route_id": route_id,
                    "status": "data_review",
                    "window_limit_min": route_window_min,
                    "original_duration_min": original.get("total_duration_min"),
                    "primary_removals": primary_entries,
                    "error": str(exc),
                }
            )
            errors.append({"scope": "route_window_recovery", "route_id": route_id, "error": str(exc)})
        progress["provider_api_calls"] = int(provider.state.get("api_calls", 0))
        progress["in_run_reuse_count"] = int(provider.state.get("cache_hits", 0))

    additional_by_stop: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in additional_occurrences:
        additional_by_stop[str(item.get("stop_key") or "")].append(item)
    for row in rows:
        matches = additional_by_stop.get(str(row.get("stop_key") or ""), [])
        if matches:
            row["additional_window_candidate"] = True
            row["additional_window_routes"] = sorted({str(item.get("route_id") or "") for item in matches})
            row["operational_category"] = "additional_window_candidate"
        category = str(row.get("operational_category") or "data_review")
        if category == "direct_over_limit":
            row["reasons"] = ["Direct travel time exceeds the configured limit."]
        elif category == "route_only_over_limit":
            row["reasons"] = ["Direct travel time fits the limit, but current shared-route ride time exceeds it."]
        elif category == "additional_window_candidate":
            row["reasons"] = ["This address is an additional removal candidate because its route still exceeds the time window."]
        elif category == "within_limit":
            row["reasons"] = ["Direct and current shared-route travel times are within the configured limit."]

    def occurrence_totals(category: str) -> tuple[int, int]:
        matching = [item for item in occurrence_lookup.values() if item.get("operational_category") == category]
        return len({str(item.get("stop_key") or "") for item in matching}), sum(_safe_int(item.get("riders")) for item in matching)

    direct_address_count, direct_rider_count = occurrence_totals("direct_over_limit")
    route_only_address_count, route_only_rider_count = occurrence_totals("route_only_over_limit")
    primary_stop_keys = {
        str(item.get("stop_key") or "")
        for item in occurrence_lookup.values()
        if item.get("operational_category") in {"direct_over_limit", "route_only_over_limit"}
    }
    primary_rider_count = direct_rider_count + route_only_rider_count
    post_primary_over = [item for item in route_window_analysis if _safe_float(item.get("post_primary_duration_min")) > route_window_min]
    final_over = [item for item in route_window_analysis if item.get("status") == "still_over_window"]
    route_data_review = [item for item in route_window_analysis if item.get("status") == "data_review"]
    additional_stop_keys = {str(item.get("stop_key") or "") for item in additional_occurrences}
    additional_rider_count = sum(_safe_int(item.get("riders")) for item in additional_occurrences)
    operational_conclusion = {
        "duration_limit_min": duration_limit,
        "route_window_min": route_window_min,
        "direct_over_limit": {
            "address_count": direct_address_count,
            "rider_count": direct_rider_count,
        },
        "route_only_over_limit": {
            "address_count": route_only_address_count,
            "rider_count": route_only_rider_count,
        },
        "primary_removal": {
            "address_count": len(primary_stop_keys),
            "rider_count": primary_rider_count,
        },
        "post_primary": {
            "route_count": len(route_window_analysis),
            "over_window_count": len(post_primary_over),
            "within_window_count": len(route_window_analysis) - len(post_primary_over) - len(route_data_review),
            "data_review_count": len(route_data_review),
        },
        "additional_removal": {
            "address_count": len(additional_stop_keys),
            "rider_count": additional_rider_count,
            "selection_strategy": ADDITIONAL_REMOVAL_STRATEGY,
        },
        "final": {
            "over_window_count": len(final_over),
            "within_window_count": len(route_window_analysis) - len(final_over) - len(route_data_review),
            "data_review_count": len(route_data_review),
            "all_measured_routes_within_window": not final_over and not route_data_review,
        },
    }

    category_priority = {
        "direct_over_limit": 0,
        "route_only_over_limit": 1,
        "additional_window_candidate": 2,
        "data_review": 3,
        "within_limit": 4,
    }
    rows.sort(
        key=lambda row: (
            category_priority.get(str(row.get("operational_category") or ""), 9),
            -max(
                (_safe_float(item.get("over_limit_min")) for item in list(row.get("route_contexts") or [])),
                default=0.0,
            ),
            -_safe_float(row.get("direct_duration_min")),
        )
    )
    resolved_rows = [row for row in rows if row.get("provider_status") == "resolved"]
    result = _base_result(
        config=config,
        provider=provider_name,
        school=school,
        rows=[{key: value for key, value in row.items() if key != "_point"} for row in rows],
        route_count=len(route_groups),
        logical_call_estimate=logical_call_estimate,
        progress={
            "phase": "complete",
            "completed": logical_call_estimate,
            "total": logical_call_estimate,
            "provider_api_calls": int(provider.state.get("api_calls", 0)),
            "in_run_reuse_count": int(provider.state.get("cache_hits", 0)),
        },
        errors=errors,
    )
    result["status"] = "partial" if errors else "complete"
    result["completed_at"] = utc_now_iso()
    result["routes"] = sorted(route_results, key=lambda row: str(row.get("route_id") or ""))
    result["route_window_analysis"] = route_window_analysis
    result["operational_conclusion"] = operational_conclusion
    result["summary"].update(
        {
            "resolved_count": len(resolved_rows),
            "failed_count": len(rows) - len(resolved_rows),
            "max_direct_duration_min": round(max((_safe_float(row.get("direct_duration_min")) for row in resolved_rows), default=0.0), 2),
            "max_direct_distance_km": round(max((_safe_float(row.get("direct_distance_km")) for row in resolved_rows), default=0.0), 3),
            "direct_over_limit_address_count": direct_address_count,
            "direct_over_limit_rider_count": direct_rider_count,
            "route_only_over_limit_address_count": route_only_address_count,
            "route_only_over_limit_rider_count": route_only_rider_count,
            "routes_over_window_after_primary_count": len(post_primary_over),
            "additional_removal_address_count": len(additional_stop_keys),
            "additional_removal_rider_count": additional_rider_count,
            "routes_over_window_final_count": len(final_over),
            "route_window_data_review_count": len(route_data_review),
            "route_window_minutes": route_window_min,
            "provider_api_calls": int(provider.state.get("api_calls", 0)),
            "in_run_reuse_count": int(provider.state.get("cache_hits", 0)),
        }
    )
    return result


def aggregate_direct_school_results(records: list[dict[str, Any]]) -> dict[str, Any]:
    samples_by_stop: dict[str, list[dict[str, Any]]] = defaultdict(list)
    run_ids: list[str] = []
    for record in records:
        result = dict(record.get("result") or {})
        if str(record.get("status") or "") != "succeeded" or not result:
            continue
        run_id = str(record.get("job_id") or "")
        run_ids.append(run_id)
        sample_at = result.get("completed_at") or record.get("finished_at") or record.get("scheduled_start_at")
        parameters = dict(result.get("parameters") or {})
        duration_limit = _safe_float(parameters.get("far_duration_minutes"), 45.0)
        for row in list(result.get("stops") or []):
            if row.get("provider_status") != "resolved":
                continue
            category = str(row.get("operational_category") or "")
            if row.get("additional_window_candidate"):
                category = "additional_window_candidate"
            elif category not in {"direct_over_limit", "route_only_over_limit", "within_limit"}:
                if _safe_float(row.get("direct_duration_min")) >= duration_limit:
                    category = "direct_over_limit"
                elif _safe_float(row.get("estimated_current_ride_min")) >= duration_limit:
                    category = "route_only_over_limit"
                else:
                    category = "within_limit"
            samples_by_stop[str(row.get("stop_key") or "")].append(
                {
                    "job_id": run_id,
                    "sample_at": sample_at,
                    "stop_key": row.get("stop_key"),
                    "address": row.get("address"),
                    "direct_duration_min": row.get("direct_duration_min"),
                    "direct_distance_km": row.get("direct_distance_km"),
                    "rider_detour_min": row.get("rider_detour_min"),
                    "operational_category": category,
                }
            )
    rows: list[dict[str, Any]] = []
    all_samples: list[dict[str, Any]] = []
    for stop_key, samples in samples_by_stop.items():
        all_samples.extend(samples)
        durations = [_safe_float(sample.get("direct_duration_min")) for sample in samples]
        distances = [_safe_float(sample.get("direct_distance_km")) for sample in samples]
        direct_over_count = sum(1 for sample in samples if sample.get("operational_category") == "direct_over_limit")
        rows.append(
            {
                "stop_key": stop_key,
                "address": samples[0].get("address"),
                "sample_count": len(samples),
                "duration_median_min": round(statistics.median(durations), 2),
                "duration_p90_min": round(_percentile(durations, 0.9) or 0.0, 2),
                "duration_max_min": round(max(durations), 2),
                "duration_variability_min": round(statistics.pstdev(durations), 2) if len(durations) > 1 else 0.0,
                "distance_median_km": round(statistics.median(distances), 3),
                "direct_over_limit_rate": round(direct_over_count / len(samples), 3),
                "persistent_direct_over_limit": len(samples) >= 2 and direct_over_count / len(samples) >= 0.6,
            }
        )
    rows.sort(key=lambda row: (_safe_float(row.get("direct_over_limit_rate")), _safe_float(row.get("duration_p90_min"))), reverse=True)
    return {
        "run_count": len(set(run_ids)),
        "run_ids": list(dict.fromkeys(run_ids)),
        "stop_count": len(rows),
        "stops": rows,
        "samples": sorted(all_samples, key=lambda row: str(row.get("sample_at") or ""), reverse=True),
    }


def build_direct_school_workbook(
    record: dict[str, Any],
    multi_day: dict[str, Any] | None = None,
) -> bytes:
    result = dict(record.get("result") or {})
    if not result:
        raise ValueError("Direct-to-school analysis result is not available.")
    conclusion = dict(result.get("operational_conclusion") or _legacy_operational_conclusion(result))
    parameters = dict(result.get("parameters") or {})
    metadata = dict(record.get("metadata") or {})
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Operational Summary"
    summary_sheet.merge_cells("A1:F1")
    summary_sheet["A1"] = "Direct-to-School Operational Conclusion / 点到校运营结论"
    summary_sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 28
    summary_sheet.append([
        "Job / 任务", metadata.get("job_name") or record.get("job_id"),
        "School / 学校", dict(result.get("school") or {}).get("address"),
        "Captured / 测算时间", result.get("completed_at"),
    ])
    summary_sheet.append([
        "Provider / 地图服务", result.get("provider"),
        "Direction / 方向", result.get("service_direction"),
        "Status / 状态", result.get("status"),
    ])
    summary_sheet.append([
        "Student trip limit / 学生单程阈值", conclusion.get("duration_limit_min"),
        "Route window / 路线时间窗", conclusion.get("route_window_min"),
        "Unit / 单位", "minutes / 分钟",
    ])
    summary_sheet.append([])
    summary_sheet.append([
        "Step / 步骤", "Operational question / 运营问题", "Students / 学生",
        "Addresses / 地址", "Routes over window / 超窗路线", "Conclusion / 结论",
    ])
    direct = dict(conclusion.get("direct_over_limit") or {})
    route_only = dict(conclusion.get("route_only_over_limit") or {})
    post_primary = dict(conclusion.get("post_primary") or {})
    additional = dict(conclusion.get("additional_removal") or {})
    final = dict(conclusion.get("final") or {})
    summary_sheet.append([
        "1", "Direct trip already exceeds the student trip limit / 直达学校已超过学生单程阈值",
        direct.get("rider_count", 0), direct.get("address_count", 0), "-",
        "Dedicated transport candidate / 优先评估专车",
    ])
    summary_sheet.append([
        "2", "Direct trip is within limit, but current route ride exceeds it / 直达未超限但随路线乘车超限",
        route_only.get("rider_count", 0), route_only.get("address_count", 0), "-",
        "Remove from shared route or redesign route / 摘出或调整路线",
    ])
    summary_sheet.append([
        "3", "After steps 1-2 are removed, do routes still exceed the route window? / 摘出前两类后路线是否仍超窗",
        additional.get("rider_count", 0), additional.get("address_count", 0),
        post_primary.get("over_window_count"),
        _final_conclusion_label(final),
    ])
    summary_sheet.append([
        "Selection rule / 补充摘除顺序",
        "Longest direct trip first; route saving breaks ties / 按直达时间从长到短；同分时优先路线节省更大的站点",
        "-", "-", "-", ADDITIONAL_REMOVAL_STRATEGY,
    ])
    summary_sheet.append([])
    summary_sheet.append(["Parameter / 参数", "Value / 数值", "Meaning / 含义"])
    parameter_rows = [
        ("Trip direction / 测算方向", result.get("service_direction"), "Direction used for live map measurements / 实时地图测算方向"),
        ("Student trip time limit (min) / 学生单程时间上限（分钟）", parameters.get("far_duration_minutes"), "Applied to both direct-trip and current-route rider classification / 用于直达及当前路线乘车分类"),
        ("Route operating window / 路线运行时间窗", f"{parameters.get('time_window_start')} - {parameters.get('time_window_end')}", "Configured route operating interval / 运行时配置的路线运行时间窗"),
        ("Per-stop dwell time (min) / 每站停靠时间（分钟）", parameters.get("stop_service_minutes"), "Added for each service stop / 每个服务站点计入"),
    ]
    for row in parameter_rows:
        summary_sheet.append(list(row))
    _style_summary_sheet(summary_sheet)

    classification_rows = _student_classification_rows(result)
    _write_readable_table(
        workbook.create_sheet("Student Classification"),
        "Student Classification / 学生分类",
        "One row per route occurrence. Student counts therefore remain correct when one address appears on multiple routes. / 每条路线 occurrence 一行，跨路线地址不会误计学生数。",
        [
            "Category / 分类", "Address / 地址", "Route / 路线", "Stop order / 站序",
            "Students / 学生", "Direct trip min / 直达分钟", "Current ride min / 当前乘车分钟",
            "Over limit min / 超限分钟", "Recommended action / 建议动作",
        ],
        classification_rows,
        [28, 42, 12, 12, 12, 18, 20, 18, 34],
    )

    route_rows = _route_outcome_rows(result)
    _write_readable_table(
        workbook.create_sheet("Route Outcomes"),
        "Route Window Outcomes / 路线时间窗复测",
        "Routes are remeasured after the first two student groups are removed. Additional removals are tested from longest to shortest direct trip until the route fits or cannot be resolved. / 摘出前两类学生后实时复测，仍超窗则按直达时间从长到短验证补充摘除。",
        [
            "Route / 路线", "Window min / 窗口", "Original min / 原始", "Primary removed students / 首轮摘出学生",
            "After primary min / 首轮后", "Still over? / 是否仍超", "Additional students / 补充摘出学生",
            "Additional addresses / 补充地址", "Final min / 最终", "Final status / 最终状态",
        ],
        route_rows,
        [12, 14, 16, 22, 18, 16, 22, 48, 14, 24],
    )

    measurement_rows = _address_measurement_rows(result)
    _write_readable_table(
        workbook.create_sheet("Address Measurements"),
        "All Address Measurements / 全部地址测算",
        "Live provider measurements captured for this run. OSRM values are free-flow references only. / 本次任务的实时地图测算；OSRM 仅为自由流参考。",
        [
            "Address / 地址", "Students / 学生", "Routes / 路线", "Operational class / 运营分类",
            "Direct min / 直达分钟", "Direct km / 直达公里", "Current ride min / 当前乘车分钟",
            "Over limit min / 超限分钟", "Additional removal routes / 补充摘站路线", "Captured / 测算时间",
        ],
        measurement_rows,
        [42, 12, 18, 28, 16, 16, 20, 20, 22, 24],
    )

    quality_rows = [
        [item.get("scope"), item.get("route_id"), item.get("address"), item.get("error")]
        for item in list(result.get("errors") or [])
    ]
    _write_readable_table(
        workbook.create_sheet("Data Quality"),
        "Data Quality and Exceptions / 数据质量与异常",
        "Empty means no recorded data-quality exception. / 空表表示本次没有记录到数据质量异常。",
        ["Scope / 环节", "Route / 路线", "Address / 地址", "Issue / 问题"],
        quality_rows,
        [24, 14, 42, 70],
    )

    daily_rows = [
        [
            item.get("sample_at"), item.get("address"), item.get("direct_duration_min"),
            item.get("direct_distance_km"), item.get("rider_detour_min"), _operational_category_label(item.get("operational_category")),
        ]
        for item in list(dict(multi_day or {}).get("samples") or [])
    ]
    _write_readable_table(
        workbook.create_sheet("Daily History"),
        "Daily Measurement History / 多日测算历史",
        "Repeated scheduled measurements for the same address. / 同一地址的多次定时测算记录。",
        ["Captured / 测算时间", "Address / 地址", "Direct min / 直达分钟", "Direct km / 直达公里", "Current minus direct min / 当前减直达分钟", "Operational class / 运营分类"],
        daily_rows,
        [24, 42, 18, 18, 18, 30],
    )
    from io import BytesIO

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _legacy_operational_conclusion(result: dict[str, Any]) -> dict[str, Any]:
    duration_limit = _safe_float(dict(result.get("parameters") or {}).get("far_duration_minutes"), 45.0)
    direct_rows: set[str] = set()
    route_rows: set[str] = set()
    direct_riders = 0
    route_riders = 0
    for row in list(result.get("stops") or []):
        contexts = list(row.get("route_contexts") or [])
        direct_over = _safe_float(row.get("direct_duration_min")) >= duration_limit
        if direct_over:
            direct_rows.add(str(row.get("stop_key") or row.get("address") or ""))
            direct_riders += sum(_safe_int(item.get("riders")) for item in contexts) or _safe_int(row.get("riders"))
        else:
            matching = [item for item in contexts if _safe_float(item.get("estimated_current_ride_min")) >= duration_limit]
            if matching:
                route_rows.add(str(row.get("stop_key") or row.get("address") or ""))
                route_riders += sum(_safe_int(item.get("riders")) for item in matching) or _safe_int(row.get("riders"))
    return {
        "duration_limit_min": duration_limit,
        "route_window_min": _window_duration_minutes(dict(result.get("parameters") or {})),
        "direct_over_limit": {"address_count": len(direct_rows), "rider_count": direct_riders},
        "route_only_over_limit": {"address_count": len(route_rows), "rider_count": route_riders},
        "post_primary": {"over_window_count": None, "data_review_count": len(list(result.get("routes") or []))},
        "additional_removal": {"address_count": 0, "rider_count": 0},
        "final": {"data_review_count": len(list(result.get("routes") or [])), "all_measured_routes_within_window": False},
        "legacy_result": True,
    }


def _final_conclusion_label(final: dict[str, Any]) -> str:
    if _safe_int(final.get("data_review_count")):
        return "Rerun required for route recovery evidence / 需重跑以生成路线摘除证据"
    if bool(final.get("all_measured_routes_within_window")):
        return "All measured routes fit the window / 所有已测路线均符合时间窗"
    return f"{_safe_int(final.get('over_window_count'))} route(s) still exceed the window / 仍有路线超窗"


def _student_classification_rows(result: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    category_labels = {
        "direct_over_limit": "1. Direct trip over limit / 直达已超限",
        "route_only_over_limit": "2. Route ride only over limit / 仅路线乘车超限",
    }
    for stop in list(result.get("stops") or []):
        direct_duration = stop.get("direct_duration_min")
        for context in list(stop.get("route_contexts") or []):
            category = str(context.get("operational_category") or "")
            if category not in category_labels:
                continue
            rows.append([
                category_labels[category], stop.get("address"), context.get("route_id"), context.get("stop_sequence"),
                context.get("riders"), direct_duration, context.get("estimated_current_ride_min"),
                context.get("over_limit_min"),
                "Dedicated transport review / 评估专车" if category == "direct_over_limit" else "Remove from shared route or redesign / 摘出或调整路线",
            ])
    for route in list(result.get("route_window_analysis") or []):
        for item in list(route.get("additional_removals") or []):
            rows.append([
                "3. Additional route-window removal / 补充摘出",
                item.get("address"), route.get("route_id"), item.get("stop_sequence"), item.get("riders"),
                item.get("direct_duration_min"), item.get("current_ride_min"), item.get("estimated_route_saving_min"),
                "Additional dedicated transport or route split / 补充专车或拆分路线",
            ])
    return rows


def _route_outcome_rows(result: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    route_results = list(result.get("route_window_analysis") or [])
    route_results.sort(
        key=lambda route: (
            0 if route.get("status") == "still_over_window" else
            1 if route.get("status") == "data_review" else
            2 if _safe_int(route.get("additional_removed_riders")) > 0 else
            3 if _safe_int(route.get("primary_removed_riders")) > 0 else 4,
            str(route.get("route_id") or ""),
        )
    )
    for route in route_results:
        additional = list(route.get("additional_removals") or [])
        rows.append([
            route.get("route_id"), route.get("window_limit_min"), route.get("original_duration_min"),
            route.get("primary_removed_riders"), route.get("post_primary_duration_min"),
            "Yes / 是" if _safe_float(route.get("post_primary_duration_min")) > _safe_float(route.get("window_limit_min")) else "No / 否",
            route.get("additional_removed_riders"),
            "; ".join(str(item.get("address") or "") for item in additional),
            route.get("final_duration_min"), route.get("status"),
        ])
    if not rows:
        rows.append(["Legacy result / 旧结果", None, None, None, None, None, None, None, None, "Rerun required / 需重跑"])
    return rows


def _address_measurement_rows(result: dict[str, Any]) -> list[list[Any]]:
    duration_limit = _safe_float(dict(result.get("parameters") or {}).get("far_duration_minutes"), 45.0)
    rows: list[list[Any]] = []
    for row in list(result.get("stops") or []):
        category = str(row.get("operational_category") or "data_review")
        direct_duration = _safe_float(row.get("direct_duration_min"))
        current_ride = _safe_float(row.get("estimated_current_ride_min"))
        over_limit = max(0.0, max(direct_duration, current_ride) - duration_limit)
        rows.append([
            row.get("address"), row.get("riders"), ", ".join(str(item) for item in list(row.get("route_ids") or [])),
            _operational_category_label(category), row.get("direct_duration_min"), row.get("direct_distance_km"),
            row.get("estimated_current_ride_min"), round(over_limit, 2),
            ", ".join(str(item) for item in list(row.get("additional_window_routes") or [])), row.get("provider_called_at"),
        ])
    return rows


def _operational_category_label(value: Any) -> str:
    return {
        "direct_over_limit": "Direct trip over limit / 直达超时",
        "route_only_over_limit": "Current route only over limit / 仅当前路线超时",
        "additional_window_candidate": "Additional removal candidate / 补充摘站候选",
        "within_limit": "Within limit / 未超时",
        "data_review": "Data review / 数据复核",
    }.get(str(value or ""), "Legacy result / 旧结果")


def _style_summary_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="DDE9E8")
    for row_index in (6, 12):
        for cell in sheet[row_index]:
            if cell.value is not None:
                cell.font = Font(bold=True, color="164E63")
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in enumerate((18, 58, 18, 22, 22, 42), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A6"


def _write_readable_table(
    sheet: Any,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[float],
) -> None:
    last_column = get_column_letter(len(headers))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color="475569")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 34
    sheet.append([])
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="DDE9E8")
    border = Border(bottom=Side(style="thin", color="94A3B8"))
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="164E63")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{last_column}{max(4, sheet.max_row)}"


def _append_rows(sheet: Any, rows: list[dict[str, Any]], columns: list[str]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([_excel_value(row.get(column)) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)
