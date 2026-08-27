from __future__ import annotations

import base64
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "backend"))

import backend_service  # noqa: E402
import direct_school_analysis as analysis  # noqa: E402


def point(address: str, lat: float, lng: float, passengers: int = 0) -> dict:
    return {
        "country": "China",
        "city": "Shanghai",
        "address": address,
        "display_address": address,
        "original_members": [address],
        "lat": lat,
        "lng": lng,
        "plot_lat": lat,
        "plot_lng": lng,
        "passenger_count": passengers,
    }


def prepared_payload() -> dict:
    school = point("School", 31.20, 121.50)
    far = point("Far stop", 31.80, 121.10, 2)
    near = point("Near stop", 31.24, 121.48, 3)
    return {
        "input_records": [
            {"country": "China", "city": "Shanghai", "address": "School", "passenger_count": 0},
            {"country": "China", "city": "Shanghai", "address": "Far stop", "passenger_count": 2},
            {"country": "China", "city": "Shanghai", "address": "Near stop", "passenger_count": 3},
        ],
        "original_points": [school, far, near],
        "current_plan": {
            "service_direction": "To School",
            "stops": [
                {"route_id": "R1", "stop_sequence": 1, "country": "China", "city": "Shanghai", "address": "Far stop", "passenger_count": 2, "is_depot": False},
                {"route_id": "R1", "stop_sequence": 2, "country": "China", "city": "Shanghai", "address": "Near stop", "passenger_count": 3, "is_depot": False},
                {"route_id": "R1", "stop_sequence": 3, "country": "China", "city": "Shanghai", "address": "School", "passenger_count": 0, "is_depot": True},
            ],
        },
    }


class FakeProvider:
    def __init__(self, provider: str, **_kwargs) -> None:
        self.provider = provider
        self.state = {"api_calls": 0, "cache_hits": 0}

    def route(self, points: list[dict]) -> dict:
        self.state["api_calls"] += 1
        addresses = [str(item.get("address") or item.get("display_address")) for item in points]
        if len(points) == 3:
            duration_s, distance_m = 5400.0, 40000.0
        elif addresses[0] == "Far stop":
            duration_s, distance_m = 3600.0, 30000.0
        else:
            duration_s, distance_m = 900.0, 5000.0
        return {
            "duration_s": duration_s,
            "distance_m": distance_m,
            "called_at": "2026-08-25T00:00:00+00:00",
        }


def fake_osrm(origin: dict, destination: dict, _cache: dict) -> dict:
    far = "Far stop" in {origin.get("address"), destination.get("address")}
    return {
        "duration_s": 3000.0 if far else 720.0,
        "distance_m": 28000.0 if far else 4500.0,
        "geometry": [
            [float(origin["lng"]), float(origin["lat"])],
            [float(destination["lng"]), float(destination["lat"])],
        ],
    }


def test_osrm_leg_uses_audit_plot_coordinates_and_snap_connectors(monkeypatch) -> None:
    origin = {
        **point("Origin", 31.2400, 121.4800),
        "plot_lat": 31.2421,
        "plot_lng": 121.4754,
    }
    destination = {
        **point("School", 31.2000, 121.5000),
        "plot_lat": 31.2022,
        "plot_lng": 121.4953,
    }

    def fake_leg_details(points: list[dict]) -> list[dict]:
        assert points[0]["lat"] == origin["plot_lat"]
        assert points[0]["lng"] == origin["plot_lng"]
        assert points[1]["lat"] == destination["plot_lat"]
        assert points[1]["lng"] == destination["plot_lng"]
        return [
            {
                "duration_s": 600,
                "distance_m": 5000,
                "geometry": [(31.2420, 121.4755), (31.2023, 121.4952)],
                "snap_connectors": [
                    {
                        "type": "origin",
                        "distance_m": 32,
                        "geometry": [(31.2421, 121.4754), (31.2420, 121.4755)],
                    }
                ],
            }
        ]

    monkeypatch.setattr(
        analysis.distance_tool,
        "compute_osrm_route_leg_details",
        fake_leg_details,
    )

    result = analysis._osrm_leg(origin, destination, {})

    assert result["coordinate_source"] == "plot_wgs84"
    assert result["geometry"] == [[121.4755, 31.2420], [121.4952, 31.2023]]
    assert result["snap_connectors"] == [
        {
            "type": "origin",
            "distance_m": 32.0,
            "geometry": [[121.4754, 31.2421], [121.4755, 31.2420]],
        }
    ]


def test_analysis_builds_three_step_operational_conclusion(monkeypatch) -> None:
    monkeypatch.setattr(analysis, "FreshRouteProvider", FakeProvider)
    monkeypatch.setattr(analysis, "_osrm_leg", fake_osrm)
    checkpoints: list[dict] = []

    result = analysis.run_direct_school_analysis(
        prepared_payload(),
        {
            "service_direction": "To School",
            "far_duration_minutes": 45,
        },
        run_seed="test",
        checkpoint=lambda payload: checkpoints.append(payload),
    )

    assert result["status"] == "complete"
    assert result["analysis_version"] == 4
    assert result["summary"]["address_count"] == 2
    assert result["summary"]["provider_api_calls"] == 4
    far = next(row for row in result["stops"] if row["address"] == "Far stop")
    near = next(row for row in result["stops"] if row["address"] == "Near stop")
    assert far["operational_category"] == "direct_over_limit"
    assert near["direct_duration_min"] == 15
    assert result["operational_conclusion"]["direct_over_limit"]["rider_count"] == 2
    assert result["operational_conclusion"]["route_only_over_limit"]["rider_count"] == 0
    assert result["operational_conclusion"]["final"]["all_measured_routes_within_window"] is True
    assert result["route_window_analysis"][0]["post_primary_duration_min"] == 16
    assert checkpoints
    assert checkpoints[0]["analysis_type"] == "direct_school"


def test_distance_and_retired_parameters_never_trigger_classification(monkeypatch) -> None:
    monkeypatch.setattr(analysis, "FreshRouteProvider", FakeProvider)
    monkeypatch.setattr(analysis, "_osrm_leg", fake_osrm)

    result = analysis.run_direct_school_analysis(
        prepared_payload(),
        {
            "service_direction": "To School",
            "far_distance_km": 1,
            "far_duration_minutes": 120,
            "burden_minutes": 200,
            "bypass_candidate_limit": 50,
            "candidate_cluster_radius_km": 100,
            "time_window_start": "06:30",
            "time_window_end": "09:30",
        },
        run_seed="time-only-threshold-test",
    )

    far = next(row for row in result["stops"] if row["address"] == "Far stop")
    assert far["direct_distance_km"] == 30
    assert far["direct_duration_min"] == 60
    assert far["operational_category"] == "within_limit"
    for retired_key in (
        "far_distance_km",
        "burden_minutes",
        "bypass_candidate_limit",
        "candidate_cluster_radius_km",
    ):
        assert retired_key not in result["parameters"]


def test_submit_config_drops_retired_candidate_scoring_parameters() -> None:
    config = backend_service._direct_school_analysis_config(
        {
            "analysis_config": {
                "far_duration_minutes": 60,
                "burden_minutes": 20,
                "bypass_candidate_limit": 25,
                "candidate_cluster_radius_km": 5,
            }
        },
        {"service_direction": "To School"},
    )

    assert config["far_duration_minutes"] == 60
    assert config["service_direction"] == "To School"
    assert not {"burden_minutes", "bypass_candidate_limit", "candidate_cluster_radius_km"} & config.keys()

    afternoon = backend_service._direct_school_analysis_config(
        {"analysis_config": {"time_window_start": "06:30", "time_window_end": "08:00"}},
        {"service_direction": "From School"},
    )
    assert afternoon["time_window_start"] == "15:40"
    assert afternoon["time_window_end"] == "17:40"


def test_analysis_counts_route_only_students_by_occurrence(monkeypatch) -> None:
    monkeypatch.setattr(analysis, "FreshRouteProvider", FakeProvider)
    monkeypatch.setattr(analysis, "_osrm_leg", fake_osrm)

    result = analysis.run_direct_school_analysis(
        prepared_payload(),
        {
            "service_direction": "To School",
            "far_duration_minutes": 16,
        },
        run_seed="occurrence-test",
    )

    conclusion = result["operational_conclusion"]
    assert conclusion["direct_over_limit"] == {"address_count": 1, "rider_count": 2}
    assert conclusion["route_only_over_limit"] == {"address_count": 1, "rider_count": 3}
    near = next(row for row in result["stops"] if row["address"] == "Near stop")
    assert near["route_contexts"][0]["operational_category"] == "route_only_over_limit"


def test_analysis_recommends_additional_removal_until_route_fits(monkeypatch) -> None:
    monkeypatch.setattr(analysis, "FreshRouteProvider", FakeProvider)
    monkeypatch.setattr(analysis, "_osrm_leg", fake_osrm)

    result = analysis.run_direct_school_analysis(
        prepared_payload(),
        {
            "service_direction": "To School",
            "far_duration_minutes": 120,
            "time_window_start": "06:30",
            "time_window_end": "07:00",
        },
        run_seed="recovery-test",
    )

    conclusion = result["operational_conclusion"]
    route = result["route_window_analysis"][0]
    assert conclusion["primary_removal"]["rider_count"] == 0
    assert conclusion["post_primary"]["over_window_count"] == 1
    assert conclusion["additional_removal"]["rider_count"] == 2
    assert route["additional_removals"][0]["address"] == "Far stop"
    assert route["additional_removals"][0]["selection_rank"] == 1
    assert route["additional_removals"][0]["selection_basis"] == analysis.ADDITIONAL_REMOVAL_STRATEGY
    assert route["additional_removal_strategy"] == analysis.ADDITIONAL_REMOVAL_STRATEGY
    assert route["final_duration_min"] == 16
    assert route["status"] == "within_window"


def test_additional_removal_ranking_prefers_longest_direct_trip() -> None:
    long_trip_rank = analysis._additional_removal_rank(
        {"direct_duration_min": 55},
        estimated_saving_min=4,
        riders=5,
        stop_sequence=8,
    )
    short_trip_rank = analysis._additional_removal_rank(
        {"direct_duration_min": 25},
        estimated_saving_min=20,
        riders=1,
        stop_sequence=2,
    )

    assert long_trip_rank > short_trip_rank


def test_multi_day_aggregation_requires_repeated_evidence() -> None:
    records = []
    for index, duration in enumerate((50.0, 55.0, 40.0), start=1):
        records.append(
            {
                "job_id": f"job-{index}",
                "status": "succeeded",
                "finished_at": f"2026-08-{20 + index}T00:00:00+00:00",
                "result": {
                    "stops": [
                        {
                            "stop_key": "far",
                            "address": "Far stop",
                            "provider_status": "resolved",
                            "direct_duration_min": duration,
                            "direct_distance_km": 30,
                        }
                    ]
                },
            }
        )

    aggregate = analysis.aggregate_direct_school_results(records)

    assert aggregate["run_count"] == 3
    assert aggregate["stops"][0]["duration_median_min"] == 50
    assert aggregate["stops"][0]["persistent_direct_over_limit"] is True
    assert aggregate["stops"][0]["direct_over_limit_rate"] == 0.667


def test_excel_export_contains_required_analysis_sheets() -> None:
    record = {
        "job_id": "job-1",
        "metadata": {"job_name": "Direct test"},
        "result": {
            "status": "complete",
            "provider": "amap",
            "service_direction": "To School",
            "school": {"address": "School"},
            "summary": {"address_count": 1},
            "parameters": {"far_duration_minutes": 45},
            "stops": [
                {
                    "stop_key": "far",
                    "address": "Far stop",
                    "operational_category": "direct_over_limit",
                    "provider_status": "resolved",
                }
            ],
            "errors": [],
        },
    }

    body = analysis.build_direct_school_workbook(
        record,
        {"samples": [{"job_id": "job-1", "stop_key": "far", "address": "Far stop"}]},
    )
    workbook = load_workbook(BytesIO(body))

    assert workbook.sheetnames == [
        "Operational Summary",
        "Student Classification",
        "Route Outcomes",
        "Address Measurements",
        "Data Quality",
        "Daily History",
    ]
    assert "Operational Conclusion" in workbook["Operational Summary"]["A1"].value
    assert workbook["Student Classification"]["A4"].value == "Category / 分类"
    assert workbook["Route Outcomes"]["J4"].value == "Final status / 最终状态"
    assert workbook["Operational Summary"].freeze_panes == "A6"
    assert workbook["Student Classification"].freeze_panes == "A5"
    assert workbook["Student Classification"].column_dimensions["B"].width >= 40
    assert workbook["Address Measurements"].auto_filter.ref.startswith("A4:")


def test_audit_template_is_the_direct_school_input_contract() -> None:
    workbook_bytes = backend_service.build_excel_template_bytes()

    _client_core, source_label, current_plan = backend_service._read_current_plan_upload(
        {
            "file_name": "audit-template.xlsx",
            "file_base64": base64.b64encode(workbook_bytes).decode("ascii"),
        }
    )

    assert source_label == "audit-template.xlsx"
    assert current_plan["summary"]["route_count"] == 2
    assert current_plan["summary"]["service_stop_count"] == 3
    assert current_plan["fleet"]
    assert current_plan["input_records"][0]["passenger_count"] == 0


def test_missing_school_geocode_is_not_replaced_by_first_student_point() -> None:
    current_plan = {
        "input_records": [
            {"country": "China", "city": "Shanghai", "address": "School"},
            {"country": "China", "city": "Shanghai", "address": "Student A"},
            {"country": "China", "city": "Shanghai", "address": "Student B"},
        ]
    }
    prepared = {
        "original_points": [
            {**point("Student A", 31.2, 121.4), "is_depot": True},
            point("Student B", 31.3, 121.5),
        ]
    }

    with pytest.raises(ValueError, match="school address could not be geocoded"):
        backend_service._direct_school_school_point(current_plan, prepared)
