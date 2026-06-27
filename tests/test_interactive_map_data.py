import importlib
import io
import os
import unittest

from openpyxl import load_workbook


class InteractiveMapDataTests(unittest.TestCase):
    def setUp(self) -> None:
        self.service = importlib.import_module("backend_service")

    def test_build_job_map_data_from_structured_result(self) -> None:
        job_record = {
            "job_id": "job-1",
            "result": {
                "service_direction": "From School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "bus_capacity": 42,
                                "comfort_capacity": 35,
                                "nodes": [0, 1],
                                "time_s": 600,
                                "distance_m": 1200,
                                "leg_details": [
                                    {
                                        "from_node": 0,
                                        "to_node": 1,
                                        "duration_s": 600,
                                        "distance_m": 1200,
                                        "geometry": [[31.2, 121.4], [31.21, 121.41]],
                                    }
                                ],
                            }
                        ],
                    }
                },
            },
        }

        payload, error = self.service._build_job_map_data(job_record, "current_plan")

        self.assertIsNone(error)
        self.assertIsNotNone(payload)
        assert payload is not None
        self.assertEqual(payload["scenario_key"], "current_plan")
        self.assertEqual(len(payload["routes"]), 1)
        self.assertEqual(len(payload["stops"]), 2)
        self.assertEqual(payload["routes"][0]["geometry"], [[121.4, 31.2], [121.41, 31.21]])
        self.assertEqual(payload["stops"][1]["address"], "Stop A")
        self.assertEqual(payload["bounds"]["min_lng"], 121.4)

    def test_china_map_payload_adds_amap_display_geometry_and_duration(self) -> None:
        old_key = os.environ.get("AMAP_API_KEY")
        old_enabled = self.service.AMAP_DISPLAY_GEOMETRY_ENABLED
        old_func = self.service._amap_display_geometry_for_route

        def restore() -> None:
            if old_key is None:
                os.environ.pop("AMAP_API_KEY", None)
            else:
                os.environ["AMAP_API_KEY"] = old_key
            self.service.AMAP_DISPLAY_GEOMETRY_ENABLED = old_enabled
            self.service._amap_display_geometry_for_route = old_func

        self.addCleanup(restore)
        os.environ["AMAP_API_KEY"] = "unit-test-key"
        self.service.AMAP_DISPLAY_GEOMETRY_ENABLED = True
        self.service._amap_display_geometry_for_route = (
            lambda _points, _nodes: (
                [[121.4001, 31.2001], [121.4101, 31.2101]],
                "amap_cn",
                "",
                700,
                1300,
            )
        )
        job_record = {
            "job_id": "job-cn-display",
            "config": {"country": "China", "city": "Shanghai"},
            "result": {
                "service_direction": "From School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "lat": 31.2019,
                                "lng": 121.4044,
                                "provider": "amap",
                                "adcode": "310000",
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "lat": 31.2119,
                                "lng": 121.4144,
                                "provider": "amap",
                                "adcode": "310000",
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "bus_capacity": 42,
                                "nodes": [0, 1],
                                "time_s": 600,
                                "traffic_adjusted_drive_time_s": 900,
                                "distance_m": 1200,
                                "leg_details": [
                                    {
                                        "duration_s": 600,
                                        "distance_m": 1200,
                                        "geometry": [[31.2, 121.4], [31.21, 121.41]],
                                    }
                                ],
                            }
                        ],
                    }
                },
            },
        }

        payload, error = self.service._build_job_map_data(job_record, "current_plan")

        self.assertIsNone(error)
        assert payload is not None
        route = payload["routes"][0]
        self.assertEqual(route["geometry"], [[121.4, 31.2], [121.41, 31.21]])
        self.assertEqual(
            route["display_geometry"], [[121.4001, 31.2001], [121.4101, 31.2101]]
        )
        self.assertEqual(route["display_geometry_source"], "amap_cn")
        self.assertEqual(route["display_duration_s"], 700)
        self.assertEqual(route["display_distance_m"], 1300)
        self.assertEqual(route["duration_s"], 700)
        self.assertEqual(route["raw_duration_s"], 600)
        self.assertEqual(route["distance_m"], 1200)
        self.assertEqual(payload["bounds"]["max_lng"], 121.4101)

    def test_non_china_map_payload_does_not_request_amap_display_geometry(self) -> None:
        old_key = os.environ.get("AMAP_API_KEY")
        old_enabled = self.service.AMAP_DISPLAY_GEOMETRY_ENABLED
        old_func = self.service._amap_display_geometry_for_route

        def fail_if_called(_points, _nodes):
            raise AssertionError("AMap display geometry should not be requested")

        def restore() -> None:
            if old_key is None:
                os.environ.pop("AMAP_API_KEY", None)
            else:
                os.environ["AMAP_API_KEY"] = old_key
            self.service.AMAP_DISPLAY_GEOMETRY_ENABLED = old_enabled
            self.service._amap_display_geometry_for_route = old_func

        self.addCleanup(restore)
        os.environ["AMAP_API_KEY"] = "unit-test-key"
        self.service.AMAP_DISPLAY_GEOMETRY_ENABLED = True
        self.service._amap_display_geometry_for_route = fail_if_called
        job_record = {
            "job_id": "job-kr-display",
            "config": {"country": "South Korea", "city": "Seoul"},
            "result": {
                "service_direction": "From School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 37.55,
                                "plot_lng": 126.98,
                                "provider": "google",
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 37.56,
                                "plot_lng": 126.99,
                                "provider": "google",
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "nodes": [0, 1],
                                "time_s": 600,
                                "distance_m": 1200,
                                "leg_details": [
                                    {
                                        "duration_s": 600,
                                        "distance_m": 1200,
                                        "geometry": [[37.55, 126.98], [37.56, 126.99]],
                                    }
                                ],
                            }
                        ],
                    }
                },
            },
        }

        payload, error = self.service._build_job_map_data(job_record, "current_plan")

        self.assertIsNone(error)
        assert payload is not None
        route = payload["routes"][0]
        self.assertEqual(route["geometry"], [[126.98, 37.55], [126.99, 37.56]])
        self.assertIsNone(route["display_geometry"])
        self.assertEqual(route["display_geometry_source"], "osrm")

    def test_from_school_time_impact_model_includes_weighted_review_fields(self) -> None:
        job_record = {
            "job_id": "job-impact-pm",
            "config": {
                "from_school_departure_time": "15:40",
                "stop_service_minutes": 0,
            },
            "result": {
                "service_direction": "From School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "nodes": [0, 1],
                                "time_s": 600,
                                "distance_m": 1200,
                                "leg_details": [
                                    {
                                        "duration_s": 600,
                                        "distance_m": 1200,
                                        "geometry": [[31.2, 121.4], [31.21, 121.41]],
                                    }
                                ],
                            }
                        ],
                    },
                    "original": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "route_id": "R6",
                                "vehicle_id": 6,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "nodes": [0, 1],
                                "time_s": 1500,
                                "distance_m": 2200,
                                "leg_details": [
                                    {
                                        "duration_s": 1500,
                                        "distance_m": 2200,
                                        "geometry": [[31.2, 121.4], [31.21, 121.41]],
                                    }
                                ],
                            }
                        ],
                    },
                },
            },
        }

        payload, error = self.service._build_job_map_data(job_record, "original")

        self.assertIsNone(error)
        assert payload is not None
        stop_impact = payload["stops"][1]["time_impact"]
        self.assertTrue(stop_impact["comparison_available"])
        self.assertEqual(stop_impact["comparison_status"], "matched")
        self.assertEqual(stop_impact["time_role"], "dropoff")
        self.assertEqual(stop_impact["current_route_id"], "Bus 1")
        self.assertEqual(stop_impact["new_route_id"], "R6")
        self.assertEqual(stop_impact["current_time_label"], "15:50")
        self.assertEqual(stop_impact["new_time_label"], "16:05")
        self.assertEqual(stop_impact["delta_minutes"], 15)
        self.assertEqual(stop_impact["adverse_delta_minutes"], 15)
        self.assertEqual(stop_impact["acceptance_threshold_minutes"], 15)
        self.assertTrue(stop_impact["within_acceptance"])
        self.assertEqual(stop_impact["acceptance_status"], "within")
        self.assertEqual(stop_impact["over_acceptance_minutes"], 0)
        self.assertEqual(stop_impact["absolute_delta_minutes"], 15)
        self.assertEqual(stop_impact["impact_direction"], "worse")
        self.assertEqual(stop_impact["change_direction"], "later")
        self.assertEqual(stop_impact["affected_rider_count"], 3)
        self.assertEqual(stop_impact["adverse_rider_minutes"], 45)
        self.assertEqual(stop_impact["matched_key"], "node:1")
        self.assertTrue(stop_impact["route_changed"])

        summary = payload["summary"]["time_impact"]
        self.assertTrue(summary["available"])
        self.assertEqual(summary["compared_stop_count"], 1)
        self.assertEqual(summary["compared_rider_count"], 3)
        self.assertEqual(summary["worse_stop_count"], 1)
        self.assertEqual(summary["worse_rider_count"], 3)
        self.assertEqual(summary["acceptance_threshold_minutes"], 15)
        self.assertEqual(summary["within_acceptance_stop_count"], 1)
        self.assertEqual(summary["within_acceptance_rider_count"], 3)
        self.assertEqual(summary["over_acceptance_stop_count"], 0)
        self.assertEqual(summary["over_acceptance_rider_count"], 0)
        self.assertEqual(summary["acceptance_rider_ratio"], 1)
        self.assertEqual(summary["weighted_avg_adverse_delta_minutes"], 15)
        self.assertEqual(summary["total_adverse_rider_minutes"], 45)
        self.assertEqual(summary["route_changed_rider_count"], 3)
        self.assertEqual(summary["top_impacted_stops"][0]["address"], "Stop A")
        self.assertEqual(payload["routes"][0]["time_impact"]["worse_stop_count"], 1)

    def test_time_impact_excel_export_contains_review_sheets(self) -> None:
        job_record = {
            "job_id": "job-impact-export",
            "config": {
                "from_school_departure_time": "15:40",
                "stop_service_minutes": 0,
            },
            "result": {
                "service_direction": "From School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "nodes": [0, 1],
                                "time_s": 600,
                                "distance_m": 1200,
                                "leg_details": [{"duration_s": 600, "distance_m": 1200}],
                            }
                        ],
                    },
                    "original": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop A",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 3,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "route_id": "R6",
                                "vehicle_id": 6,
                                "bus_type_name": "Large Bus",
                                "load": 3,
                                "nodes": [0, 1],
                                "time_s": 1500,
                                "distance_m": 2200,
                                "leg_details": [{"duration_s": 1500, "distance_m": 2200}],
                            }
                        ],
                    },
                },
            },
        }

        workbook_bytes, error = self.service._build_time_impact_workbook_export(
            job_record, "original"
        )

        self.assertIsNone(error)
        self.assertIsNotNone(workbook_bytes)
        assert workbook_bytes is not None
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary", "Routes", "Stops"])
        self.assertEqual(workbook["Summary"]["A1"].value, "Metric")
        self.assertEqual(workbook["Routes"]["A1"].value, "Scenario")
        self.assertEqual(workbook["Stops"]["D2"].value, "Stop A")
        summary_rows = {
            row[0]: row[1]
            for row in workbook["Summary"].iter_rows(min_row=2, values_only=True)
        }
        self.assertEqual(summary_rows["Acceptance threshold minutes"], 15)
        self.assertEqual(summary_rows["Within-threshold riders"], 3)
        self.assertEqual(summary_rows["Over-threshold riders"], 0)

    def test_to_school_time_impact_treats_earlier_pickup_as_adverse(self) -> None:
        job_record = {
            "job_id": "job-impact-am",
            "config": {
                "to_school_arrival_time": "08:00",
                "stop_service_minutes": 0,
            },
            "result": {
                "service_direction": "To School",
                "structured_results": {
                    "current_plan": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop B",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 2,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "vehicle_id": 1,
                                "bus_type_name": "Large Bus",
                                "load": 2,
                                "nodes": [1, 0],
                                "time_s": 1200,
                                "distance_m": 1200,
                                "leg_details": [
                                    {
                                        "duration_s": 1200,
                                        "distance_m": 1200,
                                        "geometry": [[31.21, 121.41], [31.2, 121.4]],
                                    }
                                ],
                            }
                        ],
                    },
                    "original": {
                        "points": [
                            {
                                "address": "School",
                                "plot_lat": 31.2,
                                "plot_lng": 121.4,
                                "passenger_count": 0,
                                "is_depot": True,
                            },
                            {
                                "address": "Stop B",
                                "plot_lat": 31.21,
                                "plot_lng": 121.41,
                                "passenger_count": 2,
                                "is_depot": False,
                            },
                        ],
                        "routes": [
                            {
                                "route_id": "R2",
                                "vehicle_id": 2,
                                "bus_type_name": "Large Bus",
                                "load": 2,
                                "nodes": [1, 0],
                                "time_s": 2400,
                                "distance_m": 2200,
                                "leg_details": [
                                    {
                                        "duration_s": 2400,
                                        "distance_m": 2200,
                                        "geometry": [[31.21, 121.41], [31.2, 121.4]],
                                    }
                                ],
                            }
                        ],
                    },
                },
            },
        }

        payload, error = self.service._build_job_map_data(job_record, "original")

        self.assertIsNone(error)
        assert payload is not None
        stop_impact = payload["stops"][0]["time_impact"]
        self.assertEqual(stop_impact["time_role"], "pickup")
        self.assertEqual(stop_impact["current_time_label"], "07:39")
        self.assertEqual(stop_impact["new_time_label"], "07:19")
        self.assertEqual(stop_impact["delta_minutes"], -20)
        self.assertEqual(stop_impact["adverse_delta_minutes"], 20)
        self.assertEqual(stop_impact["acceptance_threshold_minutes"], 15)
        self.assertFalse(stop_impact["within_acceptance"])
        self.assertEqual(stop_impact["acceptance_status"], "over")
        self.assertEqual(stop_impact["over_acceptance_minutes"], 5)
        self.assertEqual(stop_impact["adverse_direction"], "earlier_pickup")
        self.assertEqual(stop_impact["impact_direction"], "worse")
        self.assertEqual(stop_impact["affected_rider_count"], 2)
        self.assertEqual(stop_impact["adverse_rider_minutes"], 40)
        summary = payload["summary"]["time_impact"]
        self.assertEqual(summary["over_acceptance_stop_count"], 1)
        self.assertEqual(summary["over_acceptance_rider_count"], 2)
        self.assertEqual(summary["max_over_acceptance_delta_minutes"], 5)


if __name__ == "__main__":
    unittest.main()
