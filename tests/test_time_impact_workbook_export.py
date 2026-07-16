import importlib
import io
import sys
from pathlib import Path

from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "apps" / "backend"))
backend_service = importlib.import_module("backend_service")
planner_core = importlib.import_module("planner_core")


def test_time_impact_workbook_includes_am_window_gate(monkeypatch):
    def fake_map_payload(_job_record, _scenario_key, _map_key, *, attach_impact=False):
        return {
            "job_id": "job-1",
            "scenario_name": "15-Minute Constrained",
            "service_direction": "To School",
            "summary": {
                "time_impact": {"available": True},
                "am_arrival_gate": {
                    "status": "failed",
                    "target_arrival_label": "08:00",
                    "checked_route_count": 1,
                    "failed_route_count": 1,
                    "unavailable_route_count": 0,
                    "max_time_window_overrun_minutes": 12,
                },
            },
            "routes": [
                {
                    "id": "Bus 1",
                    "bus_type_name": "18-fbus",
                    "load": 10,
                    "stop_count": 3,
                    "time_impact": {},
                    "am_arrival_gate": {
                        "status": "failed",
                        "verified_departure_label": "06:00",
                        "verified_arrival_label": "08:12",
                        "time_window_overrun_minutes": 12,
                    },
                }
            ],
            "stops": [],
        }, None

    monkeypatch.setattr(backend_service, "_build_job_map_payload", fake_map_payload)

    content, error = backend_service._build_time_impact_workbook_export({}, "time_constrained")

    assert error is None
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    route_rows = list(workbook["Routes"].iter_rows(values_only=True))
    assert ("AM window status", "failed") in summary_rows
    assert "AM window status" in route_rows[0]
    assert route_rows[1][-4:] == ("failed", "06:00", "08:12", 12)


def test_scenario_template_workbook_includes_am_window_gate():
    content = planner_core.build_baseline_template_workbook_bytes(
        {
            "points": [
                {"node_id": 0, "address": "School", "passenger_count": 0},
                {"node_id": 1, "address": "Stop A", "passenger_count": 3},
            ],
            "routes": [
                {
                    "route_id": "Bus 1",
                    "display_route_id": "Opt Bus 1",
                    "bus_type_name": "18-fbus",
                    "bus_capacity": 18,
                    "nodes": [0, 1],
                    "am_arrival_gate": {
                        "status": "failed",
                        "verified_departure_label": "06:00",
                        "verified_arrival_label": "08:12",
                        "time_window_overrun_minutes": 12,
                    },
                }
            ],
        },
        service_direction="To School",
        source_label="15-Minute Constrained",
    )

    workbook = load_workbook(io.BytesIO(content), read_only=True)
    rows = list(workbook["current_plan_assignments"].iter_rows(values_only=True))
    assert "am window status" in rows[0]
    assert rows[1][0] == "Opt Bus 1"
    assert rows[1][-5:-1] == ("failed", "06:00", "08:12", 12)


def test_scenario_template_export_keeps_new_time_without_comparison(monkeypatch):
    def fake_map_payload(_job_record, _scenario_key, _map_key, *, attach_impact=False):
        return {
            "routes": [{"id": "Bus 1", "am_arrival_gate": {"status": "passed"}}],
            "stops": [{"node_index": 1, "scheduled_time_label": "07:10"}],
        }, None

    monkeypatch.setattr(backend_service, "_build_job_map_payload", fake_map_payload)

    content, error = backend_service._build_scenario_template_export(
        {
            "result": {
                "structured_results": {
                    "time_constrained": {
                        "points": [
                            {"node_id": 0, "address": "School", "passenger_count": 0},
                            {"node_id": 1, "address": "Stop A", "passenger_count": 3},
                        ],
                        "routes": [
                            {
                                "route_id": "Bus 1",
                                "bus_type_name": "18-fbus",
                                "bus_capacity": 18,
                                "nodes": [0, 1],
                            }
                        ],
                    }
                }
            }
        },
        "time_constrained",
    )

    assert error is None
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    rows = list(workbook["current_plan_assignments"].iter_rows(values_only=True))
    assert rows[2][8] == "07:10"
